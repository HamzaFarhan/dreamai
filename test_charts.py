#!/usr/bin/env python3
"""
Test script for OpenPyXL Charts functionality
"""

import os
import tempfile
from pathlib import Path

from openpyxl import Workbook

from excel_charts_toolset import (
    create_chart,
    create_data_table,
    create_matplotlib_chart,
    get_chart_types,
    list_charts,
)


def create_test_excel_file():
    """Create a test Excel file with sample data."""
    # Create temporary file
    temp_file = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    temp_path = temp_file.name
    temp_file.close()
    
    # Create workbook with sample data
    wb = Workbook()
    ws = wb.active
    ws.title = "TestData"
    
    # Add headers
    ws['A1'] = "Month"
    ws['B1'] = "Sales"
    ws['C1'] = "Profit"
    
    # Add sample data
    data = [
        ["Jan", 1000, 200],
        ["Feb", 1200, 250],
        ["Mar", 1100, 220],
        ["Apr", 1300, 280],
        ["May", 1400, 300],
        ["Jun", 1250, 270]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)
    
    wb.save(temp_path)
    return temp_path


def test_chart_creation():
    """Test creating various types of charts."""
    print("Testing chart creation...")
    
    # Create test file
    excel_path = create_test_excel_file()
    print(f"Created test file: {excel_path}")
    
    try:
        # Test creating a column chart
        result = create_chart(
            excel_path=excel_path,
            data_range="A1:B7",
            chart_type="column",
            title="Monthly Sales",
            chart_sheet="Charts",
            position="A1"
        )
        print(f"✓ Column chart created: {result}")
        
        # Test creating a line chart
        result = create_chart(
            excel_path=excel_path,
            data_range="A1:C7",
            chart_type="line",
            title="Sales and Profit Trend",
            chart_sheet="Charts",
            position="A15"
        )
        print(f"✓ Line chart created: {result}")
        
        # Test creating a pie chart
        result = create_chart(
            excel_path=excel_path,
            data_range="A1:B7",
            chart_type="pie",
            title="Sales Distribution",
            chart_sheet="Charts",
            position="J1"
        )
        print(f"✓ Pie chart created: {result}")
        
    except Exception as e:
        print(f"✗ Chart creation failed: {e}")
        return False
    
    finally:
        # Clean up
        if os.path.exists(excel_path):
            os.unlink(excel_path)
    
    return True


def test_data_table():
    """Test creating data tables."""
    print("\nTesting data table creation...")
    
    # Create test file
    excel_path = create_test_excel_file()
    
    try:
        # Test creating a data table
        result = create_data_table(
            excel_path=excel_path,
            sheet_name="TestData",
            data_range="A1:C7",
            table_name="SalesTable"
        )
        print(f"✓ Data table created: {result}")
        
    except Exception as e:
        print(f"✗ Data table creation failed: {e}")
        return False
    
    finally:
        # Clean up
        if os.path.exists(excel_path):
            os.unlink(excel_path)
    
    return True


def test_matplotlib_chart():
    """Test creating matplotlib charts."""
    print("\nTesting matplotlib chart creation...")
    
    # Create test file
    excel_path = create_test_excel_file()
    
    try:
        # Test creating a matplotlib bar chart
        result = create_matplotlib_chart(
            excel_path=excel_path,
            data_range="A1:B7",
            chart_type="bar",
            title="Matplotlib Bar Chart",
            chart_sheet="MatplotlibCharts",
            position="A1"
        )
        print(f"✓ Matplotlib chart created: {result}")
        
    except ImportError as e:
        print(f"⚠ Matplotlib not available: {e}")
        return True  # Not a failure, just not available
    except Exception as e:
        print(f"✗ Matplotlib chart creation failed: {e}")
        return False
    
    finally:
        # Clean up
        if os.path.exists(excel_path):
            os.unlink(excel_path)
    
    return True


def test_chart_listing():
    """Test listing charts in workbook."""
    print("\nTesting chart listing...")
    
    # Create test file
    excel_path = create_test_excel_file()
    
    try:
        # Add a chart first
        create_chart(
            excel_path=excel_path,
            data_range="A1:B7",
            chart_type="column",
            title="Test Chart",
            position="A10"
        )
        
        # List charts
        charts_info = list_charts(excel_path)
        print(f"✓ Found {charts_info['total_charts']} chart(s)")
        
        for chart in charts_info['charts']:
            print(f"  - {chart['type']}: {chart['title']} on sheet {chart['sheet']}")
        
    except Exception as e:
        print(f"✗ Chart listing failed: {e}")
        return False
    
    finally:
        # Clean up
        if os.path.exists(excel_path):
            os.unlink(excel_path)
    
    return True


def test_chart_types():
    """Test getting available chart types."""
    print("\nTesting chart types listing...")
    
    try:
        chart_types = get_chart_types()
        print(f"✓ Available OpenPyXL chart types: {', '.join(chart_types['openpyxl_charts'])}")
        
        if chart_types['matplotlib_available']:
            print(f"✓ Available Matplotlib chart types: {', '.join(chart_types['matplotlib_charts'])}")
        else:
            print("⚠ Matplotlib not available")
        
    except Exception as e:
        print(f"✗ Chart types listing failed: {e}")
        return False
    
    return True


def main():
    """Run all tests."""
    print("OpenPyXL Charts Test Suite")
    print("=" * 40)
    
    tests = [
        test_chart_types,
        test_chart_creation,
        test_data_table,
        test_matplotlib_chart,
        test_chart_listing,
    ]
    
    passed = 0
    total = len(tests)
    
    for test in tests:
        if test():
            passed += 1
    
    print(f"\n{'='*40}")
    print(f"Tests passed: {passed}/{total}")
    
    if passed == total:
        print("🎉 All tests passed!")
        return 0
    else:
        print("❌ Some tests failed.")
        return 1


if __name__ == "__main__":
    exit(main())
