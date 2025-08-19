import json
from io import BytesIO
from pathlib import Path
from typing import Any, Literal

import matplotlib.style as style
from openpyxl import load_workbook
from openpyxl.chart import (
    AreaChart,
    BarChart,
    BubbleChart,
    DoughnutChart,
    LineChart,
    PieChart,
    RadarChart,
    ScatterChart,
    StockChart,
    SurfaceChart,
)
from openpyxl.chart.bar_chart import BarChart as ColumnChart
from openpyxl.chart.reference import Reference
from openpyxl.drawing.image import Image
from openpyxl.utils import range_boundaries


# Custom Exceptions
class FileOperationError(Exception):
    """Raised when file operations fail."""

    pass


class SheetNotFoundError(Exception):
    """Raised when a specified sheet is not found in the workbook."""

    pass


class ChartError(Exception):
    """Raised when there's an error with chart parameters or creation."""

    pass


class DataError(Exception):
    """Raised when there's an error with chart data."""

    pass


# Type definitions
ChartType = Literal[
    "column", "bar", "line", "pie", "scatter", "area", "doughnut", "radar", "bubble", "stock", "surface"
]
MatplotlibChartType = Literal[
    "line", "bar", "scatter", "pie", "histogram", "box", "heatmap", "violin", "density", "subplots"
]


# Helper Functions
def _split_sheet_and_range(range_spec: str) -> tuple[str | None, str]:
    """
    Parse "Sheet1!A1:B2" or "A1:B2" -> (sheet_name or None, a1_range)

    Args:
        range_spec: Range specification in A1 notation, optionally with sheet name

    Returns:
        Tuple of (sheet_name, a1_range)
    """
    if "!" in range_spec:
        sheet, a1 = range_spec.split("!", 1)
        return sheet.strip().strip("'\""), a1.strip()
    return None, range_spec.strip()


def _validate_chart_type(chart_type: str) -> str:
    """
    Validate and normalize chart type.

    Args:
        chart_type: Chart type string

    Returns:
        Validated chart type

    Raises:
        ChartError: If chart type is invalid
    """
    valid_types = {
        "column",
        "bar",
        "line",
        "pie",
        "scatter",
        "area",
        "doughnut",
        "radar",
        "bubble",
        "stock",
        "surface",
    }

    normalized = chart_type.lower().strip()
    if normalized not in valid_types:
        valid_list = sorted(list(valid_types))
        raise ChartError(f"Invalid chart type '{chart_type}'. Valid types are: {valid_list}")

    return normalized


def _get_chart_class(chart_type: str) -> Any:
    """
    Get openpyxl chart class for given type.

    Args:
        chart_type: Type of chart

    Returns:
        Chart class from openpyxl

    Raises:
        ChartError: If chart type is not supported
    """
    chart_classes = {
        "column": ColumnChart,
        "bar": BarChart,
        "line": LineChart,
        "pie": PieChart,
        "scatter": ScatterChart,
        "area": AreaChart,
        "doughnut": DoughnutChart,
        "radar": RadarChart,
        "bubble": BubbleChart,
        "stock": StockChart,
        "surface": SurfaceChart,
    }

    if chart_type not in chart_classes:
        raise ChartError(f"Chart type '{chart_type}' not supported")

    return chart_classes[chart_type]


def _parse_data_range(data_range: str, source_sheet: Any) -> dict[str, Any]:
    """
    Parse data range and extract information.

    Args:
        data_range: Range in A1 notation
        source_sheet: Worksheet object

    Returns:
        Dictionary with parsed range information

    Raises:
        DataError: If range is invalid
    """
    try:
        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # Validate range exists in sheet
        if max_row > source_sheet.max_row or max_col > source_sheet.max_column:
            raise DataError(f"Range '{data_range}' exceeds sheet dimensions")

        return {
            "min_col": min_col,
            "min_row": min_row,
            "max_col": max_col,
            "max_row": max_row,
            "has_headers": True,  # Assume first row has headers
        }
    except ValueError as e:
        raise DataError(f"Invalid range specification '{data_range}': {e}")


# Main Chart Functions
def create_chart(
    excel_path: str,
    data_range: str,
    chart_type: ChartType,
    title: str = "Chart",
    chart_sheet: str | None = None,
    position: str = "A1",
    width: int = 15,
    height: int = 10,
    sheet_name: str | None = None,
) -> str:
    """
    Create a chart in an Excel file using openpyxl.

    Args:
        excel_path: Path to the Excel file
        data_range: Range containing chart data in A1 notation (e.g., "A1:B10" or "Sheet1!A1:B10")
        chart_type: Type of chart to create
        title: Chart title
        chart_sheet: Name of sheet to place chart (None = same as data sheet)
        position: Cell position for chart placement
        width: Chart width in characters
        height: Chart height in characters
        sheet_name: Override sheet name if not in data_range

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If specified sheet doesn't exist
        ChartError: If chart creation fails
        DataError: If data range is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path).resolve()
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Validate chart type
        validated_chart_type = _validate_chart_type(chart_type)

        # Parse sheet and range
        sheet_from_range, a1_range = _split_sheet_and_range(data_range)
        source_sheet_name = sheet_from_range or sheet_name

        # Load workbook
        wb = load_workbook(excel_file)

        # Determine source sheet
        if source_sheet_name is None:
            source_sheet_name = wb.sheetnames[0]
        elif source_sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{source_sheet_name}' not found in {excel_path}")

        source_ws = wb[source_sheet_name]

        # Parse and validate data range
        range_info = _parse_data_range(a1_range, source_ws)

        # Determine target sheet for chart
        if chart_sheet is None:
            chart_ws = source_ws
        else:
            if chart_sheet not in wb.sheetnames:
                chart_ws = wb.create_sheet(title=chart_sheet)
            else:
                chart_ws = wb[chart_sheet]

        # Create chart
        chart_class = _get_chart_class(validated_chart_type)
        chart = chart_class()
        chart.title = title
        chart.width = width
        chart.height = height

        # Create data reference
        data = Reference(
            source_ws,
            min_col=range_info["min_col"],
            min_row=range_info["min_row"],
            max_col=range_info["max_col"],
            max_row=range_info["max_row"],
        )

        # Add data to chart
        chart.add_data(data, titles_from_data=range_info["has_headers"])

        # Set categories if we have multiple columns
        if range_info["max_col"] > range_info["min_col"]:
            categories = Reference(
                source_ws,
                min_col=range_info["min_col"],
                min_row=range_info["min_row"] + 1,
                max_row=range_info["max_row"],
            )
            chart.set_categories(categories)

        # Add chart to worksheet
        chart_ws.add_chart(chart, position)

        # Save workbook
        wb.save(excel_file)
        return str(excel_file)

    except (FileNotFoundError, SheetNotFoundError, ChartError, DataError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create chart: {e}")


# Pivot table and data table functions have been moved to excel_structure_toolset.py
# to avoid duplication and maintain clear separation of concerns.


# Matplotlib Integration Functions
def create_matplotlib_chart(
    excel_path: str,
    data_range: str,
    chart_type: MatplotlibChartType,
    title: str = "Chart",
    sheet_name: str | None = None,
    chart_sheet: str | None = None,
    position: str = "A1",
    width: int = 600,
    height: int = 400,
    style_name: str = "default",
    save_image: bool = False,
) -> str:
    """
    Create a chart using matplotlib and embed it as an image in Excel.

    Args:
        excel_path: Path to the Excel file
        data_range: Range containing chart data
        chart_type: Type of matplotlib chart
        title: Chart title
        sheet_name: Source sheet name
        chart_sheet: Target sheet for chart
        position: Cell position for image
        width: Image width in pixels
        height: Image height in pixels
        style_name: Matplotlib style
        save_image: Whether to save image file separately

    Returns:
        Path to the Excel file

    Raises:
        ImportError: If matplotlib is not available
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ChartError: If chart creation fails
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path).resolve()
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Parse sheet and range
        sheet_from_range, a1_range = _split_sheet_and_range(data_range)
        source_sheet_name = sheet_from_range or sheet_name

        # Load workbook and get data
        wb = load_workbook(excel_file)

        if source_sheet_name is None:
            source_sheet_name = wb.sheetnames[0]
        elif source_sheet_name not in wb.sheetnames:
            available_sheets = wb.sheetnames
            raise SheetNotFoundError(
                f"Sheet '{source_sheet_name}' not found. Available sheets: {available_sheets}"
            )

        ws = wb[source_sheet_name]

        # Extract data from range
        min_col, min_row, max_col, max_row = range_boundaries(a1_range)

        data = []
        headers = []
        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            row_data = [cell.value for cell in row]
            if not headers:
                headers = row_data
            else:
                data.append(row_data)

        if not data:
            raise DataError("No data found in specified range")

        # Set matplotlib style
        if style and hasattr(style, "available") and style_name in style.available:
            style.use(style_name)
        elif style and hasattr(style, "available") and style_name not in style.available:
            available_styles = list(style.available)
            raise ChartError(f"Invalid matplotlib style '{style_name}'. Available styles are: {available_styles}")

        # Import matplotlib here to avoid scope issues
        import matplotlib.pyplot as plt_local

        fig, ax = plt_local.subplots(figsize=(width / 100, height / 100), dpi=100)

        if chart_type == "line":
            for i, header in enumerate(headers[1:], 1):
                y_data = [row[i] for row in data if row[i] is not None]
                x_data = list(range(len(y_data)))
                ax.plot(x_data, y_data, label=header, marker="o")
            ax.legend()

        elif chart_type == "bar":
            x_labels = [row[0] for row in data]
            y_data = [row[1] for row in data if row[1] is not None]
            ax.bar(x_labels, y_data)
            ax.set_xlabel(headers[0])
            ax.set_ylabel(headers[1])

        elif chart_type == "scatter":
            x_data = [row[0] for row in data if row[0] is not None]
            y_data = [row[1] for row in data if row[1] is not None]
            ax.scatter(x_data, y_data)
            ax.set_xlabel(headers[0])
            ax.set_ylabel(headers[1])

        elif chart_type == "pie":
            labels = [row[0] for row in data]
            sizes = [row[1] for row in data if row[1] is not None]
            ax.pie(sizes, labels=labels, autopct="%1.1f%%")

        elif chart_type == "histogram":
            data_values = [row[1] for row in data if row[1] is not None]
            ax.hist(data_values, bins=20, edgecolor="black")
            ax.set_xlabel(headers[1])
            ax.set_ylabel("Frequency")

        elif chart_type == "box":
            data_values = [[row[i] for row in data if row[i] is not None] for i in range(1, len(headers))]
            ax.boxplot(data_values, labels=headers[1:])

        elif chart_type == "heatmap":
            # Convert data to numpy array for heatmap
            numeric_data = []
            for row in data:
                numeric_row = []
                for val in row[1:]:
                    try:
                        numeric_row.append(float(val) if val is not None else 0)
                    except (ValueError, TypeError):
                        numeric_row.append(0)
                numeric_data.append(numeric_row)

            im = ax.imshow(numeric_data, cmap="viridis", aspect="auto")
            ax.set_xticks(range(len(headers[1:])))
            ax.set_xticklabels(headers[1:])
            ax.set_yticks(range(len(data)))
            ax.set_yticklabels([row[0] for row in data])
            plt_local.colorbar(im, ax=ax)

        else:
            valid_matplotlib_types = [
                "line",
                "bar",
                "scatter",
                "pie",
                "histogram",
                "box",
                "heatmap",
                "violin",
                "density",
                "subplots",
            ]
            raise ChartError(
                f"Unsupported matplotlib chart type: '{chart_type}'. Valid types are: {valid_matplotlib_types}"
            )

        ax.set_title(title)
        plt_local.tight_layout()

        # Save to memory
        img_buffer = BytesIO()
        plt_local.savefig(img_buffer, format="png", dpi=100, bbox_inches="tight")
        img_buffer.seek(0)

        # Save image file if requested
        if save_image:
            image_path = excel_file.parent / f"{excel_file.stem}_{title.replace(' ', '_')}.png"
            plt_local.savefig(image_path, format="png", dpi=100, bbox_inches="tight")

        plt_local.close()

        # Embed image in Excel
        target_sheet_name = chart_sheet or source_sheet_name
        if target_sheet_name not in wb.sheetnames:
            wb.create_sheet(title=target_sheet_name)

        target_ws = wb[target_sheet_name]

        # Create image object and add to worksheet
        img = Image(img_buffer)
        img.width = width
        img.height = height
        target_ws.add_image(img, position)

        wb.save(excel_file)
        return str(excel_file)

    except (ImportError, FileNotFoundError, SheetNotFoundError, ChartError, DataError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create matplotlib chart: {e}")


# Chart Management Functions
def list_charts(excel_path: str, sheet_name: str | None = None) -> dict[str, Any]:
    """
    List all charts in a workbook or specific sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Optional sheet name (None = all sheets)

    Returns:
        Dictionary containing chart information

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path).resolve()
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_file)

        charts_info = {"charts": [], "total_charts": 0}

        sheets_to_check = [sheet_name] if sheet_name else wb.sheetnames

        for sheet in sheets_to_check:
            if sheet not in wb.sheetnames:
                available_sheets = wb.sheetnames
                raise SheetNotFoundError(f"Sheet '{sheet}' not found. Available sheets: {available_sheets}")

            ws = wb[sheet]

            for chart in ws._charts:
                chart_info = {
                    "sheet": sheet,
                    "title": getattr(chart, "title", "Untitled"),
                    "type": type(chart).__name__,
                    "position": getattr(chart, "anchor", "Unknown"),
                }
                charts_info["charts"].append(chart_info)

        charts_info["total_charts"] = len(charts_info["charts"])
        return charts_info

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to list charts: {e}")


def delete_chart(excel_path: str, sheet_name: str, chart_index: int = 0) -> str:
    """
    Delete a chart from a worksheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet containing the chart
        chart_index: Index of the chart to delete (0-based)

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ChartError: If chart index is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path).resolve()
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_file)

        if sheet_name not in wb.sheetnames:
            available_sheets = wb.sheetnames
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")

        ws = wb[sheet_name]

        if chart_index >= len(ws._charts) or chart_index < 0:
            valid_indices = list(range(len(ws._charts))) if ws._charts else []
            raise ChartError(
                f"Chart index {chart_index} is invalid. Sheet has {len(ws._charts)} charts. Valid indices are: {valid_indices}"
            )

        # Remove chart
        del ws._charts[chart_index]

        wb.save(excel_file)
        return str(excel_file)

    except (FileNotFoundError, SheetNotFoundError, ChartError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to delete chart: {e}")


def update_chart_data(
    excel_path: str,
    sheet_name: str,
    chart_index: int,
    new_data_range: str,
) -> str:
    """
    Update the data range for an existing chart.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet containing the chart
        chart_index: Index of the chart to update
        new_data_range: New data range in A1 notation

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ChartError: If chart index is invalid
        DataError: If data range is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path).resolve()
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_file)

        if sheet_name not in wb.sheetnames:
            available_sheets = wb.sheetnames
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found. Available sheets: {available_sheets}")

        ws = wb[sheet_name]

        if chart_index >= len(ws._charts) or chart_index < 0:
            valid_indices = list(range(len(ws._charts))) if ws._charts else []
            raise ChartError(
                f"Chart index {chart_index} is invalid. Sheet has {len(ws._charts)} charts. Valid indices are: {valid_indices}"
            )

        chart = ws._charts[chart_index]

        # Parse new data range
        sheet_from_range, a1_range = _split_sheet_and_range(new_data_range)
        source_sheet_name = sheet_from_range or sheet_name

        if source_sheet_name not in wb.sheetnames:
            available_sheets = wb.sheetnames
            raise SheetNotFoundError(
                f"Source sheet '{source_sheet_name}' not found. Available sheets: {available_sheets}"
            )

        source_ws = wb[source_sheet_name]
        range_info = _parse_data_range(a1_range, source_ws)

        # Create new data reference
        new_data = Reference(
            source_ws,
            min_col=range_info["min_col"],
            min_row=range_info["min_row"],
            max_col=range_info["max_col"],
            max_row=range_info["max_row"],
        )

        # Update chart data (this is a simplified approach)
        # Full implementation would require more complex data series management
        chart.add_data(new_data, titles_from_data=range_info["has_headers"])

        wb.save(excel_file)
        return str(excel_file)

    except (FileNotFoundError, SheetNotFoundError, ChartError, DataError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to update chart data: {e}")


# Utility Functions
def get_chart_types() -> dict[str, Any]:
    """
    Get available chart types for both openpyxl and matplotlib.

    Returns:
        Dictionary with available chart types
    """
    return {
        "openpyxl_charts": [
            "column",
            "bar",
            "line",
            "pie",
            "scatter",
            "area",
            "doughnut",
            "radar",
            "bubble",
            "stock",
            "surface",
        ],
        "matplotlib_charts": [
            "line",
            "bar",
            "scatter",
            "pie",
            "histogram",
            "box",
            "heatmap",
            "violin",
            "density",
            "subplots",
        ],
    }


def create_chart_preset(name: str, chart_config: dict[str, Any]) -> str:
    """
    Create a chart configuration preset for reuse.

    Args:
        name: Name of the preset
        chart_config: Chart configuration dictionary

    Returns:
        Path to the presets file

    Raises:
        FileOperationError: If preset creation fails
    """
    try:
        # Store presets in project directory
        presets_file = Path.cwd() / ".dreamai_chart_presets.json"

        # Load existing presets
        presets = {}
        if presets_file.exists():
            try:
                with presets_file.open("r") as f:
                    presets = json.load(f)
            except (json.JSONDecodeError, IOError):
                presets = {}

        # Add new preset
        presets[name] = {"config": chart_config, "created": True}

        # Save presets
        with presets_file.open("w") as f:
            json.dump(presets, f, indent=2)

        return str(presets_file)

    except Exception as e:
        raise FileOperationError(f"Failed to create chart preset: {e}")


def apply_chart_preset(
    excel_path: str, data_range: str, preset_name: str, override_params: dict[str, Any] | None = None
) -> str:
    """
    Apply a chart preset with optional parameter overrides.

    Args:
        excel_path: Path to the Excel file
        data_range: Data range for the chart
        preset_name: Name of the preset to apply
        override_params: Parameters to override from preset

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If presets file or Excel file doesn't exist
        KeyError: If preset name not found
        FileOperationError: If operation fails
    """
    try:
        presets_file = Path.cwd() / ".dreamai_chart_presets.json"

        if not presets_file.exists():
            raise FileNotFoundError("No chart presets file found")

        with presets_file.open("r") as f:
            presets = json.load(f)

        if preset_name not in presets:
            available_presets = list(presets.keys())
            raise KeyError(f"Preset '{preset_name}' not found. Available presets are: {available_presets}")

        # Get preset configuration
        config = presets[preset_name]["config"].copy()

        # Apply overrides
        if override_params:
            config.update(override_params)

        # Determine chart creation function based on config
        if config.get("matplotlib", False):
            return create_matplotlib_chart(excel_path, data_range, **config)
        else:
            return create_chart(excel_path, data_range, **config)

    except (FileNotFoundError, KeyError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to apply chart preset: {e}")


if __name__ == "__main__":
    # Example usage
    pass
