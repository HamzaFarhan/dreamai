"""Test script for both fixed CSV functions."""

import csv
import tempfile
from pathlib import Path

from src.dreamai.toolsets.excel_structure_toolset import csv_to_excel_sheet, csvs_to_excel


def test_csvs_to_excel_fix():
    """Test the csvs_to_excel function with different delimiters."""

    # Create test CSV files with different delimiters
    csv_files = []
    test_data = [["Name", "Age", "Score"], ["Alice", "25", "95.5"], ["Bob", "30", "87"], ["Charlie", "22", "92.3"]]

    delimiters = [",", ";", "\t"]

    for i, delimiter in enumerate(delimiters):
        with tempfile.NamedTemporaryFile(mode="w", suffix=f"_test_{i}.csv", delete=False) as csv_file:
            writer = csv.writer(csv_file, delimiter=delimiter)
            writer.writerows(test_data)
            csv_files.append(csv_file.name)

    # Create Excel file from multiple CSVs
    excel_path = tempfile.mktemp(suffix=".xlsx")

    try:
        result = csvs_to_excel(csv_files, excel_path)
        print(f"✓ Success with multiple CSV files: {result}")

    except Exception as e:
        print(f"✗ Failed with multiple CSV files: {e}")

    finally:
        # Cleanup
        for csv_file in csv_files:
            Path(csv_file).unlink(missing_ok=True)
        Path(excel_path).unlink(missing_ok=True)


def create_sample_csv():
    """Create a sample CSV that matches the user's subscriptions.csv structure."""

    sample_data = [
        ["customer_id", "subscription_date", "plan", "price", "status"],
        ["CUST001", "2023-01-15", "Premium", "29.99", "Active"],
        ["CUST002", "2023-02-20", "Basic", "9.99", "Active"],
        ["CUST003", "2023-03-10", "Premium", "29.99", "Cancelled"],
        ["CUST004", "2023-04-05", "Standard", "19.99", "Active"],
        ["CUST005", "2023-05-12", "Premium", "29.99", "Active"],
    ]

    with tempfile.NamedTemporaryFile(mode="w", suffix="_subscriptions.csv", delete=False) as csv_file:
        writer = csv.writer(csv_file)
        writer.writerows(sample_data)
        return csv_file.name


def test_user_scenario():
    """Test the exact scenario from the user's error."""

    # Create sample data similar to user's subscriptions.csv
    csv_path = create_sample_csv()
    excel_path = tempfile.mktemp(suffix="_Customer_LTV_Analysis_2023.xlsx")

    try:
        result = csv_to_excel_sheet(csv_path=csv_path, excel_path=excel_path, sheet_name="subscriptions")
        print(f"✓ Success with user scenario: {result}")

        # Verify the file was created and has the correct sheet
        from openpyxl import load_workbook

        wb = load_workbook(excel_path)
        if "subscriptions" in wb.sheetnames:
            print("✓ Sheet 'subscriptions' was created successfully")
            ws = wb["subscriptions"]
            print(f"✓ Sheet has {ws.max_row} rows and {ws.max_column} columns")
        else:
            print("✗ Sheet 'subscriptions' was not found")

    except Exception as e:
        print(f"✗ Failed with user scenario: {e}")

    finally:
        # Cleanup
        Path(csv_path).unlink(missing_ok=True)
        Path(excel_path).unlink(missing_ok=True)


if __name__ == "__main__":
    print("Testing CSV to Excel functions with delimiter detection fixes...")
    print("=" * 70)

    print("1. Testing csvs_to_excel function:")
    test_csvs_to_excel_fix()
    print()

    print("2. Testing user scenario:")
    test_user_scenario()

    print("\nAll tests complete!")
