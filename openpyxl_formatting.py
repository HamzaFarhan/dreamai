"""
OpenPyXL Formatting Functions

This module provides Excel file formatting using OpenPyXL.
All functions take file_path as the first argument and return file paths.
Functions apply formatting to Excel files and save them.

Author: Generated for DreamAI formatting needs
"""

import json
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.formatting.rule import CellIsRule, FormulaRule
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import range_boundaries


# Custom Exceptions
class FileOperationError(Exception):
    """Raised when file operations fail."""

    pass


class SheetNotFoundError(Exception):
    """Raised when a specified sheet is not found in the workbook."""

    pass


class FormatError(Exception):
    """Raised when there's an error with formatting parameters."""

    pass


# Helper Functions
def _split_sheet_and_range(range_spec: str) -> tuple[Optional[str], str]:
    """
    Parse "Sheet1!A1:B2" or "A1:B2" -> (sheet_name or None, a1_range)

    Args:
        range_spec: Range specification in A1 notation, optionally with sheet name

    Returns:
        Tuple of (sheet_name, a1_range)
    """
    if "!" in range_spec:
        sheet, a1 = range_spec.split("!", 1)
        return sheet.strip().strip("'\""), a1.strip()
    return None, range_spec.strip()


def _rgb_from_google(color: Any) -> Optional[str]:
    """
    Convert Google-style color to hex string.

    Args:
        color: {'red':0..1,'green':0..1,'blue':0..1} or "#RRGGBB" string

    Returns:
        Hex color string without # prefix, or None if invalid
    """
    if not color:
        return None

    if isinstance(color, str):
        if color.startswith("#"):
            return color.lstrip("#").upper()
        return color.upper() if len(color) == 6 else None

    if isinstance(color, dict):
        try:
            r = int(color.get("red", 0) * 255)
            g = int(color.get("green", 0) * 255)
            b = int(color.get("blue", 0) * 255)
            return f"{r:02X}{g:02X}{b:02X}"
        except (ValueError, TypeError):
            return None

    return None


def _map_formatting_to_openpyxl(formatting: dict[str, Any]) -> dict[str, Any]:
    """
    Convert Google-style formatting dict to openpyxl-friendly format.

    Args:
        formatting: Google Sheets style formatting dict

    Returns:
        Dict with 'fill' and 'font' keys for openpyxl
    """
    result = {}

    # Background color
    bg_color = formatting.get("backgroundColor")
    if bg_color:
        hex_color = _rgb_from_google(bg_color)
        if hex_color:
            result["fill"] = hex_color

    # Text formatting
    text_format = formatting.get("textFormat", {})
    if text_format:
        font_attrs = {}

        if "bold" in text_format:
            font_attrs["bold"] = bool(text_format["bold"])

        if "italic" in text_format:
            font_attrs["italic"] = bool(text_format["italic"])

        if "fontSize" in text_format:
            try:
                font_attrs["size"] = int(text_format["fontSize"])
            except (ValueError, TypeError):
                pass

        if "foregroundColor" in text_format:
            color = _rgb_from_google(text_format["foregroundColor"])
            if color:
                font_attrs["color"] = color

        if font_attrs:
            result["font"] = font_attrs

    return result


def _extract_condition_value(condition: dict[str, Any]) -> Any:
    """
    Extract value from Google-style condition dict.

    Args:
        condition: Condition dictionary

    Returns:
        Extracted value for the condition
    """
    if not condition:
        return ""

    # Try different value keys
    for key in ["values", "value", "userEnteredValue"]:
        vals = condition.get(key)
        if vals is not None:
            if isinstance(vals, list) and vals:
                first = vals[0]
                if isinstance(first, dict):
                    return first.get("userEnteredValue", first.get("value", ""))
                return first
            elif isinstance(vals, dict):
                return vals.get("userEnteredValue", vals.get("value", ""))
            else:
                return vals

    return ""


def _extract_single_value(value: Any) -> Any:
    """
    Extract a single value from various formats.

    Args:
        value: Value to extract (could be dict with userEnteredValue, etc.)

    Returns:
        Extracted single value
    """
    if isinstance(value, dict):
        return value.get("userEnteredValue", value.get("value", ""))
    return value


# Main Formatting Functions
def apply_cell_formatting(
    excel_path: str, range_spec: str, formatting: dict[str, Any], sheet_name: Optional[str] = None
) -> str:
    """
    Apply formatting to a cell range in an Excel file.

    Args:
        excel_path: Path to the Excel file
        range_spec: Range in A1 notation (e.g., "A1:B5" or "Sheet1!A1:B5")
        formatting: Formatting options (backgroundColor, textFormat, etc.)
        sheet_name: Optional sheet name override

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If specified sheet doesn't exist
        FormatError: If formatting parameters are invalid
        FileOperationError: If operation fails
    """
    try:
        excel_path = Path(excel_path).resolve()
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Parse sheet and range
        sheet_from_range, a1_range = _split_sheet_and_range(range_spec)
        target_sheet = sheet_from_range or sheet_name

        # Load workbook
        wb = load_workbook(excel_path)

        # Determine target sheet
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]
        elif target_sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{target_sheet}' not found in {excel_path}")

        ws = wb[target_sheet]

        # Map formatting to openpyxl objects
        mapped_format = _map_formatting_to_openpyxl(formatting)

        fill = None
        if "fill" in mapped_format:
            fill = PatternFill(
                start_color=mapped_format["fill"], end_color=mapped_format["fill"], fill_type="solid"
            )

        font = None
        if "font" in mapped_format:
            font = Font(**mapped_format["font"])

        # Apply formatting to range
        try:
            min_col, min_row, max_col, max_row = range_boundaries(a1_range)
        except ValueError as e:
            raise FormatError(f"Invalid range specification '{a1_range}': {e}")

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                if fill:
                    cell.fill = fill
                if font:
                    cell.font = font

        # Save workbook
        wb.save(excel_path)
        return str(excel_path)

    except (FileNotFoundError, SheetNotFoundError, FormatError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to apply cell formatting: {e}")


def apply_conditional_formatting(
    excel_path: str,
    range_spec: str,
    condition: dict[str, Any],
    format_style: dict[str, Any],
    sheet_name: Optional[str] = None,
) -> str:
    """
    Apply conditional formatting to a cell range in an Excel file.

    Args:
        excel_path: Path to the Excel file
        range_spec: Range in A1 notation to apply conditional formatting
        condition: Condition for formatting (e.g., {'type': 'GREATER_THAN', 'value': 100})
        format_style: Format to apply when condition is met
        sheet_name: Optional sheet name override

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If specified sheet doesn't exist
        FormatError: If condition or format parameters are invalid
        FileOperationError: If operation fails
    """
    try:
        excel_path = Path(excel_path).resolve()
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Parse sheet and range
        sheet_from_range, a1_range = _split_sheet_and_range(range_spec)
        target_sheet = sheet_from_range or sheet_name

        # Load workbook
        wb = load_workbook(excel_path)

        # Determine target sheet
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]
        elif target_sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{target_sheet}' not found in {excel_path}")

        ws = wb[target_sheet]

        # Map condition to openpyxl rule
        condition_type = condition.get("type", "").upper()
        condition_value = _extract_condition_value(condition)

        # Create formatting style
        mapped_style = _map_formatting_to_openpyxl(format_style)

        fill = None
        if "fill" in mapped_style:
            fill = PatternFill(start_color=mapped_style["fill"], end_color=mapped_style["fill"], fill_type="solid")

        font = None
        if "font" in mapped_style:
            font = Font(**mapped_style["font"])

        # Create conditional formatting rule
        rule = None

        if condition_type == "GREATER_THAN":
            rule = CellIsRule(operator="greaterThan", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "LESS_THAN":
            rule = CellIsRule(operator="lessThan", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "GREATER_THAN_EQ":
            rule = CellIsRule(operator="greaterThanOrEqual", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "LESS_THAN_EQ":
            rule = CellIsRule(operator="lessThanOrEqual", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "EQUAL":
            rule = CellIsRule(operator="equal", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "NOT_EQUAL":
            rule = CellIsRule(operator="notEqual", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "TEXT_CONTAINS":
            rule = CellIsRule(operator="containsText", formula=[condition_value], fill=fill, font=font)
        elif condition_type == "BETWEEN":
            # Handle between condition with two values
            if isinstance(condition_value, (list, tuple)) and len(condition_value) == 2:
                rule = CellIsRule(
                    operator="between",
                    formula=[str(condition_value[0]), str(condition_value[1])],
                    fill=fill,
                    font=font,
                )
            else:
                # If single value provided, extract from condition dict
                values = condition.get("values", [])
                if not values:
                    # Try "value" key as well
                    values = condition.get("value", [])

                if isinstance(values, (list, tuple)) and len(values) >= 2:
                    val1 = _extract_single_value(values[0])
                    val2 = _extract_single_value(values[1])
                    rule = CellIsRule(operator="between", formula=[str(val1), str(val2)], fill=fill, font=font)
                else:
                    raise FormatError(f"BETWEEN condition requires two values, got: {values}")
        elif condition_type == "CUSTOM_FORMULA":
            rule = FormulaRule(formula=[condition_value], fill=fill, font=font)
        else:
            # Default to formula rule
            rule = FormulaRule(formula=[str(condition_value)], fill=fill, font=font)

        if rule:
            ws.conditional_formatting.add(a1_range, rule)

        # Save workbook
        wb.save(excel_path)
        return str(excel_path)

    except (FileNotFoundError, SheetNotFoundError, FormatError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to apply conditional formatting: {e}")


def create_formatting_preset(name: str, formatting: dict[str, Any]) -> str:
    """
    Create a formatting preset for reuse.

    Args:
        name: Name of the preset
        formatting: Formatting configuration

    Returns:
        Path to the presets file

    Raises:
        FileOperationError: If preset creation fails
    """
    try:
        # Store presets in project directory
        presets_file = Path.cwd() / ".dreamai_format_presets.json"

        # Load existing presets
        presets = {}
        if presets_file.exists():
            try:
                with presets_file.open("r") as f:
                    presets = json.load(f)
            except (json.JSONDecodeError, IOError):
                presets = {}

        # Add new preset
        presets[name] = {"formatting": formatting, "created": True}

        # Save presets
        with presets_file.open("w") as f:
            json.dump(presets, f, indent=2)

        return str(presets_file)

    except Exception as e:
        raise FileOperationError(f"Failed to create formatting preset: {e}")


def load_formatting_preset(name: str) -> dict[str, Any]:
    """
    Load a formatting preset by name.

    Args:
        name: Name of the preset to load

    Returns:
        Formatting configuration dictionary

    Raises:
        FileNotFoundError: If presets file doesn't exist
        KeyError: If preset name not found
        FileOperationError: If loading fails
    """
    try:
        presets_file = Path.cwd() / ".dreamai_format_presets.json"

        if not presets_file.exists():
            raise FileNotFoundError("No formatting presets file found")

        with presets_file.open("r") as f:
            presets = json.load(f)

        if name not in presets:
            raise KeyError(f"Preset '{name}' not found")

        return presets[name]["formatting"]

    except (FileNotFoundError, KeyError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to load formatting preset: {e}")


def list_formatting_presets() -> dict[str, dict[str, Any]]:
    """
    List all available formatting presets.

    Returns:
        Dictionary of all presets

    Raises:
        FileOperationError: If listing fails
    """
    try:
        presets_file = Path.cwd() / ".dreamai_format_presets.json"

        if not presets_file.exists():
            return {}

        with presets_file.open("r") as f:
            return json.load(f)

    except Exception as e:
        raise FileOperationError(f"Failed to list formatting presets: {e}")


def apply_preset_formatting(
    excel_path: str, range_spec: str, preset_name: str, sheet_name: Optional[str] = None
) -> str:
    """
    Apply a formatting preset to a cell range.

    Args:
        excel_path: Path to the Excel file
        range_spec: Range in A1 notation
        preset_name: Name of the formatting preset
        sheet_name: Optional sheet name override

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file or preset doesn't exist
        KeyError: If preset name not found
        FileOperationError: If operation fails
    """
    try:
        formatting = load_formatting_preset(preset_name)
        return apply_cell_formatting(excel_path, range_spec, formatting, sheet_name)

    except (FileNotFoundError, KeyError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to apply preset formatting: {e}")


# Utility Functions
def clear_formatting(excel_path: str, range_spec: str, sheet_name: Optional[str] = None) -> str:
    """
    Clear all formatting from a cell range.

    Args:
        excel_path: Path to the Excel file
        range_spec: Range in A1 notation
        sheet_name: Optional sheet name override

    Returns:
        Path to the Excel file

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If specified sheet doesn't exist
        FileOperationError: If operation fails
    """
    try:
        excel_path = Path(excel_path).resolve()
        if not excel_path.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Parse sheet and range
        sheet_from_range, a1_range = _split_sheet_and_range(range_spec)
        target_sheet = sheet_from_range or sheet_name

        # Load workbook
        wb = load_workbook(excel_path)

        # Determine target sheet
        if target_sheet is None:
            target_sheet = wb.sheetnames[0]
        elif target_sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{target_sheet}' not found in {excel_path}")

        ws = wb[target_sheet]

        # Clear formatting from range
        try:
            min_col, min_row, max_col, max_row = range_boundaries(a1_range)
        except ValueError as e:
            raise FormatError(f"Invalid range specification '{a1_range}': {e}")

        # Create default styles
        default_font = Font()
        default_fill = PatternFill()

        for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.font = default_font
                cell.fill = default_fill

        # Save workbook
        wb.save(excel_path)
        return str(excel_path)

    except (FileNotFoundError, SheetNotFoundError, FormatError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to clear formatting: {e}")


if __name__ == "__main__":
    # Example usage
    pass
