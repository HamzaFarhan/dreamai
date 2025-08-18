#!/usr/bin/env python3
"""
Example usage of OpenPyXL Charts functionality
"""

from openpyxl import Workbook
from excel_charts_toolset import create_chart, create_data_table, get_chart_types


def create_sample_excel_with_charts():
    """Create a sample Excel file with charts."""
    # Create workbook with sample data
    wb = Workbook()
    ws = wb.active
    ws.title = "SalesData"
    
    # Add headers and data
    ws['A1'] = "Product"
    ws['B1'] = "Q1"
    ws['C1'] = "Q2"
    ws['D1'] = "Q3"
    ws['E1'] = "Q4"
    
    data = [
        ["Laptops", 120, 135, 142, 158],
        ["Phones", 95, 110, 125, 140],
        ["Tablets", 80, 85, 90, 95],
        ["Accessories", 200, 220, 210, 250]
    ]
    
    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)
    
    # Save initial file
    excel_path = "sample_charts.xlsx"
    wb.save(excel_path)
    
    print(f"Created Excel file: {excel_path}")
    
    # Create data table
    create_data_table(
        excel_path=excel_path,
        sheet_name="SalesData",
        data_range="A1:E5",
        table_name="SalesTable"
    )
    print("✓ Added data table")
    
    # Create column chart for Q1 sales
    create_chart(
        excel_path=excel_path,
        data_range="A1:B5",
        chart_type="column",
        title="Q1 Sales by Product",
        chart_sheet="Charts",
        position="A1"
    )
    print("✓ Added Q1 column chart")
    
    # Create line chart for all quarters
    create_chart(
        excel_path=excel_path,
        data_range="A1:E5",
        chart_type="line",
        title="Quarterly Sales Trend",
        chart_sheet="Charts",
        position="A20"
    )
    print("✓ Added quarterly trend line chart")
    
    # Create pie chart for Q4 sales (using just Q4 data)
    create_chart(
        excel_path=excel_path,
        data_range="A1:E5",  # Use full range, pie will use first two columns
        chart_type="pie",
        title="Q4 Sales Distribution",
        chart_sheet="Charts",
        position="J1"
    )
    print("✓ Added Q4 pie chart")
    
    print(f"\n🎉 Sample Excel file with charts created: {excel_path}")
    print("Open the file to see the charts!")
    
    return excel_path


def main():
    """Main function."""
    print("OpenPyXL Charts Example")
    print("=" * 30)
    
    # Show available chart types
    chart_types = get_chart_types()
    print(f"Available chart types: {', '.join(chart_types['openpyxl_charts'])}")
    print()
    
    # Create sample file
    try:
        excel_path = create_sample_excel_with_charts()
        return 0
    except Exception as e:
        print(f"❌ Error: {e}")
        return 1


if __name__ == "__main__":
    exit(main())
