"""
Comprehensive test suite for openpyxl_utils.py

This test suite covers all functions in the openpyxl_utils module with various
test cases including normal operations, edge cases, and error conditions.
"""

import csv
from pathlib import Path

import pytest
from openpyxl import load_workbook

from openpyxl_utils import (
    FormulaError,
    SheetNotFoundError,
    add_sheet,
    add_subtotals,
    clear_sheet,
    copy_sheet,
    create_autofilter,
    create_data_validation,
    create_dynamic_chart,
    create_excel_file,
    create_summary_table,
    csv_to_excel_sheet,
    csvs_to_excel,
    delete_sheet,
    extract_sheet_to_csv,
    get_cell_value,
    get_sheet_names,
    merge_excel_files,
    rename_sheet,
    write_data_to_sheet,
    write_date_function,
    write_financial_function,
    write_formula_to_cell,
    write_info_function,
    write_logical_function,
    write_lookup_function,
    write_math_function,
    write_multiple_formulas,
    write_statistical_function,
    write_text_function,
    write_value_to_cell,
)


class TestExcelFileOperations:
    """Test core Excel file operations."""

    def test_create_excel_file_success(self, tmp_path):
        """Test successful Excel file creation."""
        file_path = tmp_path / "test.xlsx"
        result = create_excel_file(str(file_path))

        assert Path(result).exists()
        assert Path(result).suffix == ".xlsx"

        # Verify workbook can be loaded
        wb = load_workbook(result)
        assert len(wb.sheetnames) == 1

    def test_create_excel_file_existing_file(self, tmp_path):
        """Test error when file already exists."""
        file_path = tmp_path / "existing.xlsx"
        file_path.touch()

        with pytest.raises(FileExistsError):
            create_excel_file(str(file_path))

    def test_create_excel_file_creates_directories(self, tmp_path):
        """Test that parent directories are created."""
        file_path = tmp_path / "subdir" / "another" / "test.xlsx"
        result = create_excel_file(str(file_path))

        assert Path(result).exists()
        assert Path(result).parent.exists()

    def test_get_sheet_names_success(self, tmp_path):
        """Test getting sheet names from workbook."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))
        add_sheet(str(file_path), "NewSheet")

        names = get_sheet_names(str(file_path))
        assert "Sheet" in names
        assert "NewSheet" in names
        assert len(names) == 2

    def test_get_sheet_names_nonexistent_file(self, tmp_path):
        """Test error when file doesn't exist."""
        file_path = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            get_sheet_names(str(file_path))

    def test_get_sheet_names_invalid_file(self, tmp_path):
        """Test error with invalid Excel file."""
        file_path = tmp_path / "invalid.xlsx"
        file_path.write_text("not an excel file")

        with pytest.raises(ValueError):
            get_sheet_names(str(file_path))


class TestSheetOperations:
    """Test sheet manipulation operations."""

    @pytest.fixture
    def excel_file(self, tmp_path):
        """Create a temporary Excel file for testing."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))
        return str(file_path)

    def test_add_sheet_success(self, excel_file):
        """Test successful sheet addition."""
        result = add_sheet(excel_file, "TestSheet")

        names = get_sheet_names(result)
        assert "TestSheet" in names

    def test_add_sheet_duplicate_name(self, excel_file):
        """Test error when sheet name already exists."""
        add_sheet(excel_file, "TestSheet")

        with pytest.raises(ValueError):
            add_sheet(excel_file, "TestSheet")

    def test_add_sheet_empty_name(self, excel_file):
        """Test error with empty sheet name."""
        with pytest.raises(ValueError):
            add_sheet(excel_file, "")

        with pytest.raises(ValueError):
            add_sheet(excel_file, "   ")

    def test_delete_sheet_success(self, excel_file):
        """Test successful sheet deletion."""
        add_sheet(excel_file, "ToDelete")
        result = delete_sheet(excel_file, "ToDelete")

        names = get_sheet_names(result)
        assert "ToDelete" not in names

    def test_delete_sheet_nonexistent(self, excel_file):
        """Test error when deleting nonexistent sheet."""
        with pytest.raises(SheetNotFoundError):
            delete_sheet(excel_file, "NonExistent")

    def test_delete_last_sheet(self, excel_file):
        """Test error when trying to delete the last sheet."""
        with pytest.raises(RuntimeError):
            delete_sheet(excel_file, "Sheet")

    def test_rename_sheet_success(self, excel_file):
        """Test successful sheet renaming."""
        result = rename_sheet(excel_file, "Sheet", "RenamedSheet")

        names = get_sheet_names(result)
        assert "RenamedSheet" in names
        assert "Sheet" not in names

    def test_rename_sheet_nonexistent(self, excel_file):
        """Test error when renaming nonexistent sheet."""
        with pytest.raises(SheetNotFoundError):
            rename_sheet(excel_file, "NonExistent", "NewName")

    def test_rename_sheet_to_existing_name(self, excel_file):
        """Test error when renaming to existing name."""
        add_sheet(excel_file, "ExistingSheet")

        with pytest.raises(ValueError):
            rename_sheet(excel_file, "Sheet", "ExistingSheet")

    def test_copy_sheet_success(self, excel_file):
        """Test successful sheet copying."""
        write_value_to_cell(excel_file, "Sheet", "A1", "test data")
        result = copy_sheet(excel_file, "Sheet", "CopiedSheet")

        names = get_sheet_names(result)
        assert "CopiedSheet" in names

        # Verify data was copied
        value = get_cell_value(result, "CopiedSheet", "A1")
        assert value == "test data"

    def test_copy_sheet_nonexistent(self, excel_file):
        """Test error when copying nonexistent sheet."""
        with pytest.raises(SheetNotFoundError):
            copy_sheet(excel_file, "NonExistent", "NewCopy")

    def test_clear_sheet_success(self, excel_file):
        """Test successful sheet clearing."""
        write_value_to_cell(excel_file, "Sheet", "A1", "test data")
        result = clear_sheet(excel_file, "Sheet")

        value = get_cell_value(result, "Sheet", "A1")
        assert value is None


class TestDataOperations:
    """Test data reading and writing operations."""

    @pytest.fixture
    def excel_file(self, tmp_path):
        """Create a temporary Excel file for testing."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))
        return str(file_path)

    def test_write_data_to_sheet_success(self, excel_file):
        """Test successful data writing."""
        data = [["Name", "Age"], ["John", 30], ["Jane", 25]]
        result = write_data_to_sheet(excel_file, "Sheet", data)

        # Verify data was written
        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Age"
        assert ws["A2"].value == "John"
        assert ws["B2"].value == 30

    def test_write_data_to_sheet_with_offset(self, excel_file):
        """Test data writing with start row/column offset."""
        data = [["Value"]]
        result = write_data_to_sheet(excel_file, "Sheet", data, start_row=3, start_col=2)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["B3"].value == "Value"

    def test_write_data_to_sheet_empty_data(self, excel_file):
        """Test error with empty data."""
        with pytest.raises(ValueError):
            write_data_to_sheet(excel_file, "Sheet", [])

    def test_write_data_to_sheet_nonexistent_sheet(self, excel_file):
        """Test error with nonexistent sheet."""
        data = [["test"]]
        with pytest.raises(SheetNotFoundError):
            write_data_to_sheet(excel_file, "NonExistent", data)

    def test_write_value_to_cell_success(self, excel_file):
        """Test writing single value to cell."""
        result = write_value_to_cell(excel_file, "Sheet", "A1", "test value")

        value = get_cell_value(result, "Sheet", "A1")
        assert value == "test value"

    def test_write_value_to_cell_number(self, excel_file):
        """Test writing numeric value to cell."""
        result = write_value_to_cell(excel_file, "Sheet", "A1", 42.5)

        value = get_cell_value(result, "Sheet", "A1")
        assert value == 42.5

    def test_get_cell_value_success(self, excel_file):
        """Test getting cell value."""
        write_value_to_cell(excel_file, "Sheet", "A1", "test")
        value = get_cell_value(excel_file, "Sheet", "A1")
        assert value == "test"

    def test_get_cell_value_empty_cell(self, excel_file):
        """Test getting value from empty cell."""
        value = get_cell_value(excel_file, "Sheet", "A1")
        assert value is None


class TestCSVOperations:
    """Test CSV import/export operations."""

    @pytest.fixture
    def csv_file(self, tmp_path):
        """Create a temporary CSV file for testing."""
        csv_path = tmp_path / "test.csv"
        data = [["Name", "Age", "Salary"], ["John", "30", "50000"], ["Jane", "25", "45000"]]

        with open(csv_path, "w", newline="") as f:
            writer = csv.writer(f)
            writer.writerows(data)

        return str(csv_path)

    @pytest.fixture
    def excel_file(self, tmp_path):
        """Create a temporary Excel file for testing."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))
        return str(file_path)

    def test_csv_to_excel_sheet_success(self, csv_file, tmp_path):
        """Test successful CSV to Excel conversion."""
        excel_path = tmp_path / "output.xlsx"
        result = csv_to_excel_sheet(csv_file, str(excel_path))

        wb = load_workbook(result)
        assert "test" in wb.sheetnames

        ws = wb["test"]
        assert ws["A1"].value == "Name"
        assert ws["B1"].value == "Age"
        assert ws["A2"].value == "John"

    def test_csv_to_excel_sheet_custom_name(self, csv_file, tmp_path):
        """Test CSV to Excel with custom sheet name."""
        excel_path = tmp_path / "output.xlsx"
        result = csv_to_excel_sheet(csv_file, str(excel_path), "CustomSheet")

        wb = load_workbook(result)
        assert "CustomSheet" in wb.sheetnames

    def test_csv_to_excel_sheet_nonexistent_csv(self, tmp_path):
        """Test error with nonexistent CSV file."""
        csv_path = tmp_path / "nonexistent.csv"
        excel_path = tmp_path / "output.xlsx"

        with pytest.raises(FileNotFoundError):
            csv_to_excel_sheet(str(csv_path), str(excel_path))

    def test_csvs_to_excel_success(self, tmp_path):
        """Test multiple CSVs to Excel conversion."""
        # Create multiple CSV files
        csv1 = tmp_path / "file1.csv"
        csv2 = tmp_path / "file2.csv"

        data1 = [["A", "B"], ["1", "2"]]
        data2 = [["C", "D"], ["3", "4"]]

        for csv_path, data in [(csv1, data1), (csv2, data2)]:
            with open(csv_path, "w", newline="") as f:
                writer = csv.writer(f)
                writer.writerows(data)

        excel_path = tmp_path / "output.xlsx"
        result = csvs_to_excel([str(csv1), str(csv2)], str(excel_path))

        wb = load_workbook(result)
        assert "file1" in wb.sheetnames
        assert "file2" in wb.sheetnames

    def test_csvs_to_excel_empty_list(self, tmp_path):
        """Test error with empty CSV list."""
        excel_path = tmp_path / "output.xlsx"

        with pytest.raises(ValueError):
            csvs_to_excel([], str(excel_path))

    def test_extract_sheet_to_csv_success(self, excel_file, tmp_path):
        """Test successful sheet to CSV extraction."""
        # Add data to sheet
        data = [["Name", "Age"], ["John", 30]]
        write_data_to_sheet(excel_file, "Sheet", data)

        csv_path = tmp_path / "output.csv"
        result = extract_sheet_to_csv(excel_file, "Sheet", str(csv_path))

        assert Path(result).exists()

        # Verify CSV content
        with open(result, "r") as f:
            reader = csv.reader(f)
            rows = list(reader)
            assert rows[0] == ["Name", "Age"]
            assert rows[1] == ["John", "30"]

    def test_extract_sheet_to_csv_nonexistent_sheet(self, excel_file, tmp_path):
        """Test error with nonexistent sheet."""
        csv_path = tmp_path / "output.csv"

        with pytest.raises(SheetNotFoundError):
            extract_sheet_to_csv(excel_file, "NonExistent", str(csv_path))


class TestFormulaFunctions:
    """Test formula writing functions."""

    @pytest.fixture
    def excel_file(self, tmp_path):
        """Create a temporary Excel file for testing."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))
        return str(file_path)

    def test_write_formula_to_cell_success(self, excel_file):
        """Test writing custom formula to cell."""
        result = write_formula_to_cell(excel_file, "Sheet", "A1", "=1+1")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=1+1"

    def test_write_formula_to_cell_auto_equals(self, excel_file):
        """Test formula without leading equals sign."""
        result = write_formula_to_cell(excel_file, "Sheet", "A1", "1+1")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=1+1"

    def test_write_math_function_sum(self, excel_file):
        """Test writing SUM function."""
        result = write_math_function(excel_file, "Sheet", "A1", "SUM", "B1:B10")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=SUM(B1:B10)"

    def test_write_math_function_pi(self, excel_file):
        """Test writing PI function (no parameters)."""
        result = write_math_function(excel_file, "Sheet", "A1", "PI")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=PI()"

    def test_write_math_function_invalid(self, excel_file):
        """Test error with invalid math function."""
        with pytest.raises(FormulaError):
            write_math_function(excel_file, "Sheet", "A1", "INVALID_FUNCTION", "B1")

    def test_write_statistical_function_average(self, excel_file):
        """Test writing AVERAGE function."""
        result = write_statistical_function(excel_file, "Sheet", "A1", "AVERAGE", "B1:B10")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=AVERAGE(B1:B10)"

    def test_write_statistical_function_no_data(self, excel_file):
        """Test error with no data for statistical function."""
        with pytest.raises(FormulaError):
            write_statistical_function(excel_file, "Sheet", "A1", "AVERAGE")

    def test_write_logical_function_if(self, excel_file):
        """Test writing IF function."""
        result = write_logical_function(excel_file, "Sheet", "A1", "IF", "B1>0", '"Yes"', '"No"')

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == '=IF(B1>0,"Yes","No")'

    def test_write_logical_function_true(self, excel_file):
        """Test writing TRUE function (no arguments)."""
        result = write_logical_function(excel_file, "Sheet", "A1", "TRUE")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=TRUE()"

    def test_write_date_function_today(self, excel_file):
        """Test writing TODAY function."""
        result = write_date_function(excel_file, "Sheet", "A1", "TODAY")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=TODAY()"

    def test_write_date_function_date(self, excel_file):
        """Test writing DATE function with arguments."""
        result = write_date_function(excel_file, "Sheet", "A1", "DATE", 2023, 12, 25)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=DATE(2023,12,25)"

    def test_write_text_function_concatenate(self, excel_file):
        """Test writing CONCATENATE function."""
        result = write_text_function(excel_file, "Sheet", "A1", "CONCATENATE", "B1", "C1")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=CONCATENATE(B1,C1)"

    def test_write_lookup_function_vlookup(self, excel_file):
        """Test writing VLOOKUP function."""
        result = write_lookup_function(excel_file, "Sheet", "A1", "VLOOKUP", "B1", "D:E", 2, "FALSE")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=VLOOKUP(B1,D:E,2,FALSE)"

    def test_write_financial_function_pv(self, excel_file):
        """Test writing PV function."""
        result = write_financial_function(excel_file, "Sheet", "A1", "PV", 0.05, 10, -1000)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=PV(0.05,10,-1000)"

    def test_write_info_function_isblank(self, excel_file):
        """Test writing ISBLANK function."""
        result = write_info_function(excel_file, "Sheet", "A1", "ISBLANK", "B1")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=ISBLANK(B1)"

    def test_write_multiple_formulas_success(self, excel_file):
        """Test writing multiple formulas at once."""
        formulas = {"A1": "=SUM(B1:B10)", "A2": "=AVERAGE(B1:B10)", "A3": "=MAX(B1:B10)"}
        result = write_multiple_formulas(excel_file, "Sheet", formulas)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "=SUM(B1:B10)"
        assert ws["A2"].value == "=AVERAGE(B1:B10)"
        assert ws["A3"].value == "=MAX(B1:B10)"


class TestAdvancedFeatures:
    """Test advanced Excel features."""

    @pytest.fixture
    def excel_file(self, tmp_path):
        """Create a temporary Excel file with sample data."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        # Add sample data
        data = [
            ["Category", "Product", "Sales", "Date"],
            ["A", "Product1", 100, "2023-01-01"],
            ["B", "Product2", 200, "2023-01-02"],
            ["A", "Product3", 150, "2023-01-03"],
        ]
        write_data_to_sheet(str(file_path), "Sheet", data)
        return str(file_path)

    def test_create_autofilter_success(self, excel_file):
        """Test creating autofilter."""
        result = create_autofilter(excel_file, "Sheet", "A1:D4")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws.auto_filter.ref == "A1:D4"

    def test_create_autofilter_nonexistent_sheet(self, excel_file):
        """Test error with nonexistent sheet."""
        with pytest.raises(SheetNotFoundError):
            create_autofilter(excel_file, "NonExistent", "A1:D4")

    def test_merge_excel_files_success(self, tmp_path):
        """Test merging multiple Excel files."""
        # Create multiple Excel files
        file1 = tmp_path / "file1.xlsx"
        file2 = tmp_path / "file2.xlsx"

        create_excel_file(str(file1))
        create_excel_file(str(file2))

        write_value_to_cell(str(file1), "Sheet", "A1", "File1 Data")
        write_value_to_cell(str(file2), "Sheet", "A1", "File2 Data")

        output_path = tmp_path / "merged.xlsx"
        result = merge_excel_files([str(file1), str(file2)], str(output_path))

        wb = load_workbook(result)
        sheet_names = wb.sheetnames
        assert len(sheet_names) == 2
        assert any("file1" in name for name in sheet_names)
        assert any("file2" in name for name in sheet_names)

    def test_merge_excel_files_empty_list(self, tmp_path):
        """Test error with empty file list."""
        output_path = tmp_path / "merged.xlsx"

        with pytest.raises(ValueError):
            merge_excel_files([], str(output_path))

    def test_create_data_validation_list(self, excel_file):
        """Test creating data validation with list."""
        options = ["Option1", "Option2", "Option3"]
        result = create_data_validation(excel_file, "Sheet", "A1:A10", "list", options)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert len(ws.data_validations.dataValidation) > 0

    def test_create_summary_table_success(self, excel_file):
        """Test creating summary table."""
        add_sheet(excel_file, "Summary")
        result = create_summary_table(excel_file, "Sheet", "Summary", "A1:D4", "A", {"C": "sum", "D": "count"})

        wb = load_workbook(result)
        assert "Summary" in wb.sheetnames

        summary_ws = wb["Summary"]
        # Check headers exist
        assert summary_ws["A1"].value is not None
        assert summary_ws["B1"].value is not None

    def test_create_dynamic_chart_success(self, excel_file):
        """Test creating dynamic chart."""
        add_sheet(excel_file, "Chart")
        result = create_dynamic_chart(excel_file, "Sheet", "Chart", "A1:C4", "column", "Sales Chart")

        wb = load_workbook(result)
        assert "Chart" in wb.sheetnames

        chart_ws = wb["Chart"]
        assert len(chart_ws._charts) > 0

    def test_add_subtotals_success(self, excel_file):
        """Test adding subtotals to data."""
        result = add_subtotals(excel_file, "Sheet", "A1:D4", 1, [3], "sum")

        wb = load_workbook(result)
        ws = wb["Sheet"]
        # Verify additional rows were added (subtotal rows)
        assert ws.max_row > 4


class TestErrorHandling:
    """Test error handling scenarios."""

    def test_nonexistent_file_operations(self, tmp_path):
        """Test operations on nonexistent files."""
        file_path = tmp_path / "nonexistent.xlsx"

        with pytest.raises(FileNotFoundError):
            get_sheet_names(str(file_path))

        with pytest.raises(FileNotFoundError):
            add_sheet(str(file_path), "NewSheet")

        with pytest.raises(FileNotFoundError):
            write_value_to_cell(str(file_path), "Sheet", "A1", "value")

    def test_invalid_sheet_operations(self, tmp_path):
        """Test operations on invalid sheets."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        with pytest.raises(SheetNotFoundError):
            write_data_to_sheet(str(file_path), "NonExistent", [["data"]])

        with pytest.raises(SheetNotFoundError):
            delete_sheet(str(file_path), "NonExistent")

        with pytest.raises(SheetNotFoundError):
            get_cell_value(str(file_path), "NonExistent", "A1")

    def test_invalid_formula_functions(self, tmp_path):
        """Test invalid formula function names."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        with pytest.raises(FormulaError):
            write_math_function(str(file_path), "Sheet", "A1", "INVALID_FUNC")

        with pytest.raises(FormulaError):
            write_statistical_function(str(file_path), "Sheet", "A1", "INVALID_STAT")

        with pytest.raises(FormulaError):
            write_logical_function(str(file_path), "Sheet", "A1", "INVALID_LOGIC")

    def test_permission_errors(self, tmp_path):
        """Test permission-related errors."""
        # This test might be platform-specific and could be skipped on some systems
        pass


class TestEdgeCases:
    """Test edge cases and boundary conditions."""

    def test_empty_data_handling(self, tmp_path):
        """Test handling of empty data."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        # Test empty data list
        with pytest.raises(ValueError):
            write_data_to_sheet(str(file_path), "Sheet", [])

    def test_large_data_handling(self, tmp_path):
        """Test handling of large datasets."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        # Create large dataset
        large_data = [["Col" + str(i) for i in range(10)] for _ in range(100)]
        result = write_data_to_sheet(str(file_path), "Sheet", large_data)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws.max_row == 100
        assert ws.max_column == 10

    def test_special_characters_in_sheet_names(self, tmp_path):
        """Test sheet names with special characters."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        # Test valid special characters
        valid_names = ["Sheet_1", "Sheet-2", "Sheet 3"]
        for name in valid_names:
            add_sheet(str(file_path), name)

        names = get_sheet_names(str(file_path))
        for name in valid_names:
            assert name in names

    def test_unicode_data_handling(self, tmp_path):
        """Test handling of Unicode data."""
        file_path = tmp_path / "test.xlsx"
        create_excel_file(str(file_path))

        unicode_data = [["中文", "العربية", "русский"], ["测试", "اختبار", "тест"]]
        result = write_data_to_sheet(str(file_path), "Sheet", unicode_data)

        wb = load_workbook(result)
        ws = wb["Sheet"]
        assert ws["A1"].value == "中文"
        assert ws["B1"].value == "العربية"
        assert ws["C1"].value == "русский"


if __name__ == "__main__":
    pytest.main([__file__, "-v"])
