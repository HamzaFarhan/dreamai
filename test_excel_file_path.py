#!/usr/bin/env python3

import os

from openpyxl import Workbook

from src.dreamai.toolsets.excel_formula_toolset import write_and_evaluate_formula, write_math_function


def test_excel_file_path():
    """Test that excel_file path is returned in the result dictionary."""

    # Create a test Excel file
    test_file = "test_excel_path.xlsx"
    wb = Workbook()
    ws = wb.active
    if ws is not None:
        ws.title = "Sheet1"

        # Add some test data
        ws["B1"] = 10
        ws["B2"] = 20
        ws["B3"] = 30

    wb.save(test_file)

    try:
        print("Testing write_and_evaluate_formula...")

        # Test 1: Simple successful formula
        result = write_and_evaluate_formula(test_file, "Sheet1", "A1", "SUM(B1:B3)")
        print(f"Result: {result}")
        print(f"Excel file path: {result.get('excel_file')}")
        print(f"Success: {result['success']}")
        print(f"Value: {result.get('value')}")
        print()

        # Test 2: Division by zero error
        result = write_and_evaluate_formula(test_file, "Sheet1", "A2", "10/0")
        print(f"Division by zero result: {result}")
        print(f"Excel file path: {result.get('excel_file')}")
        print(f"Success: {result['success']}")
        print(f"Error: {result.get('error')}")
        print()

        # Test 3: Using write_math_function
        print("Testing write_math_function...")
        result = write_math_function(test_file, "Sheet1", "A3", "SUM", ["B1:B3"])
        print(f"Math function result: {result}")
        print(f"Excel file path: {result.get('excel_file')}")
        print(f"Success: {result['success']}")
        print(f"Value: {result.get('value')}")

    finally:
        # Clean up
        if os.path.exists(test_file):
            os.remove(test_file)
            print(f"\nCleaned up test file: {test_file}")


if __name__ == "__main__":
    test_excel_file_path()
