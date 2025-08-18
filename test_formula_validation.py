#!/usr/bin/env python3
"""Test script to verify formula validation in excel_formula_toolset."""

from pathlib import Path

from openpyxl import Workbook

from src.dreamai.toolsets.excel_formula_toolset import (
    FormulaError,
    write_arithmetic_operation,
    write_comparison_operation,
    write_conditional_formula,
    write_date_function,
    write_financial_function,
    write_info_function,
    write_logical_function,
    write_lookup_function,
    write_math_function,
    write_nested_function,
    write_statistical_function,
    write_text_function,
)


def create_test_excel():
    """Create a test Excel file with multiple sheets."""
    wb = Workbook()

    # Create sheets
    ws1 = wb.active
    ws1.title = "Sheet1"
    ws1["A1"] = 10
    ws1["B1"] = 5
    ws1["C1"] = 0
    ws1["D1"] = "Hello"
    ws1["E1"] = "World"

    ws2 = wb.create_sheet("Data")
    ws2["A1"] = 100
    ws2["B1"] = 50
    ws2["A2"] = 200
    ws2["B2"] = 75
    ws2["A3"] = 300
    ws2["B3"] = 25

    ws3 = wb.create_sheet("Results")
    wb.create_sheet("orders")
    wb.create_sheet("analysis")

    wb.save("test_validation.xlsx")
    print("✓ Created test Excel file with sheets: Sheet1, Data, Results")
    return "test_validation.xlsx"


def test_valid_formulas(excel_path: str):
    """Test valid formulas that should work."""
    print("\n=== Testing Valid Formulas ===")

    # Test math functions
    try:
        write_math_function(excel_path, "Sheet1", "F1", "SUM", ["A1:B1"])
        print("✓ Math function: SUM(A1:B1)")
    except Exception as e:
        print(f"✗ Math function failed: {e}")

    try:
        write_math_function(excel_path, "Sheet1", "F2", "MAX", ["A1", "B1", "C1"])
        print("✓ Math function: MAX(A1, B1, C1)")
    except Exception as e:
        print(f"✗ Math function failed: {e}")

    # Test statistical functions
    try:
        write_statistical_function(excel_path, "Data", "C1", "AVERAGE", ["A1:B3"])
        print("✓ Statistical function: AVERAGE(A1:B3)")
    except Exception as e:
        print(f"✗ Statistical function failed: {e}")

    # Test date functions
    try:
        write_date_function(excel_path, "Sheet1", "G1", "TODAY")
        print("✓ Date function: TODAY()")
    except Exception as e:
        print(f"✗ Date function failed: {e}")

    try:
        write_date_function(excel_path, "Sheet1", "G2", "DATE", [2024, 12, 25])
        print("✓ Date function: DATE(2024, 12, 25)")
    except Exception as e:
        print(f"✗ Date function failed: {e}")

    # Test text functions
    try:
        write_text_function(excel_path, "Sheet1", "H1", "CONCATENATE", ["D1", '" "', "E1"])
        print('✓ Text function: CONCATENATE(D1, " ", E1)')
    except Exception as e:
        print(f"✗ Text function failed: {e}")

    try:
        write_text_function(excel_path, "Sheet1", "H2", "UPPER", ["D1"])
        print("✓ Text function: UPPER(D1)")
    except Exception as e:
        print(f"✗ Text function failed: {e}")

    # Test logical functions
    try:
        write_logical_function(excel_path, "Sheet1", "I1", "IF", ["A1>B1", '"Greater"', '"Less or Equal"'])
        print('✓ Logical function: IF(A1>B1, "Greater", "Less or Equal")')
    except Exception as e:
        print(f"✗ Logical function failed: {e}")

    # Test lookup functions
    try:
        write_lookup_function(excel_path, "Data", "D1", "VLOOKUP", ["100", "A1:B3", "2", "FALSE"])
        print("✓ Lookup function: VLOOKUP(100, A1:B3, 2, FALSE)")
    except Exception as e:
        print(f"✗ Lookup function failed: {e}")

    # Test info functions
    try:
        write_info_function(excel_path, "Sheet1", "J1", "ISNUMBER", ["A1"])
        print("✓ Info function: ISNUMBER(A1)")
    except Exception as e:
        print(f"✗ Info function failed: {e}")

    # Test financial functions
    try:
        write_financial_function(excel_path, "Results", "A1", "PMT", ["0.05/12", "60", "-10000"])
        print("✓ Financial function: PMT(0.05/12, 60, -10000)")
    except Exception as e:
        print(f"✗ Financial function failed: {e}")

    # Test arithmetic operations
    try:
        write_arithmetic_operation(excel_path, "Sheet1", "K1", "ADD", ["A1", "B1"])
        print("✓ Arithmetic operation: A1+B1")
    except Exception as e:
        print(f"✗ Arithmetic operation failed: {e}")

    try:
        write_arithmetic_operation(excel_path, "Sheet1", "K2", "DIVIDE", ["A1", "B1"])
        print("✓ Arithmetic operation: A1/B1")
    except Exception as e:
        print(f"✗ Arithmetic operation failed: {e}")

    # Test comparison operations
    try:
        write_comparison_operation(excel_path, "Sheet1", "L1", "A1", ">", "B1")
        print("✓ Comparison operation: A1>B1")
    except Exception as e:
        print(f"✗ Comparison operation failed: {e}")

    # Test nested functions
    try:
        write_nested_function(excel_path, "Sheet1", "M1", "SUM", ["A1", "MAX(B1,C1)"])
        print("✓ Nested function: SUM(A1, MAX(B1,C1))")
    except Exception as e:
        print(f"✗ Nested function failed: {e}")

    # Test conditional formula
    try:
        write_conditional_formula(excel_path, "Sheet1", "N1", "A1>5", "A1*2", "A1/2")
        print("✓ Conditional formula: IF(A1>5, A1*2, A1/2)")
    except Exception as e:
        print(f"✗ Conditional formula failed: {e}")

    # Test cross-sheet reference
    try:
        write_math_function(excel_path, "Results", "B1", "SUM", ["Sheet1!A1:B1", "Data!A1"])
        print("✓ Cross-sheet reference: SUM(Sheet1!A1:B1, Data!A1)")
    except Exception as e:
        print(f"✗ Cross-sheet reference failed: {e}")


def test_invalid_formulas(excel_path: str):
    """Test invalid formulas that should raise errors."""
    print("\n=== Testing Invalid Formulas (Should Raise Errors) ===")

    # Test invalid sheet reference
    try:
        write_math_function(excel_path, "Sheet1", "Z1", "SUM", ["NonExistentSheet!A1:B1"])
        print("✗ Should have failed: Invalid sheet reference not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught invalid sheet reference: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test division by zero
    try:
        write_arithmetic_operation(excel_path, "Sheet1", "Z2", "DIVIDE", ["A1", "0"])
        print("✗ Should have failed: Division by zero not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught division by zero: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test invalid function name
    try:
        write_math_function(excel_path, "Sheet1", "Z3", "INVALID_FUNC", ["A1"])
        print("✗ Should have failed: Invalid function not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught invalid function: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test mismatched parentheses
    try:
        write_nested_function(excel_path, "Sheet1", "Z4", "SUM", ["A1", "MAX(B1,C1"])
        print("✗ Should have failed: Mismatched parentheses not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught mismatched parentheses: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test missing arguments
    try:
        write_statistical_function(excel_path, "Sheet1", "Z5", "AVERAGE", [])
        print("✗ Should have failed: Missing arguments not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught missing arguments: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test invalid operator
    try:
        write_comparison_operation(excel_path, "Sheet1", "Z6", "A1", ">>", "B1")
        print("✗ Should have failed: Invalid operator not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught invalid operator: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")


def test_edge_cases(excel_path: str):
    """Test edge cases and complex formulas."""
    print("\n=== Testing Edge Cases ===")

    # Test sheet name with spaces (needs quotes)
    try:
        wb = Workbook()
        wb.active.title = "Sheet With Spaces"
        wb.save(excel_path)
        write_math_function(excel_path, "Sheet With Spaces", "A1", "SUM", ["1", "2", "3"])
        print("✓ Sheet name with spaces handled correctly")
    except Exception as e:
        print(f"✗ Sheet name with spaces failed: {e}")

    # Recreate standard test file for remaining tests
    create_test_excel()

    # Test complex nested formula
    try:
        write_nested_function(excel_path, "Sheet1", "O1", "IF", ["SUM(A1:B1)>10"], ["MAX(A1,B1)", "MIN(A1,B1)"])
        print("✓ Complex nested formula: IF(SUM(A1:B1)>10, MAX(A1,B1), MIN(A1,B1))")
    except Exception as e:
        print(f"✗ Complex nested formula failed: {e}")

    # Test formula with multiple sheet references
    try:
        write_math_function(excel_path, "Results", "C1", "SUM", ["Sheet1!A1", "Sheet1!B1", "Data!A1", "Data!B1"])
        print("✓ Multiple sheet references: SUM(Sheet1!A1, Sheet1!B1, Data!A1, Data!B1)")
    except Exception as e:
        print(f"✗ Multiple sheet references failed: {e}")

    # Test COUNTIFS with multiple criteria
    try:
        write_math_function(excel_path, "Data", "E1", "COUNTIFS", ["A1:A3", '">100"', "B1:B3", '"<100"'])
        print("✓ COUNTIFS with multiple criteria")
    except Exception as e:
        print(f"✗ COUNTIFS failed: {e}")


def test_malformed_range(excel_path: str):
    """Test malformed range references."""
    print("\n=== Testing Malformed Range References ===")

    # Test malformed range like C:(C)
    try:
        write_math_function(excel_path, "Sheet1", "Z7", "SUM", ["C:(C)"])
        print("✗ Should have failed: Malformed range C:(C) not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught malformed range: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test malformed range with sheet name
    try:
        write_math_function(excel_path, "Sheet1", "Z8", "SUM", ["Data.G:(G)"])
        print("✗ Should have failed: Malformed range Data.G:(G) not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught malformed range with sheet: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")

    # Test the user's specific failing formula
    try:
        formula = '=SUMIFS(orders.G:(G),orders.D:(D),">="&DATE(2023,1,1),orders.D:(D),"<="&DATE(2023,12,31))/SUMPRODUCT((orders.D:(D)>=DATE(2023,1,1))*(orders.D:(D)<=DATE(2023,12,31))*(COUNTIFS(orders.B:(B),orders.B:(B),orders.D:(D),">="&DATE(2023,1,1),orders.D:(D),"<="&DATE(2023,12,31))=1))'
        # We need to use a function that takes a raw formula, or adapt one.
        # For now, we'll simulate the call that would happen inside _write_formula
        from src.dreamai.toolsets.excel_formula_toolset import _validate_formula

        _validate_formula(excel_path, "analysis", "C5", formula)
        print("✗ Should have failed: User's complex formula with malformed ranges not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught user's complex formula error: {e}")
    except Exception as e:
        print(f"✗ Unexpected error in user's formula test: {e}")


def test_invalid_sheet_separator(excel_path: str):
    """Test formulas with invalid sheet separators."""
    print("\n=== Testing Invalid Sheet Separator ===")

    # Test formula with "sheet.range"
    try:
        write_math_function(excel_path, "Results", "D1", "SUM", ["Sheet1.A1"])
        print("✗ Should have failed: Invalid sheet separator '.' not caught")
    except FormulaError as e:
        print(f"✓ Correctly caught invalid sheet separator: {e}")
    except Exception as e:
        print(f"✗ Unexpected error: {e}")


def cleanup():
    """Clean up test files."""
    test_file = Path("test_validation.xlsx")
    if test_file.exists():
        test_file.unlink()
        print("\n✓ Cleaned up test file")


def main():
    """Run all tests."""
    print("=" * 60)
    print("EXCEL FORMULA VALIDATION TEST SUITE")
    print("=" * 60)

    try:
        # Create test Excel file
        excel_path = create_test_excel()

        # Run tests
        test_valid_formulas(excel_path)
        test_invalid_formulas(excel_path)
        test_edge_cases(excel_path)
        test_malformed_range(excel_path)
        test_invalid_sheet_separator(excel_path)

        print("\n" + "=" * 60)
        print("TEST SUITE COMPLETED")
        print("=" * 60)

    finally:
        # Clean up
        cleanup()


if __name__ == "__main__":
    main()
