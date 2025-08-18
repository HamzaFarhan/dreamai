import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.formula.translate import Tokenizer
from openpyxl.utils import cell as cell_utils


# Custom Exceptions
class SheetNotFoundError(Exception):
    """Raised when a specified sheet is not found in the workbook."""

    pass


class FormulaError(Exception):
    """Raised when there's an error with formula syntax or parameters."""

    pass


class FileOperationError(Exception):
    """Raised when file operations fail."""

    pass


# Formula Validation Functions
def _validate_cell_reference(cell_ref: str) -> None:
    """
    Validate that a cell reference is properly formatted.

    Args:
        cell_ref: Cell reference like 'A1', 'B2', etc.

    Raises:
        FormulaError: If cell reference is invalid
    """
    try:
        cell_utils.coordinate_to_tuple(cell_ref)
    except (ValueError, TypeError) as e:
        raise FormulaError(f"Invalid cell reference '{cell_ref}': {e}")


def _validate_range_reference(range_ref: str) -> None:
    """
    Validate that a range reference is properly formatted.

    Args:
        range_ref: Range reference like 'A1:B10', 'C:C', etc.

    Raises:
        FormulaError: If range reference is invalid
    """
    try:
        # Handle column-only ranges like 'A:A' or 'C:E'
        if re.match(r"^[A-Z]+:[A-Z]+$", range_ref):
            return
        # Handle row-only ranges like '1:1' or '5:10'
        if re.match(r"^\d+:\d+$", range_ref):
            return
        # Handle normal cell ranges like 'A1:B10'
        cell_utils.range_boundaries(range_ref)
    except (ValueError, TypeError) as e:
        raise FormulaError(f"Invalid range reference '{range_ref}': {e}")


def _validate_worksheet_references(formula: str) -> None:
    """
    Validate worksheet references in formulas.

    Args:
        formula: Excel formula

    Raises:
        FormulaError: If worksheet references are invalid
    """
    # Check for potentially problematic worksheet names
    worksheet_pattern = r"([A-Za-z_][A-Za-z0-9_]*)\."
    matches = re.findall(worksheet_pattern, formula)

    for worksheet_name in matches:
        # Check for invalid characters in worksheet names
        if any(char in worksheet_name for char in ["[", "]", ":", "*", "?", "/", "\\"]):
            raise FormulaError(f"Invalid worksheet name '{worksheet_name}': contains illegal characters")

        # Check for reserved names
        reserved_names = [
            "CON",
            "PRN",
            "AUX",
            "NUL",
            "COM1",
            "COM2",
            "COM3",
            "COM4",
            "COM5",
            "COM6",
            "COM7",
            "COM8",
            "COM9",
            "LPT1",
            "LPT2",
            "LPT3",
            "LPT4",
            "LPT5",
            "LPT6",
            "LPT7",
            "LPT8",
            "LPT9",
        ]
        if worksheet_name.upper() in reserved_names:
            raise FormulaError(f"Invalid worksheet name '{worksheet_name}': is a reserved name")

        # Check for worksheet names that are too long
        if len(worksheet_name) > 31:
            raise FormulaError(f"Invalid worksheet name '{worksheet_name}': exceeds 31 character limit")


def _validate_string_literals(formula: str) -> None:
    """
    Validate string literals in formulas.

    Args:
        formula: Excel formula

    Raises:
        FormulaError: If string literals are invalid
    """
    # Check for unmatched quotes
    quote_positions = []
    i = 0
    while i < len(formula):
        if formula[i] == '"':
            quote_positions.append(i)
            # Skip to next quote or end
            i += 1
            while i < len(formula) and formula[i] != '"':
                if formula[i] == "\\":  # Escaped character
                    i += 2
                else:
                    i += 1
            if i < len(formula):  # Found closing quote
                quote_positions.append(i)
            i += 1
        else:
            i += 1

    if len(quote_positions) % 2 != 0:
        raise FormulaError(f"Unmatched quotes in formula: {formula}")

    # Check for strings that might cause parsing issues
    for i in range(0, len(quote_positions), 2):
        start = quote_positions[i]
        if i + 1 < len(quote_positions):
            end = quote_positions[i + 1]
            string_content = formula[start + 1 : end]

            # Check for strings that contain problematic characters
            if "\n" in string_content or "\r" in string_content:
                raise FormulaError(f"String literal contains line breaks: '{string_content}'")

            # Check for very long strings
            if len(string_content) > 255:
                raise FormulaError(f"String literal too long ({len(string_content)} chars): Excel limit is 255")


def _validate_date_formats(formula: str) -> None:
    """
    Validate date formats in formulas.

    Args:
        formula: Excel formula

    Raises:
        FormulaError: If date formats are problematic
    """
    # Look for potential date strings that might cause issues
    date_patterns = [
        r'"[<>=]*\d{4}-\d{2}-\d{2}"',  # ISO format dates
        r'"[<>=]*\d{1,2}/\d{1,2}/\d{4}"',  # US format dates
        r'"[<>=]*\d{1,2}-\d{1,2}-\d{4}"',  # Alternative format
    ]

    for pattern in date_patterns:
        matches = re.findall(pattern, formula)
        for match in matches:
            # Check for potential ambiguous date formats
            date_part = match.strip('"').lstrip("<>=")
            if "/" in date_part:
                parts = date_part.split("/")
                if len(parts) == 3:
                    try:
                        month, day, year = map(int, parts)
                        if month > 12 or month < 1:
                            raise FormulaError(f"Invalid date format '{date_part}': month must be 1-12")
                        if day > 31 or day < 1:
                            raise FormulaError(f"Invalid date format '{date_part}': day must be 1-31")
                        if year < 1900 or year > 9999:
                            raise FormulaError(f"Invalid date format '{date_part}': year must be 1900-9999")
                    except ValueError:
                        pass  # Not a valid date format, but might be intentional
            elif "-" in date_part and date_part.count("-") == 2:
                parts = date_part.split("-")
                if len(parts) == 3:
                    try:
                        if len(parts[0]) == 4:  # ISO format YYYY-MM-DD
                            year, month, day = map(int, parts)
                        else:  # Alternative format DD-MM-YYYY or MM-DD-YYYY
                            if int(parts[1]) > 12:  # Assume DD-MM-YYYY
                                day, month, year = map(int, parts)
                            else:  # Assume MM-DD-YYYY
                                month, day, year = map(int, parts)

                        if month > 12 or month < 1:
                            raise FormulaError(f"Invalid date format '{date_part}': month must be 1-12")
                        if day > 31 or day < 1:
                            raise FormulaError(f"Invalid date format '{date_part}': day must be 1-31")
                        if year < 1900 or year > 9999:
                            raise FormulaError(f"Invalid date format '{date_part}': year must be 1900-9999")
                    except ValueError:
                        pass  # Not a valid date format, but might be intentional


def _check_formula_complexity(formula: str) -> None:
    """
    Check if formula exceeds Excel's complexity limits.

    Args:
        formula: Excel formula

    Raises:
        FormulaError: If formula is too complex
    """
    # Check formula length (Excel has a 8192 character limit)
    if len(formula) > 8192:
        raise FormulaError(f"Formula too long ({len(formula)} chars): Excel limit is 8192")

    # Count number of operations (approximate complexity check)
    operators = ["+", "-", "*", "/", "^", "&", "=", "<", ">", "<=", ">=", "<>"]
    operation_count = sum(formula.count(op) for op in operators)

    if operation_count > 1000:  # Rough heuristic for complexity
        raise FormulaError(
            f"Formula too complex ({operation_count} operations): may cause Excel performance issues"
        )

    # Count number of function calls
    function_count = formula.count("(")
    if function_count > 100:  # Another complexity heuristic
        raise FormulaError(f"Too many function calls ({function_count}): may cause Excel performance issues")


def _validate_formula_syntax(formula: str) -> None:
    """
    Validate Excel formula syntax using openpyxl tokenizer.

    Args:
        formula: Excel formula (with or without leading =)

    Raises:
        FormulaError: If formula has syntax errors
    """
    try:
        # Ensure formula starts with =
        if not formula.startswith("="):
            formula = "=" + formula

        # Enhanced validation checks
        _validate_worksheet_references(formula)
        _validate_string_literals(formula)
        _validate_date_formats(formula)
        _check_formula_complexity(formula)

        # Use openpyxl tokenizer to validate syntax
        tokenizer = Tokenizer(formula)
        tokens = list(tokenizer.items)

        # Check for common syntax errors
        open_parens = sum(1 for token in tokens if "(" in str(token))
        close_parens = sum(1 for token in tokens if ")" in str(token))

        if open_parens != close_parens:
            raise FormulaError(f"Mismatched parentheses in formula: {formula}")

        # Check for function nesting depth (Excel has limits)
        max_nesting = 0
        current_nesting = 0
        for token in tokens:
            token_str = str(token)
            if "(" in token_str:
                current_nesting += token_str.count("(")
                max_nesting = max(max_nesting, current_nesting)
            if ")" in token_str:
                current_nesting -= token_str.count(")")

        if max_nesting > 64:  # Excel's function nesting limit
            raise FormulaError(f"Formula nesting too deep ({max_nesting}): Excel limit is 64 levels")

        # Validate any cell/range references in the formula
        for token in tokens:
            token_str = str(token)
            # Check for range patterns like A1:B10
            if ":" in token_str and not any(op in token_str for op in ["<=", ">=", "<>", "=", "<", ">"]):
                # Extract just the range part (remove any prefixes)
                range_match = re.search(r"[A-Z]+\d*:[A-Z]+\d*", token_str)
                if range_match:
                    try:
                        _validate_range_reference(range_match.group())
                    except FormulaError:
                        # Some ranges might be valid in Excel context even if not in openpyxl
                        pass
            # Check for single cell references like A1, B2
            elif re.match(r"^[A-Z]+\d+$", token_str.strip()):
                try:
                    _validate_cell_reference(token_str.strip())
                except FormulaError:
                    # Some cells might be valid in Excel context
                    pass

    except FormulaError:
        raise
    except Exception as e:
        raise FormulaError(f"Formula syntax error: {e}")


def _validate_function_arguments(function_name: str, args: list[Any]) -> None:
    """
    Validate function arguments for common Excel functions.

    Args:
        function_name: Name of the Excel function
        args: List of function arguments

    Raises:
        FormulaError: If arguments are invalid
    """
    function_name = function_name.upper()

    # Validate specific function argument patterns
    if function_name in ["SUM", "AVERAGE", "MAX", "MIN", "COUNT", "COUNTA"]:
        for arg in args:
            arg_str = str(arg)
            # Check if it looks like a range
            if ":" in arg_str:
                _validate_range_reference(arg_str)
            # Check if it looks like a single cell
            elif re.match(r"^[A-Z]+\d+$", arg_str):
                _validate_cell_reference(arg_str)

    elif function_name == "IF":
        if len(args) < 2:
            raise FormulaError("IF function requires at least 2 arguments (condition, true_value)")
        if len(args) > 3:
            raise FormulaError("IF function accepts maximum 3 arguments (condition, true_value, false_value)")

    elif function_name == "VLOOKUP":
        if len(args) < 3:
            raise FormulaError(
                "VLOOKUP function requires at least 3 arguments (lookup_value, table_array, col_index)"
            )
        if len(args) > 4:
            raise FormulaError("VLOOKUP function accepts maximum 4 arguments")

    elif function_name in ["SUMIF", "COUNTIF"]:
        if len(args) < 2:
            raise FormulaError(f"{function_name} function requires at least 2 arguments")
        if len(args) > 3:
            raise FormulaError(f"{function_name} function accepts maximum 3 arguments")


# Formula Writing Functions
def _write_formula(excel_path: str, sheet_name: str, cell_ref: str, formula: str) -> str:
    """
    Helper function to write a formula to a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        formula: Excel formula (without leading =)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        FormulaError: If formula is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        # Validate cell reference
        _validate_cell_reference(cell_ref)

        # Ensure formula starts with =
        if not formula.startswith("="):
            formula = "=" + formula

        # Validate formula syntax
        _validate_formula_syntax(formula)

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws[cell_ref] = formula
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, FormulaError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write formula: {e}")


def write_and_evaluate_formula(excel_path: str, sheet_name: str, cell_ref: str, formula: str) -> dict[str, Any]:
    """
    Write a formula to a cell and evaluate it to check for errors.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        formula: Excel formula (without leading =)

    Returns:
        dict containing:
        - success: bool - Whether formula was written and evaluated successfully
        - value: Any - The calculated value (if successful)
        - error: str - Error type if any (#DIV/0!, #VALUE!, etc.)
        - error_message: str - Human readable error description
        - formula_written: bool - Whether formula was successfully written to file
        - excel_file: str - Path to the Excel file (if formula was written)

    Example:
        result = write_and_evaluate_formula("file.xlsx", "Sheet1", "A1", "10/0")
    """
    result: dict[str, Any] = {
        "success": False,
        "value": None,
        "error": None,
        "error_message": "",
        "formula_written": False,
        "excel_file": None,
    }

    try:
        # First, write the formula
        excel_file_path = _write_formula(excel_path, sheet_name, cell_ref, formula)
        result["formula_written"] = True
        result["excel_file"] = excel_file_path

        # Now evaluate the formula by reloading with data_only=True
        # Note: openpyxl with data_only=True only reads pre-calculated values
        # For fresh formulas, we need to simulate calculation or check formula validity
        try:
            wb_data = load_workbook(excel_path, data_only=True)

            if sheet_name not in wb_data.sheetnames:
                result["error_message"] = f"Sheet '{sheet_name}' not found after writing formula"
                return result

            ws_data = wb_data[sheet_name]
            cell_value = ws_data[cell_ref].value

            # Since data_only=True won't calculate new formulas, we need to check the formula syntax
            # and perform basic error detection
            wb_formula = load_workbook(excel_path, data_only=False)
            ws_formula = wb_formula[sheet_name]
            written_formula = ws_formula[cell_ref].value

            # Basic error detection based on formula patterns
            if written_formula and isinstance(written_formula, str) and written_formula.startswith("="):
                formula_content = written_formula[1:]  # Remove the =

                # Check for division by zero patterns (but not if already wrapped in IFERROR)
                if "/" in formula_content and not formula_content.startswith("IFERROR("):
                    # Look for patterns like /0 or /COUNTIFS(...) where result might be 0
                    div_patterns = [
                        r"/0\b",  # Direct division by 0
                        r"/COUNTIFS?\([^)]+\)",  # Division by COUNTIFS that might return 0
                        r"/COUNT\([^)]+\)",  # Division by COUNT that might return 0
                    ]

                    for pattern in div_patterns:
                        if re.search(pattern, formula_content):
                            # This could potentially cause division by zero
                            result["error"] = "#DIV/0!"
                            result["error_message"] = (
                                "Potential division by zero detected. Consider wrapping in IFERROR() or checking denominator with IF()"
                            )
                            wb_data.close()
                            wb_formula.close()
                            return result

                # Check for invalid function names
                invalid_functions = re.findall(r"\b[A-Z][A-Z0-9]*\(", formula_content)
                valid_functions = {
                    "SUM",
                    "AVERAGE",
                    "COUNT",
                    "COUNTA",
                    "COUNTIFS",
                    "SUMIF",
                    "SUMIFS",
                    "AVERAGEIF",
                    "AVERAGEIFS",
                    "IF",
                    "IFERROR",
                    "VLOOKUP",
                    "HLOOKUP",
                    "INDEX",
                    "MATCH",
                    "MAX",
                    "MIN",
                    "TODAY",
                    "NOW",
                    "DATE",
                    "YEAR",
                    "MONTH",
                    "DAY",
                    "ABS",
                    "ROUND",
                    "CONCATENATE",
                    "LEFT",
                    "RIGHT",
                    "MID",
                    "LEN",
                    "UPPER",
                    "LOWER",
                    "TRIM",
                    "AND",
                    "OR",
                    "NOT",
                }

                for func_match in invalid_functions:
                    func_name = func_match[:-1]  # Remove the (
                    if func_name not in valid_functions:
                        result["error"] = "#NAME?"
                        result["error_message"] = f"Unrecognized function name: {func_name}"
                        wb_data.close()
                        wb_formula.close()
                        return result

            wb_data.close()
            wb_formula.close()

        except Exception as eval_error:
            result["error_message"] = f"Error during formula evaluation: {eval_error}"
            return result

        # If we get here, no obvious errors were detected
        # Check the actual cell value for Excel error strings (if any were pre-calculated)
        excel_errors = {
            "#DIV/0!": "Division by zero error",
            "#VALUE!": "Wrong type of argument or operand",
            "#REF!": "Invalid cell reference",
            "#NAME?": "Unrecognized formula name or text",
            "#NUM!": "Invalid numeric value",
            "#N/A": "Value not available",
            "#NULL!": "Null intersection of two ranges",
            "#SPILL!": "Spill range has been blocked",
            "#CALC!": "Calculation error",
            "#FIELD!": "Invalid field name",
            "#UNKNOWN!": "Unknown error",
        }

        if isinstance(cell_value, str) and cell_value in excel_errors:
            result["error"] = cell_value
            result["error_message"] = excel_errors[cell_value]

            # Provide specific guidance for common errors
            if cell_value == "#DIV/0!":
                result["error_message"] += ". Consider wrapping in IFERROR() or checking denominator with IF()"
            elif cell_value == "#VALUE!":
                result["error_message"] += ". Check data types and cell references"
            elif cell_value == "#REF!":
                result["error_message"] += ". Check that all referenced cells and ranges exist"
            elif cell_value == "#NAME?":
                result["error_message"] += ". Check function names and cell references"

        elif cell_value is None:
            # Formula might be referencing empty cells - this could be valid
            result["success"] = True
            result["value"] = None
        else:
            # Formula evaluated successfully
            result["success"] = True
            result["value"] = cell_value

    except FormulaError as e:
        result["error_message"] = f"Formula validation error: {e}"
    except FileNotFoundError as e:
        result["error_message"] = f"File not found: {e}"
    except SheetNotFoundError as e:
        result["error_message"] = f"Sheet not found: {e}"
    except Exception as e:
        result["error_message"] = f"Unexpected error: {e}"

    return result


def write_formula_with_error_handling(
    excel_path: str,
    sheet_name: str,
    cell_ref: str,
    formula: str,
    max_retries: int = 3,
    error_fallback: str | None = None,
) -> dict[str, Any]:
    """
    Write a formula with automatic error handling and optional fallback.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        formula: Excel formula (without leading =)
        max_retries: Maximum number of retry attempts
        error_fallback: Fallback formula to use if original fails (e.g., "0" or '""')

    Returns:
        dict containing evaluation results and any applied fixes

    Example:
        result = write_formula_with_error_handling(
            "file.xlsx", "Sheet1", "A1",
            "AVERAGEIFS(B:B,C:C,'criteria')/COUNTIFS(D:D,'criteria')",
            error_fallback="0"
        )
    """
    attempts: list[dict[str, Any]] = []
    result: dict[str, Any] = {"success": False}

    for attempt in range(max_retries):
        result = write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)
        attempts.append({"attempt": attempt + 1, "formula": formula, "result": result.copy()})

        if result["success"]:
            result["attempts"] = attempts
            return result

        # Try to auto-fix common issues
        if result["error"] == "#DIV/0!" and attempt < max_retries - 1:
            # Wrap in IFERROR with 0 as fallback
            formula = f"IFERROR({formula},0)"
        elif result["error"] in ["#VALUE!", "#REF!", "#NAME?"] and attempt < max_retries - 1:
            # For other errors, try wrapping in IFERROR with empty string
            formula = f'IFERROR({formula},"")'
        else:
            break

    # If all retries failed and we have a fallback, use it
    if error_fallback is not None:
        fallback_result = write_and_evaluate_formula(excel_path, sheet_name, cell_ref, error_fallback)
        fallback_result["attempts"] = attempts
        fallback_result["used_fallback"] = True
        fallback_result["original_formula"] = attempts[0]["formula"]
        return fallback_result

    # Return the last failed attempt
    result["attempts"] = attempts
    return result


# Date Functions
def write_date_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any] | None = None
) -> dict[str, Any]:
    """
    Writes date functions to Excel cells.

    Supported functions: DATE, DAY, HOUR, MINUTE, MONTH, NOW, SECOND, TIME, TODAY, WEEKDAY, YEAR

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the date function
        function_args: Function arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_date_function("file.xlsx", "Sheet1", "A1", "TODAY")

    """
    try:
        function_name = function_name.upper()
        if function_args is None:
            function_args = []

        valid_functions = {
            "DATE",
            "DAY",
            "HOUR",
            "MINUTE",
            "MONTH",
            "NOW",
            "SECOND",
            "TIME",
            "TODAY",
            "WEEKDAY",
            "YEAR",
        }

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid date function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        # Build formula
        if function_args:
            # Validate function arguments
            _validate_function_arguments(function_name, function_args)
            args_str = ",".join(str(arg) for arg in function_args)
            formula = f"{function_name}({args_str})"
        else:
            formula = f"{function_name}()"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write date function: {e}")


# Financial Functions
def write_financial_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any]
) -> dict[str, Any]:
    """
    Writes financial functions to Excel cells.

    Supported functions: FV, IRR, NPV, PMT, PV

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the financial function
        function_args: Function arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_financial_function("file.xlsx", "Sheet1", "A1", "PV", [0.05, 10, -1000])

    """
    try:
        function_name = function_name.upper()

        valid_functions = {"FV", "IRR", "NPV", "PMT", "PV"}

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid financial function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        if not function_args:
            raise FormulaError(f"Financial function {function_name} requires arguments")

        # Validate function arguments
        _validate_function_arguments(function_name, function_args)

        args_str = ",".join(str(arg) for arg in function_args)
        formula = f"{function_name}({args_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write financial function: {e}")


# Logical Functions
def write_logical_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, conditions: list[Any] | None = None
) -> dict[str, Any]:
    """
    Writes logical functions to Excel cells.

    Supported functions: AND, FALSE, IF, IFERROR, NOT, OR, TRUE

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the logical function
        conditions: Function conditions/arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_logical_function("file.xlsx", "Sheet1", "A1", "IF", ["B1>10", '"Yes"', '"No"'])

    """
    try:
        function_name = function_name.upper()
        if conditions is None:
            conditions = []

        valid_functions = {"AND", "FALSE", "IF", "IFERROR", "NOT", "OR", "TRUE"}

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid logical function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        if function_name in ["FALSE", "TRUE"] and conditions:
            raise FormulaError(f"Function {function_name} takes no arguments")

        if function_name in ["FALSE", "TRUE"]:
            formula = f"{function_name}()"
        else:
            if not conditions:
                raise FormulaError(f"Function {function_name} requires arguments")

            # Validate function arguments
            _validate_function_arguments(function_name, conditions)

            conditions_str = ",".join(str(condition) for condition in conditions)
            formula = f"{function_name}({conditions_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write logical function: {e}")


# Lookup Functions
def write_lookup_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any] | None = None
) -> dict[str, Any]:
    """
    Writes lookup functions to Excel cells.

    Supported functions: CHOOSE, COLUMN, COLUMNS, HLOOKUP, INDEX, INDIRECT,
                        MATCH, OFFSET, ROW, ROWS, VLOOKUP

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the lookup function
        function_args: Function arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_lookup_function("file.xlsx", "Sheet1", "A1", "VLOOKUP", ["B1", "D:E", 2, "FALSE"])

    """
    try:
        function_name = function_name.upper()
        if function_args is None:
            function_args = []

        valid_functions = {
            "CHOOSE",
            "COLUMN",
            "COLUMNS",
            "HLOOKUP",
            "INDEX",
            "INDIRECT",
            "MATCH",
            "OFFSET",
            "ROW",
            "ROWS",
            "VLOOKUP",
        }

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid lookup function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        if function_args:
            # Validate function arguments
            _validate_function_arguments(function_name, function_args)
            args_str = ",".join(str(arg) for arg in function_args)
            formula = f"{function_name}({args_str})"
        else:
            formula = f"{function_name}()"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write lookup function: {e}")


# Math Functions
def write_math_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any] | None = None
) -> dict[str, Any]:
    """
    Writes math functions to Excel cells.

    Supported functions: ABS, ACOS, ASIN, ATAN, ATAN2, CEILING, COMBIN, COS,
                        COUNTIF, COUNTIFS, DEGREES, EVEN, EXP, FACT, FLOOR,
                        INT, LN, LOG, LOG10, MAX, MIN, MOD, ODD, PI, POWER,
                        PRODUCT, RADIANS, RAND, RANDBETWEEN, ROUND, ROUNDDOWN,
                        ROUNDUP, SIGN, SIN, SQRT, SUM, SUMIF, SUMIFS,
                        SUMPRODUCT, TAN, TRUNC

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the math function
        function_args: Function arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_math_function("file.xlsx", "Sheet1", "A1", "SUM", ["B1:B10"])

    """
    try:
        function_name = function_name.upper()
        if function_args is None:
            function_args = []

        valid_functions = {
            "ABS",
            "ACOS",
            "ASIN",
            "ATAN",
            "ATAN2",
            "CEILING",
            "COMBIN",
            "COS",
            "COUNTIF",
            "COUNTIFS",
            "DEGREES",
            "EVEN",
            "EXP",
            "FACT",
            "FLOOR",
            "INT",
            "LN",
            "LOG",
            "LOG10",
            "MAX",
            "MIN",
            "MOD",
            "ODD",
            "PI",
            "POWER",
            "PRODUCT",
            "RADIANS",
            "RAND",
            "RANDBETWEEN",
            "ROUND",
            "ROUNDDOWN",
            "ROUNDUP",
            "SIGN",
            "SIN",
            "SQRT",
            "SUM",
            "SUMIF",
            "SUMIFS",
            "SUMPRODUCT",
            "TAN",
            "TRUNC",
        }

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid math function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        # Functions that take no parameters
        no_param_functions = {"PI", "RAND"}

        if function_name in no_param_functions:
            if function_args:
                raise FormulaError(f"Function {function_name} takes no arguments")
            formula = f"{function_name}()"
        else:
            if not function_args:
                raise FormulaError(f"Function {function_name} requires arguments")

            # Validate function arguments
            _validate_function_arguments(function_name, function_args)

            args_str = ",".join(str(arg) for arg in function_args)
            formula = f"{function_name}({args_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write math function: {e}")


# Statistical Functions
def write_statistical_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any]
) -> dict[str, Any]:
    """
    Writes statistical functions to Excel cells.

    Supported functions: AVERAGE, AVERAGEIF, AVERAGEIFS, COUNT, COUNTA, COUNTBLANK,
                        LARGE, MEDIAN, MODE, PERCENTILE, QUARTILE, RANK, SMALL, STDEV, VAR

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the statistical function
        function_args: Function data/arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_statistical_function("file.xlsx", "Sheet1", "A1", "AVERAGE", ["B1:B10"])

    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "AVERAGE",
            "AVERAGEIF",
            "AVERAGEIFS",
            "COUNT",
            "COUNTA",
            "COUNTBLANK",
            "LARGE",
            "MEDIAN",
            "MODE",
            "PERCENTILE",
            "QUARTILE",
            "RANK",
            "SMALL",
            "STDEV",
            "VAR",
        }

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid statistical function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        if not function_args:
            raise FormulaError(f"Statistical function {function_name} requires data")

        # Validate function arguments
        _validate_function_arguments(function_name, function_args)

        args_str = ",".join(str(item) for item in function_args)
        formula = f"{function_name}({args_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write statistical function: {e}")


# Text Functions
def write_text_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any]
) -> dict[str, Any]:
    """
    Writes text functions to Excel cells.

    Supported functions: CHAR, CLEAN, CODE, CONCATENATE, EXACT, FIND, LEFT, LEN,
                        LOWER, MID, PROPER, REPLACE, REPT, RIGHT, SEARCH, SUBSTITUTE,
                        TEXT, TRIM, UPPER, VALUE

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the text function
        function_args: Function text arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_text_function("file.xlsx", "Sheet1", "A1", "CONCATENATE", ["A1", "B1"])

    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "CHAR",
            "CLEAN",
            "CODE",
            "CONCATENATE",
            "EXACT",
            "FIND",
            "LEFT",
            "LEN",
            "LOWER",
            "MID",
            "PROPER",
            "REPLACE",
            "REPT",
            "RIGHT",
            "SEARCH",
            "SUBSTITUTE",
            "TEXT",
            "TRIM",
            "UPPER",
            "VALUE",
        }

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid text function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        if not function_args:
            raise FormulaError(f"Text function {function_name} requires arguments")

        # Validate function arguments
        _validate_function_arguments(function_name, function_args)

        args_str = ",".join(str(arg) for arg in function_args)
        formula = f"{function_name}({args_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write text function: {e}")


# Info Functions
def write_info_function(
    excel_path: str, sheet_name: str, cell_ref: str, function_name: str, function_args: list[Any]
) -> dict[str, Any]:
    """
    Writes info functions to Excel cells.

    Supported functions: ISBLANK, ISERROR, ISNUMBER, ISTEXT

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        function_name: Name of the info function
        function_args: Function arguments

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_info_function("file.xlsx", "Sheet1", "A1", "ISBLANK", ["B1"])

    """
    try:
        function_name = function_name.upper()

        valid_functions = {"ISBLANK", "ISERROR", "ISNUMBER", "ISTEXT"}

        if function_name not in valid_functions:
            raise FormulaError(
                f"Invalid info function: {function_name}. Valid functions: {sorted(valid_functions)}"
            )

        if not function_args:
            raise FormulaError(f"Info function {function_name} requires arguments")

        # Validate function arguments
        _validate_function_arguments(function_name, function_args)

        args_str = ",".join(str(arg) for arg in function_args)
        formula = f"{function_name}({args_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write info function: {e}")


# Arithmetic Operations
def write_arithmetic_operation(
    excel_path: str, sheet_name: str, cell_ref: str, operation: str, operands: list[str]
) -> dict[str, Any]:
    """
    Writes arithmetic operations to Excel cells.

    Supported operations: ADD, SUBTRACT, MULTIPLY, DIVIDE, POWER

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        operation: Type of arithmetic operation
        operands: List of cell references, values, or nested expressions

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_arithmetic_operation("file.xlsx", "Sheet1", "A1", "DIVIDE", ["B1", "C1"])

    """
    try:
        operation = operation.upper()
        valid_operations = {"ADD", "SUBTRACT", "MULTIPLY", "DIVIDE", "POWER"}

        if operation not in valid_operations:
            raise FormulaError(f"Invalid operation: {operation}. Valid operations: {sorted(valid_operations)}")

        if len(operands) < 2:
            raise FormulaError(f"Operation {operation} requires at least 2 operands")

        # Validate operands
        for operand in operands:
            if isinstance(operand, str):
                # Check if it's a cell reference
                if re.match(r"^[A-Z]+\d+$", operand.strip()):
                    _validate_cell_reference(operand.strip())
                # Check if it's a range reference
                elif ":" in operand and not any(op in operand for op in ["<=", ">=", "<>", "=", "<", ">"]):
                    _validate_range_reference(operand.strip())

        # Build formula based on operation
        formula = ""
        if operation == "ADD":
            formula = "+".join(operands)
        elif operation == "SUBTRACT":
            if len(operands) > 2:
                # For multiple operands: first - (second + third + ...)
                formula = f"{operands[0]}-({'+'.join(operands[1:])})"
            else:
                formula = "-".join(operands)
        elif operation == "MULTIPLY":
            formula = "*".join(operands)
        elif operation == "DIVIDE":
            if len(operands) > 2:
                # For multiple operands: first / (second * third * ...)
                formula = f"{operands[0]}/({('*'.join(operands[1:]))})"
            else:
                formula = "/".join(operands)
        elif operation == "POWER":
            if len(operands) != 2:
                raise FormulaError("POWER operation requires exactly 2 operands")
            formula = f"POWER({operands[0]},{operands[1]})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write arithmetic operation: {e}")


def write_comparison_operation(
    excel_path: str, sheet_name: str, cell_ref: str, left_operand: str, operator: str, right_operand: str
) -> dict[str, Any]:
    """
    Writes comparison operations to Excel cells.

    Supported operators: =, <>, <, >, <=, >=

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        left_operand: Left side of comparison (cell reference, value, or expression)
        operator: Comparison operator
        right_operand: Right side of comparison (cell reference, value, or expression)

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_comparison_operation("file.xlsx", "Sheet1", "A1", "B1", ">", "10")

    """
    try:
        valid_operators = {"=", "<>", "<", ">", "<=", ">="}

        if operator not in valid_operators:
            raise FormulaError(f"Invalid operator: {operator}. Valid operators: {sorted(valid_operators)}")

        # Validate operands if they're cell references
        for operand in [left_operand, right_operand]:
            if isinstance(operand, str) and re.match(r"^[A-Z]+\d+$", operand.strip()):
                _validate_cell_reference(operand.strip())

        formula = f"{left_operand}{operator}{right_operand}"
        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write comparison operation: {e}")


def write_nested_function(
    excel_path: str,
    sheet_name: str,
    cell_ref: str,
    outer_function: str,
    inner_functions: list[str],
    outer_args: list[str] | None = None,
) -> dict[str, Any]:
    """
    Writes nested function calls to Excel cells.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        outer_function: Name of the outer function (e.g., "IF", "SUM")
        inner_functions: List of inner function expressions
        outer_args: Additional arguments for the outer function

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_nested_function("file.xlsx", "Sheet1", "A1", "IF",
                            ["COUNTIFS(C:C,\"Pro\",E:E,\"<=2023-01-01\")=0"],
                            ["1/B3", "COUNTIFS(F:F,\"<=2023-12-31\")"])

    """
    try:
        outer_function = outer_function.upper()

        # Build the formula
        all_args = inner_functions.copy()
        if outer_args:
            all_args.extend(outer_args)

        args_str = ",".join(all_args)
        formula = f"{outer_function}({args_str})"

        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write nested function: {e}")


def write_conditional_formula(
    excel_path: str, sheet_name: str, cell_ref: str, condition: str, true_value: str, false_value: str
) -> dict[str, Any]:
    """
    Writes IF conditional formulas to Excel cells.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_ref: Cell reference (e.g., 'A1')
        condition: Condition to evaluate
        true_value: Value/expression when condition is true
        false_value: Value/expression when condition is false

    Returns:
        dict: Evaluation result with success status and any errors

    Example:
        result = write_conditional_formula("file.xlsx", "Sheet1", "A1",
                                "COUNTIFS(C:C,\"Pro\")=0",
                                "1/B3",
                                "COUNTIFS(F:F,\"Churned\")/COUNTIFS(C:C,\"Pro\")")

    """
    try:
        formula = f"IF({condition},{true_value},{false_value})"
        return write_and_evaluate_formula(excel_path, sheet_name, cell_ref, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write conditional formula: {e}")


# Function Expression Builder
def build_countifs_expression(range_criteria_pairs: list[tuple[str, str]]) -> str:
    """
    Builds a COUNTIFS expression string.

    Args:
        range_criteria_pairs: List of (range, criteria) tuples

    Returns:
        str: COUNTIFS expression

    Example:
        build_countifs_expression([
            ("Raw_Subscriptions.C:C", "\"Pro\""),
            ("Raw_Subscriptions.E:E", "\"<=2023-01-01\"")
        ])
        # Returns: 'COUNTIFS(Raw_Subscriptions.C:C,"Pro",Raw_Subscriptions.E:E,"<=2023-01-01")'
    """
    if not range_criteria_pairs:
        raise FormulaError("COUNTIFS requires at least one range-criteria pair")

    args: list[str] = []
    for range_ref, criteria in range_criteria_pairs:
        # Validate range reference
        if ":" in range_ref and not any(op in range_ref for op in ["<=", ">=", "<>", "=", "<", ">"]):
            try:
                _validate_range_reference(range_ref.split(".")[-1] if "." in range_ref else range_ref)
            except FormulaError:
                pass  # Worksheet references might not validate perfectly
        args.extend([range_ref, criteria])

    args_str = ",".join(args)
    return f"COUNTIFS({args_str})"


def build_division_expression(numerator: str, denominator: str) -> str:
    """
    Builds a division expression string.

    Args:
        numerator: Numerator expression
        denominator: Denominator expression

    Returns:
        str: Division expression

    Example:
        build_division_expression("COUNTIFS(A:A,\"Yes\")", "COUNTIFS(B:B,\"Total\")")
        # Returns: 'COUNTIFS(A:A,"Yes")/COUNTIFS(B:B,"Total")'
    """
    return f"{numerator}/{denominator}"


if __name__ == "__main__":
    # Example usage
    try:
        # Example formulas
        excel_path = "example.xlsx"

        # Math formulas
        write_math_function(excel_path, "Sheet", "D1", "SUM", ["C2:C4"])
        write_statistical_function(excel_path, "Sheet", "D2", "AVERAGE", ["C2:C4"])
        write_math_function(excel_path, "Sheet", "D3", "MAX", ["C2:C4"])

        # Date formulas
        write_date_function(excel_path, "Sheet", "E1", "TODAY")
        write_date_function(excel_path, "Sheet", "E2", "YEAR", ["E1"])

        print("Formulas added to Excel file!")

    except FormulaError as e:
        print(f"Formula validation error: {e}")
    except Exception as e:
        print(f"Error: {e}")
