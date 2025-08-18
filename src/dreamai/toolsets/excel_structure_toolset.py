import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.pivot.cache import CacheDefinition, WorksheetSource


# Custom Exceptions
class SheetNotFoundError(Exception):
    """Raised when a specified sheet is not found in the workbook."""

    pass


class InvalidSheetNameError(Exception):
    """Raised when an invalid sheet name is provided."""

    pass


class FileOperationError(Exception):
    """Raised when file operations fail."""

    pass


# Core File Operations
def create_excel_file(file_path: str) -> str:
    """
    Creates a new Excel workbook at the specified path.

    Args:
        file_path: Path where the Excel file will be created

    Returns:
        str: Path of the created file

    Raises:
        FileExistsError: If file already exists
        PermissionError: If write access is denied
        FileOperationError: If file creation fails
    """
    try:
        path = Path(file_path)

        if path.exists():
            raise FileExistsError(f"File already exists: {file_path}")

        # Ensure directory exists
        path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()
        wb.save(file_path)

        return str(path.absolute())

    except FileExistsError:
        raise
    except PermissionError as e:
        raise PermissionError(f"Permission denied creating file {file_path}: {e}")
    except Exception as e:
        raise FileOperationError(f"Failed to create Excel file {file_path}: {e}")


def csv_to_excel_sheet(csv_path: str, excel_path: str, sheet_name: str | None = None) -> str:
    """
    Adds a CSV file as a new sheet in an Excel workbook.

    Args:
        csv_path: Path to the CSV file
        excel_path: Path to the Excel file (created if doesn't exist)
        sheet_name: Optional sheet name (default: CSV filename)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If CSV file doesn't exist
        ValueError: For invalid CSV format
        FileOperationError: If operation fails
    """
    try:
        csv_file = Path(csv_path)
        if not csv_file.exists():
            raise FileNotFoundError(f"CSV file not found: {csv_path}")

        excel_file = Path(excel_path)

        # Determine sheet name
        if sheet_name is None:
            sheet_name = csv_file.stem

        # Load or create workbook
        if excel_file.exists():
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            # Remove default sheet if we're adding a named sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create new sheet
        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        ws = wb.create_sheet(title=sheet_name)

        # Read CSV and write to sheet
        with open(csv_path, "r", encoding="utf-8") as csvfile:
            # Try to detect delimiter with fallback
            sample = csvfile.read(1024)
            csvfile.seek(0)

            delimiter = ","  # Default fallback
            try:
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample, delimiters=",;\t|").delimiter
            except csv.Error:
                # If sniffer fails, try common delimiters in order of preference
                for test_delimiter in [",", ";", "\t", "|"]:
                    if test_delimiter in sample:
                        delimiter = test_delimiter
                        break

            reader = csv.reader(csvfile, delimiter=delimiter)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    # Skip empty values
                    if not value or value.strip() == "":
                        continue

                    # Try to convert to number if possible
                    try:
                        if "." in value:
                            ws.cell(row=row_idx, column=col_idx, value=float(value))
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=int(value))
                    except (ValueError, TypeError):
                        ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except FileNotFoundError:
        raise
    except ValueError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to convert CSV to Excel: {e}")


def csvs_to_excel(csv_paths: list[str], excel_path: str) -> str:
    """
    Adds multiple CSV files as sheets in a single Excel workbook.

    Args:
        csv_paths: List of CSV file paths
        excel_path: Output Excel file path

    Returns:
        str: Excel file path

    Raises:
        ValueError: If no CSV paths provided
        FileNotFoundError: If any CSV file is missing
        FileOperationError: If operation fails
    """
    try:
        if not csv_paths:
            raise ValueError("No CSV paths provided")

        excel_file = Path(excel_path)

        # Create new workbook
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for csv_path in csv_paths:
            csv_file = Path(csv_path)
            if not csv_file.exists():
                raise FileNotFoundError(f"CSV file not found: {csv_path}")

            sheet_name = csv_file.stem

            # Ensure unique sheet name
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1

            ws = wb.create_sheet(title=sheet_name)

            # Read CSV and write to sheet
            with open(csv_path, "r", encoding="utf-8") as csvfile:
                sample = csvfile.read(1024)
                csvfile.seek(0)

                delimiter = ","  # Default fallback
                try:
                    sniffer = csv.Sniffer()
                    delimiter = sniffer.sniff(sample, delimiters=",;\t|").delimiter
                except csv.Error:
                    # If sniffer fails, try common delimiters in order of preference
                    for test_delimiter in [",", ";", "\t", "|"]:
                        if test_delimiter in sample:
                            delimiter = test_delimiter
                            break

                reader = csv.reader(csvfile, delimiter=delimiter)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        # Skip empty values
                        if not value or value.strip() == "":
                            continue

                        try:
                            if "." in value:
                                ws.cell(row=row_idx, column=col_idx, value=float(value))
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=int(value))
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (ValueError, FileNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to convert CSVs to Excel: {e}")


def add_sheet(excel_path: str, sheet_name: str) -> str:
    """
    Adds a new empty sheet to an Excel workbook.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the new sheet

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet name exists or is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        if not sheet_name or len(sheet_name.strip()) == 0:
            raise ValueError("Sheet name cannot be empty")

        wb.create_sheet(title=sheet_name)
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to add sheet: {e}")


def write_data_to_sheet(
    excel_path: str, sheet_name: str, data: list[list[Any]], start_row: int = 1, start_col: int = 1
) -> str:
    """
    Writes 2D data to a specific sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data: 2D list of values
        start_row: Starting row (1-based)
        start_col: Starting column (1-based)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If data is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        if not data:
            raise ValueError("Data cannot be empty")

        ws = wb[sheet_name]

        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write data to sheet: {e}")


def get_sheet_names(excel_path: str) -> list[str]:
    """
    Returns list of sheet names in workbook.

    Args:
        excel_path: Path to the Excel file

    Returns:
        List[str]: List of sheet names

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If invalid Excel file
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)
        return wb.sheetnames

    except FileNotFoundError:
        raise
    except Exception as e:
        # Handle all zip/file format related errors
        if "zip" in str(e).lower() or "bad zip" in str(e).lower():
            raise ValueError(f"Invalid Excel file: {excel_path}")
        raise FileOperationError(f"Failed to get sheet names: {e}")


def delete_sheet(excel_path: str, sheet_name: str) -> str:
    """
    Deletes a sheet from workbook.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to delete

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        RuntimeError: If trying to delete the last sheet
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        if len(wb.sheetnames) == 1:
            raise RuntimeError("Cannot delete the last sheet in workbook")

        wb.remove(wb[sheet_name])
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, RuntimeError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to delete sheet: {e}")


def rename_sheet(excel_path: str, old_name: str, new_name: str) -> str:
    """
    Renames an existing sheet.

    Args:
        excel_path: Path to the Excel file
        old_name: Current sheet name
        new_name: New sheet name

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If old sheet doesn't exist
        ValueError: If new name is invalid or exists
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if old_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{old_name}' not found")

        if new_name in wb.sheetnames:
            raise ValueError(f"Sheet '{new_name}' already exists")

        if not new_name or len(new_name.strip()) == 0:
            raise ValueError("New sheet name cannot be empty")

        wb[old_name].title = new_name
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to rename sheet: {e}")


def copy_sheet(excel_path: str, sheet_name: str, new_name: str) -> str:
    """
    Copies a sheet within the same workbook.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to copy
        new_name: Name for the copied sheet

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If source sheet doesn't exist
        ValueError: If new name exists or is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        if new_name in wb.sheetnames:
            raise ValueError(f"Sheet '{new_name}' already exists")

        if not new_name or len(new_name.strip()) == 0:
            raise ValueError("New sheet name cannot be empty")

        source_sheet = wb[sheet_name]
        copied_sheet = wb.copy_worksheet(source_sheet)
        copied_sheet.title = new_name

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to copy sheet: {e}")


def update_sheet_properties(
    excel_path: str,
    sheet_name: str,
    new_title: str | None = None,
    row_count: int | None = None,
    column_count: int | None = None,
) -> str:
    """
    Updates properties of an existing sheet (title and/or dimensions).

    Args:
        excel_path: Path to the Excel file
        sheet_name: Current name of the sheet to update
        new_title: New title for the sheet (optional)
        row_count: New number of rows (optional)
        column_count: New number of columns (optional)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If new title exists or no properties provided
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        if new_title is None and row_count is None and column_count is None:
            raise ValueError("At least one property (new_title, row_count, column_count) must be provided")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in {excel_path}")

        # Check if new title conflicts with existing sheet names
        if new_title is not None and new_title != sheet_name and new_title in wb.sheetnames:
            raise ValueError(f"Sheet with title '{new_title}' already exists")

        ws = wb[sheet_name]

        # Update title if provided
        if new_title is not None and new_title != sheet_name:
            ws.title = new_title

        # Update dimensions if provided
        if row_count is not None or column_count is not None:
            current_max_row = ws.max_row
            current_max_col = ws.max_column

            # Expand rows if needed
            if row_count is not None and row_count > current_max_row:
                ws.cell(row=row_count, column=1, value="")

            # Expand columns if needed
            if column_count is not None and column_count > current_max_col:
                ws.cell(row=1, column=column_count, value="")

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to update sheet properties: {e}")


def merge_excel_files(file_paths: list[str], output_path: str) -> str:
    """
    Combines multiple Excel files into one.

    Args:
        file_paths: List of Excel file paths to merge
        output_path: Output file path

    Returns:
        str: Output file path

    Raises:
        ValueError: If no files provided
        FileNotFoundError: If any input file is missing
        FileOperationError: If operation fails
    """
    try:
        if not file_paths:
            raise ValueError("No file paths provided")

        output_file = Path(output_path)

        # Create new workbook
        merged_wb = Workbook()
        # Remove default sheet
        if "Sheet" in merged_wb.sheetnames:
            merged_wb.remove(merged_wb["Sheet"])

        for file_path in file_paths:
            file_obj = Path(file_path)
            if not file_obj.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(file_path)
            file_prefix = file_obj.stem

            for sheet_name in wb.sheetnames:
                # Create unique sheet name
                new_sheet_name = f"{file_prefix}_{sheet_name}"
                counter = 1
                while new_sheet_name in merged_wb.sheetnames:
                    new_sheet_name = f"{file_prefix}_{sheet_name}_{counter}"
                    counter += 1

                # Copy sheet data
                source_sheet = wb[sheet_name]
                new_sheet = merged_wb.create_sheet(title=new_sheet_name)

                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_sheet[cell.coordinate].value = cell.value

        merged_wb.save(output_path)
        return str(output_file.absolute())

    except (ValueError, FileNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to merge Excel files: {e}")


def extract_sheet_to_csv(excel_path: str, sheet_name: str, csv_path: str) -> str:
    """
    Exports a sheet to CSV file.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to export
        csv_path: Output CSV file path

    Returns:
        str: CSV file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        PermissionError: If can't write CSV
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        csv_file = Path(csv_path)
        csv_file.parent.mkdir(parents=True, exist_ok=True)

        ws = wb[sheet_name]

        with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in ws.iter_rows(values_only=True):
                # Convert None values to empty strings
                cleaned_row = ["" if cell is None else cell for cell in row]
                writer.writerow(cleaned_row)

        return str(csv_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except PermissionError as e:
        raise PermissionError(f"Permission denied writing CSV {csv_path}: {e}")
    except Exception as e:
        raise FileOperationError(f"Failed to extract sheet to CSV: {e}")


def update_sheet_dimensions(
    excel_path: str, sheet_name: str, row_count: int | None = None, column_count: int | None = None
) -> str:
    """
    Updates the dimensions (row and column count) of an existing sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to update
        row_count: New number of rows (optional)
        column_count: New number of columns (optional)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If neither row_count nor column_count provided
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        if row_count is None and column_count is None:
            raise ValueError("At least one of row_count or column_count must be provided")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in {excel_path}")

        ws = wb[sheet_name]

        # Get current dimensions
        current_max_row = ws.max_row
        current_max_col = ws.max_column

        # Expand rows if needed
        if row_count is not None and row_count > current_max_row:
            # Add empty rows by writing to the last cell in the new range
            ws.cell(row=row_count, column=1, value="")

        # Expand columns if needed
        if column_count is not None and column_count > current_max_col:
            # Add empty columns by writing to the last cell in the new range
            ws.cell(row=1, column=column_count, value="")

        # Note: OpenPyXL doesn't support shrinking sheets directly
        # The sheet will maintain its current size if the new dimensions are smaller

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to update sheet dimensions: {e}")


def get_spreadsheet_metadata(excel_path: str) -> dict[str, Any]:
    """
    Gets comprehensive metadata about an Excel spreadsheet and all its sheets.

    Args:
        excel_path: Path to the Excel file

    Returns:
        Dict containing spreadsheet and sheet metadata

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If invalid Excel file
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path, read_only=True)

        # Get file stats
        file_stats = excel_file.stat()

        # Build sheet information
        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            # Count non-empty cells (approximate)
            used_cells = 0
            if max_row and max_col:
                # Sample approach to avoid loading entire sheet
                sample_size = min(100, max_row)
                for row_num in range(1, sample_size + 1):
                    for col_num in range(1, min(26, max_col) + 1):  # Check first 26 columns
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            used_cells += 1

            sheet_info = {
                "sheet_name": sheet_name,
                "max_row": max_row or 0,
                "max_column": max_col or 0,
                "estimated_used_cells": used_cells,
                "is_active": sheet_name == wb.active.title if wb.active else False,
            }
            sheets_info.append(sheet_info)

        wb.close()

        metadata = {
            "file_path": str(excel_file.absolute()),
            "file_name": excel_file.name,
            "file_size_bytes": file_stats.st_size,
            "created_time": file_stats.st_ctime,
            "modified_time": file_stats.st_mtime,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_info,
            "active_sheet": wb.active.title if wb.active else None,
        }

        return metadata

    except FileNotFoundError:
        raise
    except Exception as e:
        # Handle all zip/file format related errors
        if "zip" in str(e).lower() or "bad zip" in str(e).lower():
            raise ValueError(f"Invalid Excel file format: {excel_path}")
        raise FileOperationError(f"Failed to get spreadsheet metadata: {e}")


def duplicate_sheet_to_file(
    source_excel_path: str, source_sheet_name: str, target_excel_path: str, target_sheet_name: str | None = None
) -> str:
    """
    Copies a sheet from one Excel file to another Excel file.

    Args:
        source_excel_path: Path to the source Excel file
        source_sheet_name: Name of the sheet to copy
        target_excel_path: Path to the target Excel file (created if doesn't exist)
        target_sheet_name: Name for the copied sheet (defaults to source name)

    Returns:
        str: Target Excel file path

    Raises:
        FileNotFoundError: If source Excel file doesn't exist
        SheetNotFoundError: If source sheet doesn't exist
        ValueError: If target sheet name exists or is invalid
        FileOperationError: If operation fails
    """
    try:
        source_file = Path(source_excel_path)
        if not source_file.exists():
            raise FileNotFoundError(f"Source Excel file not found: {source_excel_path}")

        target_file = Path(target_excel_path)

        # Use source sheet name if target name not provided
        if target_sheet_name is None:
            target_sheet_name = source_sheet_name

        # Load source workbook
        source_wb = load_workbook(source_excel_path)

        if source_sheet_name not in source_wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{source_sheet_name}' not found in source file")

        source_ws = source_wb[source_sheet_name]

        # Load or create target workbook
        if target_file.exists():
            target_wb = load_workbook(target_excel_path)

            if target_sheet_name in target_wb.sheetnames:
                raise ValueError(f"Sheet '{target_sheet_name}' already exists in target file")
        else:
            target_wb = Workbook()
            # Remove default sheet if we're creating a new workbook
            if "Sheet" in target_wb.sheetnames:
                target_wb.remove(target_wb["Sheet"])

        # Create new sheet in target workbook
        target_ws = target_wb.create_sheet(title=target_sheet_name)

        # Copy all data from source to target
        for row in source_ws.iter_rows():
            target_row = []
            for cell in row:
                target_row.append(cell.value)
            target_ws.append(target_row)

        # Copy column widths
        for col_letter, col_dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_letter] = col_dim

        # Copy row heights
        for row_num, row_dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_num] = row_dim

        # Save target workbook
        target_wb.save(target_excel_path)

        return str(target_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to duplicate sheet to file: {e}")


def get_sheet_info(excel_path: str, sheet_name: str) -> dict[str, Any]:
    """
    Gets detailed information about a specific sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet

    Returns:
        Dict containing sheet information

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path, read_only=True)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in {excel_path}")

        ws = wb[sheet_name]

        # Get sheet dimensions and data info
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Count non-empty cells (sample approach)
        used_cells = 0
        if max_row and max_col:
            for row in ws.iter_rows(max_row=min(100, max_row), max_col=min(26, max_col)):
                for cell in row:
                    if cell.value is not None:
                        used_cells += 1

        # Check if sheet has data
        has_data = max_row > 0 and max_col > 0

        wb.close()

        sheet_info = {
            "sheet_name": sheet_name,
            "max_row": max_row,
            "max_column": max_col,
            "estimated_used_cells": used_cells,
            "has_data": has_data,
            "is_active": sheet_name == wb.active.title if wb.active else False,
            "sheet_index": wb.sheetnames.index(sheet_name),
        }

        return sheet_info

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to get sheet info: {e}")


# Pivot Table Functions
def create_pivot_table(
    excel_path: str,
    source_sheet: str,
    pivot_sheet: str,
    data_range: str,
    rows: list[str] | None = None,
    columns: list[str] | None = None,
    values: dict[str, str] | None = None,
    filters: list[str] | None = None,
    pivot_location: str = "A1",
    table_name: str | None = None,
    show_grand_totals_rows: bool = True,
    show_grand_totals_cols: bool = True,
    compact_layout: bool = True,
    outline_layout: bool = False,
    show_headers: bool = True,
    auto_format: bool = False,
    preserve_formatting: bool = True,
    enable_drill: bool = True,
    show_field_list: bool = True,
    pivot_table_style: str | None = None,
    advanced_config: dict[str, Any] | None = None,
) -> str:
    """
    Creates a fully featured pivot table from source data with basic or advanced configuration.

    Args:
        excel_path: Path to the Excel file
        source_sheet: Name of the source data sheet
        pivot_sheet: Name of the sheet for the pivot table
        data_range: Range of source data (e.g., 'A1:D100')

        # Basic configuration (used when advanced_config is None)
        rows: List of field names for row area (optional)
        columns: List of field names for column area (optional)
        values: Dict mapping field names to aggregation types (optional)
                {'field_name': 'sum'|'count'|'countNums'|'average'|'max'|'min'|'product'|'stdDev'|'stdDevp'|'var'|'varp'}
        filters: List of field names for filter/page area (optional)
        pivot_location: Cell reference where pivot table starts (default: 'A1')
        table_name: Custom name for the pivot table (optional)
        show_grand_totals_rows: Show grand totals for rows (default: True)
        show_grand_totals_cols: Show grand totals for columns (default: True)
        compact_layout: Use compact layout (default: True)
        outline_layout: Use outline layout (default: False)
        show_headers: Show field headers (default: True)
        auto_format: Use auto formatting (default: False)
        preserve_formatting: Preserve cell formatting (default: True)
        enable_drill: Enable drill-down functionality (default: True)
        show_field_list: Show field list (default: True)
        pivot_table_style: Name of pivot table style to apply (optional)

        # Advanced configuration (overrides basic parameters when provided)
        advanced_config: Advanced configuration dict for complex pivot tables (optional)
                        When provided, overrides basic parameters with detailed configuration:
                        {
                            "table_name": str (optional),
                            "location": str (default: "A1"),
                            "rows": [
                                {
                                    "field": str,
                                    "subtotals": bool (default: True),
                                    "sort_order": "asc"|"desc"|None,
                                    "show_all_items": bool (default: True)
                                }
                            ],
                            "columns": [
                                {
                                    "field": str,
                                    "subtotals": bool (default: True),
                                    "sort_order": "asc"|"desc"|None
                                }
                            ],
                            "values": [
                                {
                                    "field": str,
                                    "aggregation": str,
                                    "name": str (optional),
                                    "number_format": str (optional),
                                    "show_as": str (optional) - "normal"|"percent"|"difference"|etc.
                                }
                            ],
                            "filters": [str],
                            "layout": {
                                "compact": bool (default: True),
                                "outline": bool (default: False),
                                "tabular": bool (default: False),
                                "repeat_labels": bool (default: False),
                                "blank_rows": bool (default: False)
                            },
                            "totals": {
                                "grand_totals_rows": bool (default: True),
                                "grand_totals_cols": bool (default: True),
                                "subtotals_top": bool (default: False)
                            },
                            "display": {
                                "show_headers": bool (default: True),
                                "show_field_list": bool (default: True),
                                "show_tooltips": bool (default: True),
                                "enable_drill": bool (default: True),
                                "show_empty_rows": bool (default: False),
                                "show_empty_cols": bool (default: False)
                            },
                            "formatting": {
                                "auto_format": bool (default: False),
                                "preserve_formatting": bool (default: True),
                                "style_name": str (optional),
                                "apply_number_formats": bool (default: False),
                                "apply_border_formats": bool (default: False),
                                "apply_font_formats": bool (default: False),
                                "apply_pattern_formats": bool (default: False),
                                "apply_alignment_formats": bool (default: False)
                            }
                        }

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If source sheet doesn't exist
        ValueError: If invalid parameters provided
        FileOperationError: If operation fails

    Examples:
        # Basic usage
        create_pivot_table(
            "file.xlsx", "Data", "Pivot", "A1:D100",
            rows=["Category"], columns=["Month"], values={"Sales": "sum"}
        )

        # Advanced usage with config dict
        create_pivot_table(
            "file.xlsx", "Data", "Pivot", "A1:D100",
            advanced_config={
                "table_name": "SalesAnalysis",
                "location": "B2",
                "rows": [
                    {"field": "Category", "subtotals": True, "sort_order": "asc"},
                    {"field": "Subcategory", "subtotals": False}
                ],
                "columns": [{"field": "Quarter", "subtotals": True}],
                "values": [
                    {
                        "field": "Sales",
                        "aggregation": "sum",
                        "name": "Total Sales",
                        "number_format": "$#,##0.00"
                    }
                ],
                "layout": {"compact": False, "outline": True},
                "formatting": {"style_name": "PivotStyleMedium2"}
            }
        )
    """
    try:
        from openpyxl.pivot.table import (
            DataField,
            FieldItem,
            Location,
            PageField,
            PivotField,
            PivotTableStyle,
            RowColField,
            TableDefinition,
        )
        from openpyxl.utils import range_boundaries

        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if source_sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Source sheet '{source_sheet}' not found")

        # Process advanced configuration if provided
        if advanced_config:
            # Extract configuration with defaults
            table_name = advanced_config.get("table_name", table_name)
            pivot_location = advanced_config.get("location", pivot_location)

            rows_config = advanced_config.get("rows", [])
            columns_config = advanced_config.get("columns", [])
            values_config = advanced_config.get("values", [])
            filters_config = advanced_config.get("filters", [])

            layout = advanced_config.get("layout", {})
            totals = advanced_config.get("totals", {})
            display = advanced_config.get("display", {})
            formatting = advanced_config.get("formatting", {})

            # Convert to simple format for processing
            rows = [row["field"] if isinstance(row, dict) else row for row in rows_config]
            columns = [col["field"] if isinstance(col, dict) else col for col in columns_config]
            filters = filters_config

            values = {}
            for value_config in values_config:
                if isinstance(value_config, dict):
                    field = value_config["field"]
                    aggregation = value_config["aggregation"]
                    values[field] = aggregation
                else:
                    # Fallback for simple string format
                    values[value_config] = "sum"

            # Override basic parameters with advanced config
            show_grand_totals_rows = totals.get("grand_totals_rows", show_grand_totals_rows)
            show_grand_totals_cols = totals.get("grand_totals_cols", show_grand_totals_cols)
            compact_layout = layout.get("compact", compact_layout)
            outline_layout = layout.get("outline", outline_layout)
            show_headers = display.get("show_headers", show_headers)
            auto_format = formatting.get("auto_format", auto_format)
            preserve_formatting = formatting.get("preserve_formatting", preserve_formatting)
            enable_drill = display.get("enable_drill", enable_drill)
            show_field_list = display.get("show_field_list", show_field_list)
            pivot_table_style = formatting.get("style_name", pivot_table_style)

        # Get source worksheet to analyze field names
        source_ws = wb[source_sheet]

        # Parse data range to get field names from header row
        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # Extract field names from header row
        field_names = []
        for col in range(min_col, max_col + 1):
            cell_value = source_ws.cell(row=min_row, column=col).value
            if cell_value:
                field_names.append(str(cell_value))

        if not field_names:
            raise ValueError("No field names found in header row")

        # Validate field names in parameters
        all_specified_fields = set()
        if rows:
            all_specified_fields.update(rows)
        if columns:
            all_specified_fields.update(columns)
        if values:
            all_specified_fields.update(values.keys())
        if filters:
            all_specified_fields.update(filters)

        invalid_fields = all_specified_fields - set(field_names)
        if invalid_fields:
            raise ValueError(f"Invalid field names: {invalid_fields}. Available fields: {field_names}")

        # Create pivot sheet if it doesn't exist
        if pivot_sheet not in wb.sheetnames:
            wb.create_sheet(title=pivot_sheet)

        pivot_ws = wb[pivot_sheet]

        # Create pivot cache
        from openpyxl.pivot.cache import CacheSource

        worksheet_source = WorksheetSource(ref=data_range, sheet=source_sheet)
        cache_source = CacheSource(worksheetSource=worksheet_source)
        cache = CacheDefinition(cacheSource=cache_source)

        # Generate unique cache ID
        cache_id = 1

        # Set table name
        if table_name is None:
            table_name = f"PivotTable_{pivot_sheet}"

        # Create pivot table definition
        pivot_table = TableDefinition(
            name=table_name,
            cacheId=cache_id,
            compact=compact_layout,
            outline=outline_layout,
            showHeaders=show_headers,
            useAutoFormatting=auto_format,
            preserveFormatting=preserve_formatting,
            enableDrill=enable_drill,
            disableFieldList=not show_field_list,
            rowGrandTotals=show_grand_totals_rows,
            colGrandTotals=show_grand_totals_cols,
        )

        # Set pivot table cache
        pivot_table.cache = cache

        # Create location for pivot table
        location = Location(ref=pivot_location, firstHeaderRow=1, firstDataRow=2, firstDataCol=1)
        pivot_table.location = location

        # Create pivot fields for all available fields
        pivot_fields = []
        for i, field_name in enumerate(field_names):
            # Determine field axis based on parameters
            axis = None
            if rows and field_name in rows:
                axis = "axisRow"
            elif columns and field_name in columns:
                axis = "axisCol"
            elif filters and field_name in filters:
                axis = "axisPage"
            elif values and field_name in values:
                axis = "axisValues"

            # Create field items (simplified - just default item)
            field_items = [FieldItem(t="default")]

            pivot_field = PivotField(
                name=field_name, axis=axis, items=field_items, compact=compact_layout, outline=outline_layout
            )
            pivot_fields.append(pivot_field)

        pivot_table.pivotFields = pivot_fields

        # Add row fields
        if rows:
            row_fields = []
            for field_name in rows:
                field_index = field_names.index(field_name)
                row_fields.append(RowColField(x=field_index))
            pivot_table.rowFields = row_fields

        # Add column fields
        if columns:
            col_fields = []
            for field_name in columns:
                field_index = field_names.index(field_name)
                col_fields.append(RowColField(x=field_index))
            pivot_table.colFields = col_fields

        # Add page/filter fields
        if filters:
            page_fields = []
            for field_name in filters:
                field_index = field_names.index(field_name)
                page_field = PageField(fld=field_index, name=field_name)
                page_fields.append(page_field)
            pivot_table.pageFields = page_fields

        # Add data fields (values)
        if values:
            data_fields = []
            valid_aggregations = {
                "sum",
                "count",
                "countNums",
                "average",
                "max",
                "min",
                "product",
                "stdDev",
                "stdDevp",
                "var",
                "varp",
            }

            for field_name, aggregation in values.items():
                if aggregation not in valid_aggregations:
                    raise ValueError(
                        f"Invalid aggregation type '{aggregation}' for field '{field_name}'. Valid types: {sorted(valid_aggregations)}"
                    )

                field_index = field_names.index(field_name)

                # Map aggregation names to openpyxl format
                aggregation_map = {"average": "avg", "stdDev": "stdDev", "stdDevp": "stdDevP", "varp": "varP"}
                openpyxl_aggregation = aggregation_map.get(aggregation, aggregation)

                data_field = DataField(
                    name=f"{aggregation.title()} of {field_name}", fld=field_index, subtotal=openpyxl_aggregation
                )
                data_fields.append(data_field)

            pivot_table.dataFields = data_fields

        # Apply pivot table style if specified
        if pivot_table_style:
            style_info = PivotTableStyle(
                name=pivot_table_style,
                showRowHeaders=True,
                showColHeaders=True,
                showRowStripes=False,
                showColStripes=False,
                showLastColumn=False,
            )
            pivot_table.pivotTableStyleInfo = style_info

        # Add pivot table to worksheet
        pivot_ws.add_pivot(pivot_table)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create pivot table: {e}")


# Utility Functions
def write_value_to_cell(excel_path: str, sheet_name: str, cell: str, value: Any) -> str:
    """
    Writes a value to a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        value: Value to write

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write value to cell: {e}")


def get_cell_value(excel_path: str, sheet_name: str, cell: str) -> Any:
    """
    Gets the value from a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')

    Returns:
        Any: Cell value
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        return ws[cell].value

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to get cell value: {e}")


def clear_sheet(excel_path: str, sheet_name: str) -> str:
    """
    Clears all data from a sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to clear

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        # Remove and recreate sheet to clear all data
        wb.remove(wb[sheet_name])
        wb.create_sheet(title=sheet_name)
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to clear sheet: {e}")


# Advanced Data Processing Functions
def create_autofilter(excel_path: str, sheet_name: str, data_range: str) -> str:
    """
    Adds autofilter to a data range.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data_range: Range to apply autofilter (e.g., 'A1:D100')

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws.auto_filter.ref = data_range
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create autofilter: {e}")


def add_subtotals(
    excel_path: str,
    sheet_name: str,
    data_range: str,
    group_column: int,
    summary_columns: list[int],
    function_type: str = "sum",
) -> str:
    """
    Adds subtotals to grouped data.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data_range: Range containing the data
        group_column: Column number (1-based) to group by
        summary_columns: List of column numbers to summarize
        function_type: Type of summary function ('sum', 'average', 'count', 'max', 'min')

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If invalid parameters provided
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Parse data range
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # Read data to identify groups
        data = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
        ):
            data.append(list(row))

        if not data:
            raise ValueError("No data found in specified range")

        # Group data by the specified column
        groups = {}
        for row_idx, row in enumerate(data[1:], 1):  # Skip header
            group_key = row[group_column - min_col]
            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append((row_idx + min_row, row))

        # Insert subtotal rows
        offset = 0
        for group_key, group_rows in groups.items():
            last_row_idx = group_rows[-1][0] + offset

            # Insert subtotal row
            ws.insert_rows(last_row_idx + 1)
            offset += 1

            # Add subtotal formulas
            subtotal_row = last_row_idx + 1

            # Add group label
            ws.cell(row=subtotal_row, column=group_column, value=f"{group_key} Total")

            # Add subtotal formulas for each summary column
            for col_num in summary_columns:
                start_row = group_rows[0][0] + offset - 1
                end_row = last_row_idx

                col_letter = ws.cell(row=1, column=col_num).column_letter

                valid_function_types = {"sum", "average", "count", "max", "min"}
                function_lower = function_type.lower()

                if function_lower not in valid_function_types:
                    raise ValueError(
                        f"Invalid function_type: {function_type}. Valid types: {sorted(valid_function_types)}"
                    )

                if function_lower == "sum":
                    formula = f"=SUBTOTAL(9,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_lower == "average":
                    formula = f"=SUBTOTAL(1,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_lower == "count":
                    formula = f"=SUBTOTAL(3,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_lower == "max":
                    formula = f"=SUBTOTAL(4,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_lower == "min":
                    formula = f"=SUBTOTAL(5,{col_letter}{start_row}:{col_letter}{end_row})"

                ws.cell(row=subtotal_row, column=col_num, value=formula)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to add subtotals: {e}")


def create_data_table(
    excel_path: str,
    sheet_name: str,
    data_range: str,
    table_name: str,
    table_style: str = "TableStyleMedium9",
) -> str:
    """
    Create a formatted data table from a range.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data_range: Range to convert to table
        table_name: Name for the table
        table_style: Excel table style name

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If data range is invalid
        FileOperationError: If operation fails
    """
    try:
        from openpyxl.utils import range_boundaries
        from openpyxl.worksheet.table import Table, TableStyleInfo

        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Validate range
        try:
            range_boundaries(data_range)  # Just validate, don't need the values
        except ValueError as e:
            raise ValueError(f"Invalid range specification '{data_range}': {e}")

        # Create table
        table = Table(displayName=table_name, ref=data_range)

        # Add table style
        style = TableStyleInfo(
            name=table_style,
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=True,
        )
        table.tableStyleInfo = style

        # Add table to worksheet
        ws.add_table(table)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create data table: {e}")


# Chart creation functions have been moved to excel_charts_toolset.py
# to avoid duplication and maintain clear separation of concerns.


if __name__ == "__main__":
    # Example usage
    try:
        # Create a new Excel file
        file_path = create_excel_file("example.xlsx")
        print(f"Created Excel file: {file_path}")

        # Add some data
        data = [["Name", "Age", "Salary"], ["John", 30, 50000], ["Jane", 25, 45000], ["Bob", 35, 60000]]
        write_data_to_sheet(file_path, "Sheet", data)

        # Add advanced features
        create_autofilter(file_path, "Sheet", "A1:C4")

        print("Excel file created with data and advanced features!")

    except Exception as e:
        print(f"Error: {e}")
