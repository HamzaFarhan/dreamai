#!/usr/bin/env python3
"""
Simple demonstration of how to use the new formula error handling for AI agents.
"""

import os

from openpyxl import Workbook

from src.dreamai.toolsets.excel_formula_toolset import (
    write_and_evaluate_formula,
    write_formula_with_error_handling,
)


def create_sample_data():
    """Create sample Excel file with data that will cause division by zero."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Raw_Orders"

    # Headers
    ws["A1"] = "Customer"
    ws["B1"] = "Plan"
    ws["C1"] = "Amount"
    ws["D1"] = "Date"
    ws["E1"] = "Active"  # This column will have 0 counts for certain criteria

    # Sample data
    data = [
        ["Customer1", "Pro", 100, "2023-06-01", 1],
        ["Customer2", "Basic", 50, "2023-07-01", 1],
        ["Customer3", "Pro", 150, "2023-08-01", 1],
        ["Customer4", "Enterprise", 500, "2023-09-01", 0],  # No active enterprise customers
    ]

    for i, row in enumerate(data, 2):
        for j, value in enumerate(row, 1):
            ws.cell(row=i, column=j, value=value)

    filename = "sample_orders.xlsx"
    wb.save(filename)
    wb.close()
    return filename


def ai_agent_formula_workflow():
    """Demonstrate how an AI agent would iteratively build formulas."""
    print("🤖 AI Agent: Starting formula building workflow...")

    # Create test data
    excel_file = create_sample_data()
    print(f"📊 Created sample data in: {excel_file}")

    # Target: Calculate average amount for Pro customers divided by count of active Enterprise customers
    # This will cause division by zero since there are no active Enterprise customers

    print("\n--- AI Agent Formula Building ---")

    # Step 1: AI tries initial formula
    initial_formula = 'AVERAGEIFS(C:C,B:B,"Pro")/COUNTIFS(B:B,"Enterprise",E:E,1)'
    print(f"\n🤖 AI Agent: Trying formula: {initial_formula}")

    result = write_and_evaluate_formula(excel_file, "Raw_Orders", "F1", initial_formula)

    if result["success"]:
        print(f"✅ Success! Value: {result['value']}")
    else:
        print(f"❌ Error detected: {result['error']} - {result['error_message']}")

        # Step 2: AI agent decides to fix the error
        print("\n🤖 AI Agent: Detected division by zero. Let me fix this...")

        # AI tries different approaches
        fixed_formulas = [
            'IF(COUNTIFS(B:B,"Enterprise",E:E,1)=0,0,AVERAGEIFS(C:C,B:B,"Pro")/COUNTIFS(B:B,"Enterprise",E:E,1))',
            'IFERROR(AVERAGEIFS(C:C,B:B,"Pro")/COUNTIFS(B:B,"Enterprise",E:E,1),"No Enterprise customers")',
            'AVERAGEIFS(C:C,B:B,"Pro")',  # Fallback to simpler calculation
        ]

        for i, formula in enumerate(fixed_formulas, 1):
            print(f"\n🤖 AI Agent: Attempt {i}: {formula}")

            result = write_and_evaluate_formula(excel_file, "Raw_Orders", f"F{i + 1}", formula)

            if result["success"]:
                print(f"✅ Success! Value: {result['value']}")
                print("🎯 AI Agent: Found working formula! Saving as final result.")

                # Write the working formula to the main result cell
                write_and_evaluate_formula(excel_file, "Raw_Orders", "G1", formula)
                break
            else:
                print(f"❌ Still has error: {result['error']} - {result['error_message']}")

    print(f"\n📁 Final Excel file: {excel_file}")
    print("🔍 You can open this file to see the formulas that were written and their results.")

    # Show what's in the file
    from openpyxl import load_workbook

    wb = load_workbook(excel_file, data_only=False)
    ws = wb["Raw_Orders"]

    print("\n📋 Formulas written to Excel:")
    for cell in ["F1", "F2", "F3", "F4", "G1"]:
        if ws[cell].value:
            print(f"   {cell}: {ws[cell].value}")

    wb.close()

    # Cleanup option
    response = input("\n🗑️  Delete test file? (y/n): ")
    if response.lower() == "y":
        os.remove(excel_file)
        print("🗑️  Test file deleted.")
    else:
        print(f"📁 Test file saved as: {excel_file}")


def demonstrate_automatic_error_handling():
    """Show the automatic error handling function."""
    print("\n" + "=" * 60)
    print("🤖 Demonstrating Automatic Error Handling")
    print("=" * 60)

    excel_file = create_sample_data()

    # This will automatically try to fix the division by zero error
    result = write_formula_with_error_handling(
        excel_file,
        "Raw_Orders",
        "H1",
        'AVERAGEIFS(C:C,B:B,"Pro")/COUNTIFS(B:B,"Enterprise",E:E,1)',
        max_retries=3,
        error_fallback="0",
    )

    print("\n📊 Automatic error handling result:")
    print(f"   Success: {result['success']}")
    print(f"   Final value: {result['value']}")
    print(f"   Used fallback: {result.get('used_fallback', False)}")
    print(f"   Number of attempts: {len(result.get('attempts', []))}")

    if "attempts" in result:
        print("\n🔄 Attempts made:")
        for attempt in result["attempts"]:
            print(f"   {attempt['attempt']}: {attempt['formula']}")
            print(f"      → {attempt['result']['success']} ({attempt['result'].get('error', 'OK')})")

    os.remove(excel_file)


if __name__ == "__main__":
    ai_agent_formula_workflow()
    demonstrate_automatic_error_handling()
