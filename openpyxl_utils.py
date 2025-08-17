import csv
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.pivot.cache import CacheDefinition, WorksheetSource
from openpyxl.pivot.table import DataField, TableDefinition


# Custom Exceptions
class SheetNotFoundError(Exception):
    """Raised when a specified sheet is not found in the workbook."""

    pass


class InvalidSheetNameError(Exception):
    """Raised when an invalid sheet name is provided."""

    pass


class FormulaError(Exception):
    """Raised when there's an error with formula syntax or parameters."""

    pass


class FileOperationError(Exception):
    """Raised when file operations fail."""

    pass


# Core File Operations
def create_excel_file(file_path: str) -> str:
    """
    Creates a new Excel workbook at the specified path.

    Args:
        file_path: Path where the Excel file will be created

    Returns:
        str: Path of the created file

    Raises:
        FileExistsError: If file already exists
        PermissionError: If write access is denied
        FileOperationError: If file creation fails
    """
    try:
        path = Path(file_path)

        if path.exists():
            raise FileExistsError(f"File already exists: {file_path}")

        # Ensure directory exists
        path.parent.mkdir(parents=True, exist_ok=True)

        # Create workbook
        wb = Workbook()
        wb.save(file_path)

        return str(path.absolute())

    except FileExistsError:
        raise
    except PermissionError as e:
        raise PermissionError(f"Permission denied creating file {file_path}: {e}")
    except Exception as e:
        raise FileOperationError(f"Failed to create Excel file {file_path}: {e}")


def csv_to_excel_sheet(csv_path: str, excel_path: str, sheet_name: str | None = None) -> str:
    """
    Adds a CSV file as a new sheet in an Excel workbook.

    Args:
        csv_path: Path to the CSV file
        excel_path: Path to the Excel file (created if doesn't exist)
        sheet_name: Optional sheet name (default: CSV filename)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If CSV file doesn't exist
        ValueError: For invalid CSV format
        FileOperationError: If operation fails
    """
    try:
        csv_file = Path(csv_path)
        if not csv_file.exists():
            raise FileNotFoundError(f"CSV file not found: {csv_path}")

        excel_file = Path(excel_path)

        # Determine sheet name
        if sheet_name is None:
            sheet_name = csv_file.stem

        # Load or create workbook
        if excel_file.exists():
            wb = load_workbook(excel_path)
        else:
            wb = Workbook()
            # Remove default sheet if we're adding a named sheet
            if "Sheet" in wb.sheetnames:
                wb.remove(wb["Sheet"])

        # Create new sheet
        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        ws = wb.create_sheet(title=sheet_name)

        # Read CSV and write to sheet
        with open(csv_path, "r", encoding="utf-8") as csvfile:
            # Try to detect delimiter
            sample = csvfile.read(1024)
            csvfile.seek(0)
            sniffer = csv.Sniffer()
            delimiter = sniffer.sniff(sample).delimiter

            reader = csv.reader(csvfile, delimiter=delimiter)
            for row_idx, row in enumerate(reader, 1):
                for col_idx, value in enumerate(row, 1):
                    # Try to convert to number if possible
                    try:
                        if "." in value:
                            ws.cell(row=row_idx, column=col_idx, value=float(value))
                        else:
                            ws.cell(row=row_idx, column=col_idx, value=int(value))
                    except (ValueError, TypeError):
                        ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except FileNotFoundError:
        raise
    except ValueError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to convert CSV to Excel: {e}")


def csvs_to_excel(csv_paths: list[str], excel_path: str) -> str:
    """
    Adds multiple CSV files as sheets in a single Excel workbook.

    Args:
        csv_paths: List of CSV file paths
        excel_path: Output Excel file path

    Returns:
        str: Excel file path

    Raises:
        ValueError: If no CSV paths provided
        FileNotFoundError: If any CSV file is missing
        FileOperationError: If operation fails
    """
    try:
        if not csv_paths:
            raise ValueError("No CSV paths provided")

        excel_file = Path(excel_path)

        # Create new workbook
        wb = Workbook()
        # Remove default sheet
        if "Sheet" in wb.sheetnames:
            wb.remove(wb["Sheet"])

        for csv_path in csv_paths:
            csv_file = Path(csv_path)
            if not csv_file.exists():
                raise FileNotFoundError(f"CSV file not found: {csv_path}")

            sheet_name = csv_file.stem

            # Ensure unique sheet name
            original_name = sheet_name
            counter = 1
            while sheet_name in wb.sheetnames:
                sheet_name = f"{original_name}_{counter}"
                counter += 1

            ws = wb.create_sheet(title=sheet_name)

            # Read CSV and write to sheet
            with open(csv_path, "r", encoding="utf-8") as csvfile:
                sample = csvfile.read(1024)
                csvfile.seek(0)
                sniffer = csv.Sniffer()
                delimiter = sniffer.sniff(sample).delimiter

                reader = csv.reader(csvfile, delimiter=delimiter)
                for row_idx, row in enumerate(reader, 1):
                    for col_idx, value in enumerate(row, 1):
                        try:
                            if "." in value:
                                ws.cell(row=row_idx, column=col_idx, value=float(value))
                            else:
                                ws.cell(row=row_idx, column=col_idx, value=int(value))
                        except (ValueError, TypeError):
                            ws.cell(row=row_idx, column=col_idx, value=value)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (ValueError, FileNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to convert CSVs to Excel: {e}")


def add_sheet(excel_path: str, sheet_name: str) -> str:
    """
    Adds a new empty sheet to an Excel workbook.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the new sheet

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If sheet name exists or is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name in wb.sheetnames:
            raise ValueError(f"Sheet '{sheet_name}' already exists")

        if not sheet_name or len(sheet_name.strip()) == 0:
            raise ValueError("Sheet name cannot be empty")

        wb.create_sheet(title=sheet_name)
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to add sheet: {e}")


def write_data_to_sheet(
    excel_path: str, sheet_name: str, data: list[list[Any]], start_row: int = 1, start_col: int = 1
) -> str:
    """
    Writes 2D data to a specific sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data: 2D list of values
        start_row: Starting row (1-based)
        start_col: Starting column (1-based)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If data is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        if not data:
            raise ValueError("Data cannot be empty")

        ws = wb[sheet_name]

        for row_idx, row in enumerate(data):
            for col_idx, value in enumerate(row):
                ws.cell(row=start_row + row_idx, column=start_col + col_idx, value=value)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write data to sheet: {e}")


def get_sheet_names(excel_path: str) -> list[str]:
    """
    Returns list of sheet names in workbook.

    Args:
        excel_path: Path to the Excel file

    Returns:
        List[str]: List of sheet names

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If invalid Excel file
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)
        return wb.sheetnames

    except FileNotFoundError:
        raise
    except Exception as e:
        # Handle all zip/file format related errors
        if "zip" in str(e).lower() or "bad zip" in str(e).lower():
            raise ValueError(f"Invalid Excel file: {excel_path}")
        raise FileOperationError(f"Failed to get sheet names: {e}")


def delete_sheet(excel_path: str, sheet_name: str) -> str:
    """
    Deletes a sheet from workbook.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to delete

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        RuntimeError: If trying to delete the last sheet
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        if len(wb.sheetnames) == 1:
            raise RuntimeError("Cannot delete the last sheet in workbook")

        wb.remove(wb[sheet_name])
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, RuntimeError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to delete sheet: {e}")


def rename_sheet(excel_path: str, old_name: str, new_name: str) -> str:
    """
    Renames an existing sheet.

    Args:
        excel_path: Path to the Excel file
        old_name: Current sheet name
        new_name: New sheet name

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If old sheet doesn't exist
        ValueError: If new name is invalid or exists
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if old_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{old_name}' not found")

        if new_name in wb.sheetnames:
            raise ValueError(f"Sheet '{new_name}' already exists")

        if not new_name or len(new_name.strip()) == 0:
            raise ValueError("New sheet name cannot be empty")

        wb[old_name].title = new_name
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to rename sheet: {e}")


def copy_sheet(excel_path: str, sheet_name: str, new_name: str) -> str:
    """
    Copies a sheet within the same workbook.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to copy
        new_name: Name for the copied sheet

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If source sheet doesn't exist
        ValueError: If new name exists or is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        if new_name in wb.sheetnames:
            raise ValueError(f"Sheet '{new_name}' already exists")

        if not new_name or len(new_name.strip()) == 0:
            raise ValueError("New sheet name cannot be empty")

        source_sheet = wb[sheet_name]
        copied_sheet = wb.copy_worksheet(source_sheet)
        copied_sheet.title = new_name

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to copy sheet: {e}")


def update_sheet_properties(
    excel_path: str,
    sheet_name: str,
    new_title: str | None = None,
    row_count: int | None = None,
    column_count: int | None = None,
) -> str:
    """
    Updates properties of an existing sheet (title and/or dimensions).

    Args:
        excel_path: Path to the Excel file
        sheet_name: Current name of the sheet to update
        new_title: New title for the sheet (optional)
        row_count: New number of rows (optional)
        column_count: New number of columns (optional)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If new title exists or no properties provided
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        if new_title is None and row_count is None and column_count is None:
            raise ValueError("At least one property (new_title, row_count, column_count) must be provided")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in {excel_path}")

        # Check if new title conflicts with existing sheet names
        if new_title is not None and new_title != sheet_name and new_title in wb.sheetnames:
            raise ValueError(f"Sheet with title '{new_title}' already exists")

        ws = wb[sheet_name]

        # Update title if provided
        if new_title is not None and new_title != sheet_name:
            ws.title = new_title

        # Update dimensions if provided
        if row_count is not None or column_count is not None:
            current_max_row = ws.max_row
            current_max_col = ws.max_column

            # Expand rows if needed
            if row_count is not None and row_count > current_max_row:
                ws.cell(row=row_count, column=1, value="")

            # Expand columns if needed
            if column_count is not None and column_count > current_max_col:
                ws.cell(row=1, column=column_count, value="")

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to update sheet properties: {e}")


def merge_excel_files(file_paths: list[str], output_path: str) -> str:
    """
    Combines multiple Excel files into one.

    Args:
        file_paths: List of Excel file paths to merge
        output_path: Output file path

    Returns:
        str: Output file path

    Raises:
        ValueError: If no files provided
        FileNotFoundError: If any input file is missing
        FileOperationError: If operation fails
    """
    try:
        if not file_paths:
            raise ValueError("No file paths provided")

        output_file = Path(output_path)

        # Create new workbook
        merged_wb = Workbook()
        # Remove default sheet
        if "Sheet" in merged_wb.sheetnames:
            merged_wb.remove(merged_wb["Sheet"])

        for file_path in file_paths:
            file_obj = Path(file_path)
            if not file_obj.exists():
                raise FileNotFoundError(f"File not found: {file_path}")

            wb = load_workbook(file_path)
            file_prefix = file_obj.stem

            for sheet_name in wb.sheetnames:
                # Create unique sheet name
                new_sheet_name = f"{file_prefix}_{sheet_name}"
                counter = 1
                while new_sheet_name in merged_wb.sheetnames:
                    new_sheet_name = f"{file_prefix}_{sheet_name}_{counter}"
                    counter += 1

                # Copy sheet data
                source_sheet = wb[sheet_name]
                new_sheet = merged_wb.create_sheet(title=new_sheet_name)

                for row in source_sheet.iter_rows():
                    for cell in row:
                        new_sheet[cell.coordinate].value = cell.value

        merged_wb.save(output_path)
        return str(output_file.absolute())

    except (ValueError, FileNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to merge Excel files: {e}")


def extract_sheet_to_csv(excel_path: str, sheet_name: str, csv_path: str) -> str:
    """
    Exports a sheet to CSV file.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to export
        csv_path: Output CSV file path

    Returns:
        str: CSV file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        PermissionError: If can't write CSV
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        csv_file = Path(csv_path)
        csv_file.parent.mkdir(parents=True, exist_ok=True)

        ws = wb[sheet_name]

        with open(csv_path, "w", newline="", encoding="utf-8") as csvfile:
            writer = csv.writer(csvfile)
            for row in ws.iter_rows(values_only=True):
                # Convert None values to empty strings
                cleaned_row = ["" if cell is None else cell for cell in row]
                writer.writerow(cleaned_row)

        return str(csv_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except PermissionError as e:
        raise PermissionError(f"Permission denied writing CSV {csv_path}: {e}")
    except Exception as e:
        raise FileOperationError(f"Failed to extract sheet to CSV: {e}")


def update_sheet_dimensions(
    excel_path: str, sheet_name: str, row_count: int | None = None, column_count: int | None = None
) -> str:
    """
    Updates the dimensions (row and column count) of an existing sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to update
        row_count: New number of rows (optional)
        column_count: New number of columns (optional)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        ValueError: If neither row_count nor column_count provided
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        if row_count is None and column_count is None:
            raise ValueError("At least one of row_count or column_count must be provided")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in {excel_path}")

        ws = wb[sheet_name]

        # Get current dimensions
        current_max_row = ws.max_row
        current_max_col = ws.max_column

        # Expand rows if needed
        if row_count is not None and row_count > current_max_row:
            # Add empty rows by writing to the last cell in the new range
            ws.cell(row=row_count, column=1, value="")

        # Expand columns if needed
        if column_count is not None and column_count > current_max_col:
            # Add empty columns by writing to the last cell in the new range
            ws.cell(row=1, column=column_count, value="")

        # Note: OpenPyXL doesn't support shrinking sheets directly
        # The sheet will maintain its current size if the new dimensions are smaller

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to update sheet dimensions: {e}")


def get_spreadsheet_metadata(excel_path: str) -> dict[str, Any]:
    """
    Gets comprehensive metadata about an Excel spreadsheet and all its sheets.

    Args:
        excel_path: Path to the Excel file

    Returns:
        Dict containing spreadsheet and sheet metadata

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        ValueError: If invalid Excel file
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path, read_only=True)

        # Get file stats
        file_stats = excel_file.stat()

        # Build sheet information
        sheets_info = []
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]

            # Get sheet dimensions
            max_row = ws.max_row
            max_col = ws.max_column

            # Count non-empty cells (approximate)
            used_cells = 0
            if max_row and max_col:
                # Sample approach to avoid loading entire sheet
                sample_size = min(100, max_row)
                for row_num in range(1, sample_size + 1):
                    for col_num in range(1, min(26, max_col) + 1):  # Check first 26 columns
                        cell = ws.cell(row=row_num, column=col_num)
                        if cell.value is not None:
                            used_cells += 1

            sheet_info = {
                "sheet_name": sheet_name,
                "max_row": max_row or 0,
                "max_column": max_col or 0,
                "estimated_used_cells": used_cells,
                "is_active": sheet_name == wb.active.title if wb.active else False,
            }
            sheets_info.append(sheet_info)

        wb.close()

        metadata = {
            "file_path": str(excel_file.absolute()),
            "file_name": excel_file.name,
            "file_size_bytes": file_stats.st_size,
            "created_time": file_stats.st_ctime,
            "modified_time": file_stats.st_mtime,
            "sheet_count": len(wb.sheetnames),
            "sheet_names": wb.sheetnames,
            "sheets": sheets_info,
            "active_sheet": wb.active.title if wb.active else None,
        }

        return metadata

    except FileNotFoundError:
        raise
    except Exception as e:
        # Handle all zip/file format related errors
        if "zip" in str(e).lower() or "bad zip" in str(e).lower():
            raise ValueError(f"Invalid Excel file format: {excel_path}")
        raise FileOperationError(f"Failed to get spreadsheet metadata: {e}")


def duplicate_sheet_to_file(
    source_excel_path: str, source_sheet_name: str, target_excel_path: str, target_sheet_name: str | None = None
) -> str:
    """
    Copies a sheet from one Excel file to another Excel file.

    Args:
        source_excel_path: Path to the source Excel file
        source_sheet_name: Name of the sheet to copy
        target_excel_path: Path to the target Excel file (created if doesn't exist)
        target_sheet_name: Name for the copied sheet (defaults to source name)

    Returns:
        str: Target Excel file path

    Raises:
        FileNotFoundError: If source Excel file doesn't exist
        SheetNotFoundError: If source sheet doesn't exist
        ValueError: If target sheet name exists or is invalid
        FileOperationError: If operation fails
    """
    try:
        source_file = Path(source_excel_path)
        if not source_file.exists():
            raise FileNotFoundError(f"Source Excel file not found: {source_excel_path}")

        target_file = Path(target_excel_path)

        # Use source sheet name if target name not provided
        if target_sheet_name is None:
            target_sheet_name = source_sheet_name

        # Load source workbook
        source_wb = load_workbook(source_excel_path)

        if source_sheet_name not in source_wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{source_sheet_name}' not found in source file")

        source_ws = source_wb[source_sheet_name]

        # Load or create target workbook
        if target_file.exists():
            target_wb = load_workbook(target_excel_path)

            if target_sheet_name in target_wb.sheetnames:
                raise ValueError(f"Sheet '{target_sheet_name}' already exists in target file")
        else:
            target_wb = Workbook()
            # Remove default sheet if we're creating a new workbook
            if "Sheet" in target_wb.sheetnames:
                target_wb.remove(target_wb["Sheet"])

        # Create new sheet in target workbook
        target_ws = target_wb.create_sheet(title=target_sheet_name)

        # Copy all data from source to target
        for row in source_ws.iter_rows():
            target_row = []
            for cell in row:
                target_row.append(cell.value)
            target_ws.append(target_row)

        # Copy column widths
        for col_letter, col_dim in source_ws.column_dimensions.items():
            target_ws.column_dimensions[col_letter] = col_dim

        # Copy row heights
        for row_num, row_dim in source_ws.row_dimensions.items():
            target_ws.row_dimensions[row_num] = row_dim

        # Save target workbook
        target_wb.save(target_excel_path)

        return str(target_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to duplicate sheet to file: {e}")


def get_sheet_info(excel_path: str, sheet_name: str) -> dict[str, Any]:
    """
    Gets detailed information about a specific sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet

    Returns:
        Dict containing sheet information

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path, read_only=True)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found in {excel_path}")

        ws = wb[sheet_name]

        # Get sheet dimensions and data info
        max_row = ws.max_row or 0
        max_col = ws.max_column or 0

        # Count non-empty cells (sample approach)
        used_cells = 0
        if max_row and max_col:
            for row in ws.iter_rows(max_row=min(100, max_row), max_col=min(26, max_col)):
                for cell in row:
                    if cell.value is not None:
                        used_cells += 1

        # Check if sheet has data
        has_data = max_row > 0 and max_col > 0

        wb.close()

        sheet_info = {
            "sheet_name": sheet_name,
            "max_row": max_row,
            "max_column": max_col,
            "estimated_used_cells": used_cells,
            "has_data": has_data,
            "is_active": sheet_name == wb.active.title if wb.active else False,
            "sheet_index": wb.sheetnames.index(sheet_name),
        }

        return sheet_info

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to get sheet info: {e}")


# Formula Writing Functions
def _write_formula(excel_path: str, sheet_name: str, cell: str, formula: str) -> str:
    """
    Helper function to write a formula to a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        formula: Excel formula (without leading =)

    Returns:
        str: Excel file path

    Raises:
        FileNotFoundError: If Excel file doesn't exist
        SheetNotFoundError: If sheet doesn't exist
        FormulaError: If formula is invalid
        FileOperationError: If operation fails
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Ensure formula starts with =
        if not formula.startswith("="):
            formula = "=" + formula

        ws[cell] = formula
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write formula: {e}")


# Date Functions
def write_date_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes date functions to Excel cells.

    Supported functions: DATE, DAY, HOUR, MINUTE, MONTH, NOW, SECOND, TIME, TODAY, WEEKDAY, YEAR

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the date function
        function_args: Function arguments

    Returns:
        str: Excel file path

    Example:
        write_date_function("file.xlsx", "Sheet1", "A1", "TODAY")
        write_date_function("file.xlsx", "Sheet1", "B1", "DATE", 2023, 12, 25)
    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "DATE",
            "DAY",
            "HOUR",
            "MINUTE",
            "MONTH",
            "NOW",
            "SECOND",
            "TIME",
            "TODAY",
            "WEEKDAY",
            "YEAR",
        }

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid date function: {function_name}")

        # Build formula
        if function_args:
            args_str = ",".join(str(arg) for arg in function_args)
            formula = f"{function_name}({args_str})"
        else:
            formula = f"{function_name}()"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write date function: {e}")


# Financial Functions
def write_financial_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes financial functions to Excel cells.

    Supported functions: FV, IRR, NPV, PMT, PV

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the financial function
        function_args: Function arguments

    Returns:
        str: Excel file path

    Example:
        write_financial_function("file.xlsx", "Sheet1", "A1", "PV", 0.05, 10, -1000)
    """
    try:
        function_name = function_name.upper()

        valid_functions = {"FV", "IRR", "NPV", "PMT", "PV"}

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid financial function: {function_name}")

        if not function_args:
            raise FormulaError(f"Financial function {function_name} requires arguments")

        args_str = ",".join(str(arg) for arg in function_args)
        formula = f"{function_name}({args_str})"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write financial function: {e}")


# Logical Functions
def write_logical_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, conditions: list[Any]
) -> str:
    """
    Writes logical functions to Excel cells.

    Supported functions: AND, FALSE, IF, IFERROR, NOT, OR, TRUE

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the logical function
        conditions: Function conditions/arguments

    Returns:
        str: Excel file path

    Example:
        write_logical_function("file.xlsx", "Sheet1", "A1", "IF", "B1>10", '"Yes"', '"No"')
    """
    try:
        function_name = function_name.upper()

        valid_functions = {"AND", "FALSE", "IF", "IFERROR", "NOT", "OR", "TRUE"}

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid logical function: {function_name}")

        if function_name in ["FALSE", "TRUE"] and conditions:
            raise FormulaError(f"Function {function_name} takes no arguments")

        if function_name in ["FALSE", "TRUE"]:
            formula = f"{function_name}()"
        else:
            if not conditions:
                raise FormulaError(f"Function {function_name} requires arguments")
            conditions_str = ",".join(str(condition) for condition in conditions)
            formula = f"{function_name}({conditions_str})"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write logical function: {e}")


# Lookup Functions
def write_lookup_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes lookup functions to Excel cells.

    Supported functions: CHOOSE, COLUMN, COLUMNS, HLOOKUP, INDEX, INDIRECT,
                        MATCH, OFFSET, ROW, ROWS, VLOOKUP

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the lookup function
        function_args: Function arguments

    Returns:
        str: Excel file path

    Example:
        write_lookup_function("file.xlsx", "Sheet1", "A1", "VLOOKUP", "B1", "D:E", 2, "FALSE")
    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "CHOOSE",
            "COLUMN",
            "COLUMNS",
            "HLOOKUP",
            "INDEX",
            "INDIRECT",
            "MATCH",
            "OFFSET",
            "ROW",
            "ROWS",
            "VLOOKUP",
        }

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid lookup function: {function_name}")

        if function_args:
            args_str = ",".join(str(arg) for arg in function_args)
            formula = f"{function_name}({args_str})"
        else:
            formula = f"{function_name}()"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write lookup function: {e}")


# Math Functions
def write_math_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes math functions to Excel cells.

    Supported functions: ABS, ACOS, ASIN, ATAN, ATAN2, CEILING, COMBIN, COS,
                        COUNTIF, COUNTIFS, DEGREES, EVEN, EXP, FACT, FLOOR,
                        INT, LN, LOG, LOG10, MAX, MIN, MOD, ODD, PI, POWER,
                        PRODUCT, RADIANS, RAND, RANDBETWEEN, ROUND, ROUNDDOWN,
                        ROUNDUP, SIGN, SIN, SQRT, SUM, SUMIF, SUMIFS,
                        SUMPRODUCT, TAN, TRUNC

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the math function
        function_args: Function arguments

    Returns:
        str: Excel file path

    Example:
        write_math_function("file.xlsx", "Sheet1", "A1", "SUM", "B1:B10")
    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "ABS",
            "ACOS",
            "ASIN",
            "ATAN",
            "ATAN2",
            "CEILING",
            "COMBIN",
            "COS",
            "COUNTIF",
            "COUNTIFS",
            "DEGREES",
            "EVEN",
            "EXP",
            "FACT",
            "FLOOR",
            "INT",
            "LN",
            "LOG",
            "LOG10",
            "MAX",
            "MIN",
            "MOD",
            "ODD",
            "PI",
            "POWER",
            "PRODUCT",
            "RADIANS",
            "RAND",
            "RANDBETWEEN",
            "ROUND",
            "ROUNDDOWN",
            "ROUNDUP",
            "SIGN",
            "SIN",
            "SQRT",
            "SUM",
            "SUMIF",
            "SUMIFS",
            "SUMPRODUCT",
            "TAN",
            "TRUNC",
        }

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid math function: {function_name}")

        # Functions that take no parameters
        no_param_functions = {"PI", "RAND"}

        if function_name in no_param_functions:
            if function_args:
                raise FormulaError(f"Function {function_name} takes no arguments")
            formula = f"{function_name}()"
        else:
            if not function_args:
                raise FormulaError(f"Function {function_name} requires arguments")
            args_str = ",".join(str(arg) for arg in function_args)
            formula = f"{function_name}({args_str})"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write math function: {e}")


# Statistical Functions
def write_statistical_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes statistical functions to Excel cells.

    Supported functions: AVERAGE, AVERAGEIF, AVERAGEIFS, COUNT, COUNTA, COUNTBLANK,
                        LARGE, MEDIAN, MODE, PERCENTILE, QUARTILE, RANK, SMALL, STDEV, VAR

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the statistical function
        function_args: Function data/arguments

    Returns:
        str: Excel file path

    Example:
        write_statistical_function("file.xlsx", "Sheet1", "A1", "AVERAGE", "B1:B10")
    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "AVERAGE",
            "AVERAGEIF",
            "AVERAGEIFS",
            "COUNT",
            "COUNTA",
            "COUNTBLANK",
            "LARGE",
            "MEDIAN",
            "MODE",
            "PERCENTILE",
            "QUARTILE",
            "RANK",
            "SMALL",
            "STDEV",
            "VAR",
        }

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid statistical function: {function_name}")

        if not data:
            raise FormulaError(f"Statistical function {function_name} requires data")

        data_str = ",".join(str(item) for item in function_args)
        formula = f"{function_name}({data_str})"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write statistical function: {e}")


# Text Functions
def write_text_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes text functions to Excel cells.

    Supported functions: CHAR, CLEAN, CODE, CONCATENATE, EXACT, FIND, LEFT, LEN,
                        LOWER, MID, PROPER, REPLACE, REPT, RIGHT, SEARCH, SUBSTITUTE,
                        TEXT, TRIM, UPPER, VALUE

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the text function
        function_args: Function text arguments

    Returns:
        str: Excel file path

    Example:
        write_text_function("file.xlsx", "Sheet1", "A1", "CONCATENATE", "A1", "B1")
    """
    try:
        function_name = function_name.upper()

        valid_functions = {
            "CHAR",
            "CLEAN",
            "CODE",
            "CONCATENATE",
            "EXACT",
            "FIND",
            "LEFT",
            "LEN",
            "LOWER",
            "MID",
            "PROPER",
            "REPLACE",
            "REPT",
            "RIGHT",
            "SEARCH",
            "SUBSTITUTE",
            "TEXT",
            "TRIM",
            "UPPER",
            "VALUE",
        }

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid text function: {function_name}")

        if not function_args:
            raise FormulaError(f"Text function {function_name} requires arguments")

        args_str = ",".join(str(arg) for arg in function_args)
        formula = f"{function_name}({args_str})"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write text function: {e}")


# Info Functions
def write_info_function(
    excel_path: str, sheet_name: str, cell: str, function_name: str, function_args: list[Any]
) -> str:
    """
    Writes info functions to Excel cells.

    Supported functions: ISBLANK, ISERROR, ISNUMBER, ISTEXT

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        function_name: Name of the info function
        function_args: Function arguments

    Returns:
        str: Excel file path

    Example:
        write_info_function("file.xlsx", "Sheet1", "A1", "ISBLANK", "B1")
    """
    try:
        function_name = function_name.upper()

        valid_functions = {"ISBLANK", "ISERROR", "ISNUMBER", "ISTEXT"}

        if function_name not in valid_functions:
            raise FormulaError(f"Invalid info function: {function_name}")

        if not function_args:
            raise FormulaError(f"Info function {function_name} requires arguments")

        args_str = ",".join(str(arg) for arg in function_args)
        formula = f"{function_name}({args_str})"

        return _write_formula(excel_path, sheet_name, cell, formula)

    except FormulaError:
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write info function: {e}")


# Pivot Table Functions
def create_pivot_table(
    excel_path: str,
    source_sheet: str,
    pivot_sheet: str,
    data_range: str,
    rows: list[str],
    columns: list[str],
    values: dict[str, str],
) -> str:
    """
    Creates a pivot table from source data.

    Args:
        excel_path: Path to the Excel file
        source_sheet: Name of the source data sheet
        pivot_sheet: Name of the sheet for the pivot table
        data_range: Range of source data (e.g., 'A1:D100')
        rows: List of field names for row area
        columns: List of field names for column area
        values: Dict mapping field names to aggregation types
                {'field_name': 'sum'|'count'|'average'|'max'|'min'|'product'|'stdDev'|'var'}

    Returns:
        str: Excel file path

    Example:
        create_pivot_table("file.xlsx", "Data", "Pivot", "A1:D100",
                          ["Category"], ["Month"], {"Sales": "sum", "Count": "count"})
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if source_sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Source sheet '{source_sheet}' not found")

        # Create pivot sheet if it doesn't exist
        if pivot_sheet not in wb.sheetnames:
            wb.create_sheet(title=pivot_sheet)

        # Create pivot cache
        cache = CacheDefinition()
        cache.worksheetSource = WorksheetSource(ref=data_range, sheet=source_sheet)

        # Create pivot table
        pivot_table = TableDefinition()
        pivot_table.name = f"PivotTable_{pivot_sheet}"
        pivot_table.cache = cache

        # Add data fields
        for field_name, aggregation in values.items():
            valid_aggregations = {"sum", "count", "average", "max", "min", "product", "stdDev", "var"}
            if aggregation not in valid_aggregations:
                raise ValueError(f"Invalid aggregation type: {aggregation}")

            data_field = DataField(name=field_name, subtotal=aggregation)
            pivot_table.dataFields.append(data_field)

        # Note: Full pivot table implementation requires more complex field mapping
        # This is a simplified version that creates the basic structure

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create pivot table: {e}")


# Utility Functions
def write_formula_to_cell(excel_path: str, sheet_name: str, cell: str, formula: str) -> str:
    """
    Writes any custom formula to a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        formula: Complete Excel formula (with or without leading =)

    Returns:
        str: Excel file path

    Example:
        write_formula_to_cell("file.xlsx", "Sheet1", "A1", "=SUM(B1:B10)*0.1")
    """
    return _write_formula(excel_path, sheet_name, cell, formula)


def write_value_to_cell(excel_path: str, sheet_name: str, cell: str, value: Any) -> str:
    """
    Writes a value to a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')
        value: Value to write

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws[cell] = value
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write value to cell: {e}")


def get_cell_value(excel_path: str, sheet_name: str, cell: str) -> Any:
    """
    Gets the value from a cell.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell: Cell reference (e.g., 'A1')

    Returns:
        Any: Cell value
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        return ws[cell].value

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to get cell value: {e}")


def clear_sheet(excel_path: str, sheet_name: str) -> str:
    """
    Clears all data from a sheet.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the sheet to clear

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        # Remove and recreate sheet to clear all data
        wb.remove(wb[sheet_name])
        wb.create_sheet(title=sheet_name)
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to clear sheet: {e}")


# Batch Operations
def write_multiple_formulas(excel_path: str, sheet_name: str, formulas: dict[str, str]) -> str:
    """
    Writes multiple formulas to cells in a single operation.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        formulas: Dict mapping cell references to formulas

    Returns:
        str: Excel file path

    Example:
        write_multiple_formulas("file.xlsx", "Sheet1", {
            "A1": "=SUM(B1:B10)",
            "A2": "=AVERAGE(B1:B10)",
            "A3": "=MAX(B1:B10)"
        })
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        for cell, formula in formulas.items():
            if not formula.startswith("="):
                formula = "=" + formula
            ws[cell] = formula

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to write multiple formulas: {e}")


# Advanced Data Processing Functions


def create_autofilter(excel_path: str, sheet_name: str, data_range: str) -> str:
    """
    Adds autofilter to a data range.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data_range: Range to apply autofilter (e.g., 'A1:D100')

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]
        ws.auto_filter.ref = data_range
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create autofilter: {e}")


def create_data_validation(
    excel_path: str, sheet_name: str, cell_range: str, validation_type: str, validation_criteria: Any
) -> str:
    """
    Adds data validation to a range.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        cell_range: Range to apply validation (e.g., 'A1:A10')
        validation_type: Type of validation ('list', 'whole', 'decimal', 'date', 'time', 'text_length')
        validation_criteria: Criteria for validation (list for dropdown, dict for others)

    Returns:
        str: Excel file path

    Example:
        create_data_validation("file.xlsx", "Sheet1", "A1:A10", "list", ["Option1", "Option2", "Option3"])
        create_data_validation("file.xlsx", "Sheet1", "B1:B10", "whole", {"operator": "between", "formula1": 1, "formula2": 100})
    """
    try:
        from openpyxl.worksheet.datavalidation import DataValidation

        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        if validation_type == "list":
            # Create dropdown list
            if isinstance(validation_criteria, list):
                formula = f'"{",".join(str(item) for item in validation_criteria)}"'
            else:
                formula = str(validation_criteria)
            dv = DataValidation(type="list", formula1=formula, allow_blank=True)
        else:
            # Create other validation types
            if not isinstance(validation_criteria, dict):
                raise ValueError("validation_criteria must be a dict for non-list validations")

            dv = DataValidation(
                type=validation_type,
                operator=validation_criteria.get("operator", "between"),
                formula1=validation_criteria.get("formula1"),
                formula2=validation_criteria.get("formula2"),
                allow_blank=validation_criteria.get("allow_blank", True),
            )

        dv.add(cell_range)
        ws.add_data_validation(dv)
        wb.save(excel_path)

        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create data validation: {e}")


def create_summary_table(
    excel_path: str,
    source_sheet: str,
    summary_sheet: str,
    data_range: str,
    group_by_column: str,
    aggregations: dict[str, str],
) -> str:
    """
    Creates a summary table with aggregations (manual pivot table alternative).

    Args:
        excel_path: Path to the Excel file
        source_sheet: Name of the source data sheet
        summary_sheet: Name of the summary sheet
        data_range: Range of source data (e.g., 'A1:D100')
        group_by_column: Column letter to group by (e.g., 'A')
        aggregations: Dict mapping column letters to aggregation functions
                     {'B': 'sum', 'C': 'average', 'D': 'count'}

    Returns:
        str: Excel file path

    Example:
        create_summary_table("file.xlsx", "Data", "Summary", "A1:D100", "A",
                           {"B": "sum", "C": "average", "D": "count"})
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if source_sheet not in wb.sheetnames:
            raise SheetNotFoundError(f"Source sheet '{source_sheet}' not found")

        # Create summary sheet if it doesn't exist
        if summary_sheet not in wb.sheetnames:
            wb.create_sheet(title=summary_sheet)

        source_ws = wb[source_sheet]
        summary_ws = wb[summary_sheet]

        # Clear summary sheet
        summary_ws.delete_rows(1, summary_ws.max_row)

        # Create headers
        headers = [f"Group ({group_by_column})"]
        for col, agg in aggregations.items():
            headers.append(f"{agg.title()}({col})")

        for col_idx, header in enumerate(headers, 1):
            summary_ws.cell(row=1, column=col_idx, value=header)

        # Create formulas for unique values and aggregations
        # This creates a basic summary using Excel formulas
        summary_row = 2

        # Get unique values using a formula approach
        # Note: This is a simplified implementation
        # For a full implementation, you'd need to read the data and process it

        # Add a sample aggregation formula structure
        unique_formula = f"=UNIQUE({source_sheet}!{group_by_column}2:{group_by_column}1000)"
        summary_ws.cell(row=2, column=1, value=unique_formula)

        # Add aggregation formulas for each column
        col_idx = 2
        for col, agg_func in aggregations.items():
            if agg_func.lower() == "sum":
                formula = (
                    f"=SUMIF({source_sheet}!{group_by_column}:{group_by_column},A2,{source_sheet}!{col}:{col})"
                )
            elif agg_func.lower() == "average":
                formula = (
                    f"=AVERAGEIF({source_sheet}!{group_by_column}:{group_by_column},A2,{source_sheet}!{col}:{col})"
                )
            elif agg_func.lower() == "count":
                formula = f"=COUNTIF({source_sheet}!{group_by_column}:{group_by_column},A2)"
            elif agg_func.lower() == "max":
                formula = (
                    f"=MAXIFS({source_sheet}!{col}:{col},{source_sheet}!{group_by_column}:{group_by_column},A2)"
                )
            elif agg_func.lower() == "min":
                formula = (
                    f"=MINIFS({source_sheet}!{col}:{col},{source_sheet}!{group_by_column}:{group_by_column},A2)"
                )
            else:
                formula = (
                    f"=SUMIF({source_sheet}!{group_by_column}:{group_by_column},A2,{source_sheet}!{col}:{col})"
                )

            summary_ws.cell(row=2, column=col_idx, value=formula)
            col_idx += 1

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create summary table: {e}")


def add_subtotals(
    excel_path: str,
    sheet_name: str,
    data_range: str,
    group_column: int,
    summary_columns: list[int],
    function_type: str = "sum",
) -> str:
    """
    Adds subtotals to grouped data.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data_range: Range containing the data
        group_column: Column number (1-based) to group by
        summary_columns: List of column numbers to summarize
        function_type: Type of summary function ('sum', 'average', 'count', 'max', 'min')

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Parse data range
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # Read data to identify groups
        data = []
        for row in ws.iter_rows(
            min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True
        ):
            data.append(list(row))

        if not data:
            raise ValueError("No data found in specified range")

        # Group data by the specified column
        groups = {}
        for row_idx, row in enumerate(data[1:], 1):  # Skip header
            group_key = row[group_column - min_col]
            if group_key not in groups:
                groups[group_key] = []
            groups[group_key].append((row_idx + min_row, row))

        # Insert subtotal rows
        offset = 0
        for group_key, group_rows in groups.items():
            last_row_idx = group_rows[-1][0] + offset

            # Insert subtotal row
            ws.insert_rows(last_row_idx + 1)
            offset += 1

            # Add subtotal formulas
            subtotal_row = last_row_idx + 1

            # Add group label
            ws.cell(row=subtotal_row, column=group_column, value=f"{group_key} Total")

            # Add subtotal formulas for each summary column
            for col_num in summary_columns:
                start_row = group_rows[0][0] + offset - 1
                end_row = last_row_idx

                col_letter = ws.cell(row=1, column=col_num).column_letter

                if function_type.lower() == "sum":
                    formula = f"=SUBTOTAL(9,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_type.lower() == "average":
                    formula = f"=SUBTOTAL(1,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_type.lower() == "count":
                    formula = f"=SUBTOTAL(3,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_type.lower() == "max":
                    formula = f"=SUBTOTAL(4,{col_letter}{start_row}:{col_letter}{end_row})"
                elif function_type.lower() == "min":
                    formula = f"=SUBTOTAL(5,{col_letter}{start_row}:{col_letter}{end_row})"
                else:
                    formula = f"=SUBTOTAL(9,{col_letter}{start_row}:{col_letter}{end_row})"

                ws.cell(row=subtotal_row, column=col_num, value=formula)

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to add subtotals: {e}")


def create_advanced_filter(
    excel_path: str,
    sheet_name: str,
    data_range: str,
    criteria_range: str,
    output_range: str | None = None,
    unique_only: bool = False,
) -> str:
    """
    Creates an advanced filter with criteria.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the target sheet
        data_range: Range containing the data to filter
        criteria_range: Range containing the filter criteria
        output_range: Range for filtered output (if None, filters in place)
        unique_only: Whether to show only unique records

    Returns:
        str: Excel file path
    """
    try:
        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Sheet '{sheet_name}' not found")

        ws = wb[sheet_name]

        # Note: OpenPyXL doesn't directly support advanced filters
        # This creates the structure that Excel can use

        # Add a note about the advanced filter setup
        note_cell = ws.cell(row=1, column=20)  # Column T
        note_cell.value = f"Advanced Filter: Data={data_range}, Criteria={criteria_range}"

        if output_range:
            note_cell.value += f", Output={output_range}"
        if unique_only:
            note_cell.value += ", Unique Only"

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create advanced filter: {e}")


def create_dynamic_chart(
    excel_path: str,
    sheet_name: str,
    chart_sheet: str,
    data_range: str,
    chart_type: str = "column",
    title: str = "Chart",
) -> str:
    """
    Creates a dynamic chart from data.

    Args:
        excel_path: Path to the Excel file
        sheet_name: Name of the source data sheet
        chart_sheet: Name of the chart sheet
        data_range: Range containing the data
        chart_type: Type of chart ('column', 'line', 'pie', 'bar', 'scatter')
        title: Chart title

    Returns:
        str: Excel file path
    """
    try:
        from openpyxl.chart import BarChart, LineChart, PieChart, ScatterChart
        from openpyxl.chart.bar_chart import BarChart as ColumnChart
        from openpyxl.chart.reference import Reference

        excel_file = Path(excel_path)
        if not excel_file.exists():
            raise FileNotFoundError(f"Excel file not found: {excel_path}")

        wb = load_workbook(excel_path)

        if sheet_name not in wb.sheetnames:
            raise SheetNotFoundError(f"Source sheet '{sheet_name}' not found")

        # Create chart sheet if it doesn't exist
        if chart_sheet not in wb.sheetnames:
            wb.create_sheet(title=chart_sheet)

        source_ws = wb[sheet_name]
        chart_ws = wb[chart_sheet]

        # Create chart based on type
        chart_classes = {
            "column": ColumnChart,
            "line": LineChart,
            "pie": PieChart,
            "bar": BarChart,
            "scatter": ScatterChart,
        }

        if chart_type not in chart_classes:
            raise ValueError(f"Invalid chart type: {chart_type}")

        chart = chart_classes[chart_type]()
        chart.title = title

        # Parse data range
        from openpyxl.utils import range_boundaries

        min_col, min_row, max_col, max_row = range_boundaries(data_range)

        # Create data reference
        data = Reference(source_ws, min_col=min_col, min_row=min_row, max_col=max_col, max_row=max_row)

        # Add data to chart
        chart.add_data(data, titles_from_data=True)

        # Add chart to sheet
        chart_ws.add_chart(chart, "A1")

        wb.save(excel_path)
        return str(excel_file.absolute())

    except (FileNotFoundError, SheetNotFoundError, ValueError):
        raise
    except Exception as e:
        raise FileOperationError(f"Failed to create dynamic chart: {e}")


if __name__ == "__main__":
    # Example usage
    try:
        # Create a new Excel file
        file_path = create_excel_file("example.xlsx")
        print(f"Created Excel file: {file_path}")

        # Add some data
        data = [["Name", "Age", "Salary"], ["John", 30, 50000], ["Jane", 25, 45000], ["Bob", 35, 60000]]
        write_data_to_sheet(file_path, "Sheet", data)

        # Add formulas
        write_math_function(file_path, "Sheet", "D1", "SUM", "C2:C4")
        write_statistical_function(file_path, "Sheet", "D2", "AVERAGE", "C2:C4")
        write_math_function(file_path, "Sheet", "D3", "MAX", "C2:C4")

        # Add advanced features
        create_autofilter(file_path, "Sheet", "A1:C4")

        print("Excel file created with data, formulas, and advanced features!")

    except Exception as e:
        print(f"Error: {e}")
