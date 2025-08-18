#!/usr/bin/env python3
"""
Test script demonstrating the new formula error handling functionality.
"""

import os

from openpyxl import Workbook

from src.dreamai.toolsets.excel_formula_toolset import (
    write_and_evaluate_formula,
    write_formula_with_error_handling,
)


def create_test_excel():
    """Create a test Excel file with some sample data."""
    wb = Workbook()
    ws = wb.active
    ws.title = "TestSheet"

    # Add some sample data
    ws["A1"] = "Category"
    ws["B1"] = "Value"
    ws["C1"] = "Date"

    ws["A2"] = "Pro"
    ws["B2"] = 100
    ws["C2"] = "2023-06-01"

    ws["A3"] = "Basic"
    ws["B3"] = 50
    ws["C3"] = "2023-07-01"

    ws["A4"] = "Pro"
    ws["B4"] = 150
    ws["C4"] = "2023-08-01"

    # Create an empty column that will cause division by zero
    ws["D1"] = "Empty"
    # D2:D4 are left empty (will be 0 in formulas)

    test_file = "test_formulas.xlsx"
    wb.save(test_file)
    wb.close()
    return test_file


def test_basic_error_detection():
    """Test basic error detection functionality."""
    print("=" * 60)
    print("Testing Basic Error Detection")
    print("=" * 60)

    test_file = create_test_excel()

    # Test 1: Division by zero error
    print("\n1. Testing division by zero error:")
    result = write_and_evaluate_formula(
        test_file, "TestSheet", "E1", 'AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0")'
    )
    print('   Formula: AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0")')
    print(f"   Success: {result['success']}")
    print(f"   Error: {result['error']}")
    print(f"   Message: {result['error_message']}")

    # Test 2: Valid formula
    print("\n2. Testing valid formula:")
    result = write_and_evaluate_formula(test_file, "TestSheet", "E2", 'AVERAGEIFS(B:B,A:A,"Pro")')
    print('   Formula: AVERAGEIFS(B:B,A:A,"Pro")')
    print(f"   Success: {result['success']}")
    print(f"   Value: {result['value']}")

    # Test 3: Reference error
    print("\n3. Testing reference error:")
    result = write_and_evaluate_formula(
        test_file,
        "TestSheet",
        "E3",
        "SUM(Z99:Z100)",  # Non-existent range
    )
    print("   Formula: SUM(Z99:Z100)")
    print(f"   Success: {result['success']}")
    print(f"   Value: {result['value']}")  # Should be 0 for empty range

    # Test 4: Name error
    print("\n4. Testing name error:")
    result = write_and_evaluate_formula(test_file, "TestSheet", "E4", "INVALIDFUNCTION(A1:A4)")
    print("   Formula: INVALIDFUNCTION(A1:A4)")
    print(f"   Success: {result['success']}")
    print(f"   Error: {result['error']}")
    print(f"   Message: {result['error_message']}")

    # Cleanup
    os.remove(test_file)


def test_automatic_error_handling():
    """Test automatic error handling with retries."""
    print("\n" + "=" * 60)
    print("Testing Automatic Error Handling")
    print("=" * 60)

    test_file = create_test_excel()

    # Test 1: Auto-fix division by zero
    print("\n1. Testing auto-fix for division by zero:")
    result = write_formula_with_error_handling(
        test_file, "TestSheet", "F1", 'AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0")', max_retries=3
    )
    print('   Original: AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0")')
    print(f"   Success: {result['success']}")
    print(f"   Value: {result['value']}")
    print(f"   Attempts: {len(result['attempts'])}")

    for attempt in result["attempts"]:
        print(f"     Attempt {attempt['attempt']}: {attempt['formula']}")
        print(f"       Result: {attempt['result']['success']} - {attempt['result'].get('error', 'OK')}")

    # Test 2: Use fallback for unfixable error
    print("\n2. Testing fallback for unfixable error:")
    result = write_formula_with_error_handling(
        test_file, "TestSheet", "F2", "INVALIDFUNCTION(A1:A4)", max_retries=2, error_fallback="0"
    )
    print("   Original: INVALIDFUNCTION(A1:A4)")
    print(f"   Success: {result['success']}")
    print(f"   Value: {result['value']}")
    print(f"   Used fallback: {result.get('used_fallback', False)}")
    if result.get("used_fallback"):
        print(f"   Original formula: {result['original_formula']}")

    # Cleanup
    os.remove(test_file)


def demonstrate_ai_agent_workflow():
    """Demonstrate how an AI agent would use this for iterative formula building."""
    print("\n" + "=" * 60)
    print("AI Agent Workflow Demonstration")
    print("=" * 60)

    test_file = create_test_excel()

    # Simulate AI agent trying to create a complex formula
    formulas_to_try = [
        # First attempt - will cause division by zero
        'AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0")',
        # Second attempt - AI realizes it needs to handle empty denominator
        'IF(COUNTIFS(D:D,">0")=0,0,AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0"))',
        # Third attempt - AI uses a different approach
        'IFERROR(AVERAGEIFS(B:B,A:A,"Pro")/COUNTIFS(D:D,">0"),0)',
    ]

    print("\nSimulating AI agent iterating through formula attempts:")

    for i, formula in enumerate(formulas_to_try, 1):
        print(f"\n--- AI Agent Attempt {i} ---")
        print(f"Trying formula: {formula}")

        result = write_and_evaluate_formula(test_file, "TestSheet", "G1", formula)

        print(f"Success: {result['success']}")
        if result["success"]:
            print(f"Value: {result['value']}")
            print("✅ AI Agent: Formula works! Saving to final location.")
            break
        else:
            print(f"Error: {result['error']} - {result['error_message']}")
            print("🤖 AI Agent: Need to fix this formula...")

    print("\nFinal result in Excel file:")
    print(f"Cell G1 contains the working formula: {formulas_to_try[i - 1]}")

    # Show the final Excel file content
    from openpyxl import load_workbook

    wb = load_workbook(test_file, data_only=True)
    ws = wb["TestSheet"]
    print(f"Calculated value: {ws['G1'].value}")
    wb.close()

    # Cleanup
    os.remove(test_file)


if __name__ == "__main__":
    test_basic_error_detection()
    test_automatic_error_handling()
    demonstrate_ai_agent_workflow()

    print("\n" + "=" * 60)
    print("All tests completed!")
    print("=" * 60)
